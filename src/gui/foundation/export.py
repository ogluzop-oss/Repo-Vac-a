"""
Exportación ÚNICA del sistema Enterprise (Foundation). Un solo `exportar_excel(...)` reutilizable por
`EnterpriseToolbar` y por cualquier pantalla, para eliminar los exportadores duplicados. Replica el
comportamiento ya establecido (estilo de cabecera cian, autoajuste de columnas y registro en
Documentos → Exportaciones) de forma centralizada. No implementa lógica de negocio: solo materializa
datos ya calculados por los servicios.
"""

import logging
import os
from datetime import datetime

logger = logging.getLogger("gui.foundation.export")


def _a_dataframe(datos, columnas=None):
    import pandas as pd
    if hasattr(datos, "to_excel"):        # ya es un DataFrame
        return datos
    filas = list(datos or [])
    if filas and isinstance(filas[0], dict):
        cols = columnas or list(filas[0].keys())
        return pd.DataFrame(filas, columns=cols)
    return pd.DataFrame(filas, columns=columnas)


def exportar_excel(datos, nombre_base, *, columnas=None, hoja="Datos", referencia=None,
                   subcarpeta="informes", registrar=True) -> dict:
    """Exporta `datos` (lista de dicts / DataFrame) a un .xlsx con estilo Enterprise.

    Devuelve {"ok", "ruta", "nombre", "registrado", "error"}. Best-effort: nunca lanza.
    """
    try:
        import pandas as pd  # noqa: F401
        from openpyxl.styles import Alignment, Font, PatternFill

        df = _a_dataframe(datos, columnas)
        ruta_dir = os.path.join(os.getcwd(), "documentos", subcarpeta)
        os.makedirs(ruta_dir, exist_ok=True)
        nombre = f"{nombre_base}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        ruta = os.path.join(ruta_dir, nombre)

        import pandas as pd
        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=hoja[:31])
            ws = writer.sheets[hoja[:31]]
            fill = PatternFill("solid", fgColor="00FFC6")
            hf = Font(name="Segoe UI", bold=True, color="0E1117")
            for c in ws[1]:
                c.fill = fill; c.font = hf; c.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=2):
                for c in row:
                    c.font = Font(name="Segoe UI"); c.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ml = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = max(16, ml + 4)

        registrado = False
        if registrar:
            try:
                from src.db.documentos import registrar_documento
                registrado = bool(registrar_documento(
                    ruta, tipo="exportacion", nombre=nombre, referencia=(referencia or nombre_base)))
            except Exception as e:
                logger.debug("registrar_documento: %s", e)
        return {"ok": True, "ruta": ruta, "nombre": nombre, "registrado": registrado, "error": None}
    except Exception as e:
        logger.error("exportar_excel: %s", e)
        return {"ok": False, "ruta": None, "nombre": None, "registrado": False, "error": str(e)}
