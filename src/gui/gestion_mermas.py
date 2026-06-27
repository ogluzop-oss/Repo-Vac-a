import os
from datetime import datetime

import pandas as pd
from PyQt6.QtCore import QStringListModel, Qt
from PyQt6.QtWidgets import (
    QComboBox,
    QCompleter,
    QDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QPushButton,
    QSizePolicy,
    QSpinBox,
    QStackedWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from assets.estilo_global import (
    construir_tabla_estilizada,
    mostrar_mensaje,
)
from src.db.conexion import _ajustar_stock_articulo_por_tipo, obtener_articulo
from src.db.mermas import (
    eliminar_merma,
    modificar_merma,
    obtener_mermas,
    registrar_merma,
)
from src.utils import i18n
from src.utils.i18n import tr

# Valores lógicos del origen de stock (NO se traducen; son claves internas).
_TIPO_LINEAL = "STOCK LINEAL"
_TIPO_ALMACEN = "STOCK ALMACÉN"


def _tipo_stock_label(tipo):
    """Etiqueta visible (traducida) para un valor lógico de origen de stock."""
    if tipo == _TIPO_LINEAL:
        return tr("merma.stock_shelf", default="STOCK LINEAL")
    return tr("merma.stock_warehouse", default="STOCK ALMACÉN")

# ---------------------------------------------------------------------------
# CONSTANTES Y ESTILOS NEÓN
# ---------------------------------------------------------------------------
_CIAN = "#00FFC6"
_FONDO = "#0E1117"
_PANEL_BG = "#161B22"
_BORDE = "#30363D"

_NEON_BUTTON_INPUT_SS = f"""
QPushButton {{
    background-color: #161B22;
    color: #FFFFFF;
    border: 2px solid {_CIAN};
    border-radius: 12px;
    padding: 12px 20px;
    font-size: 16px;
    font-family: 'Segoe UI';
    font-weight: bold;
}}
QPushButton:hover {{
    border: 2px solid #FFFFFF;
    background-color: #1A2230;
}}
QPushButton:pressed {{
    background-color: {_CIAN};
    color: #0E1117;
    border: 2px solid {_CIAN};
}}
"""
_NEON_INPUT_SS = f"""
QLineEdit, QSpinBox, QComboBox {{
    background-color: #161B22;
    color: #FFFFFF;
    border: 2px solid {_CIAN};
    border-radius: 12px;
    padding: 12px 20px;
    font-size: 16px;
    font-family: 'Segoe UI';
    font-weight: bold;
}}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus {{
    border: 2px solid {_CIAN};
    background-color: #1A2230;
}}
"""

_BTN_CIAN_SS = f"""
QPushButton {{
    background-color: #0E1117;
    color: {_CIAN};
    font-weight: bold;
    border-radius: 14px;
    padding: 15px 30px;
    font-size: 14px;
    font-family: 'Segoe UI';
    border: 2px solid {_CIAN};
}}
QPushButton:hover {{
    background-color: {_CIAN};
    color: #0E1117;
    border: 2px solid {_CIAN};
}}
"""

# ---------------------------------------------------------------------------
# COMPONENTES AUXILIARES
# ---------------------------------------------------------------------------


class _SidebarBtn(QPushButton):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setObjectName("btn_sidebar")
        self.setCheckable(True)
        self.setAutoExclusive(True)
        self.setFixedHeight(55)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: #FFFFFF;
                border: none;
                border-left: 4px solid transparent;
                border-radius: 0px; /* Esquinas puntiagudas */
                font-size: 12px;
                font-family: 'Segoe UI';
                font-weight: 900;
                text-align: left;
                padding-left: 28px;
            }}
            QPushButton:hover {{ background-color: #FFFFFF; color: #0E1117; border-radius: 0px; }} /* Hover para el botón de salir */
            QPushButton:checked {{
                background-color: #1A2230;
                border-left: 4px solid {_CIAN};
                color: {_CIAN};
                border-radius: 0px;

            }}
        """)


class _StockSourceDialog(QDialog):
    """Diálogo neón para elegir entre STOCK LINEAL o ALMACÉN."""

    def __init__(self, parent, titulo, mensaje, color="#00FFC6"):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.resultado = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        container = QFrame()
        container.setStyleSheet(
            f"QFrame {{ background-color: #0D1117; border: 2px solid {color}; border-radius: 20px; }}"
        )
        layout.addWidget(container)

        inner = QVBoxLayout(container)
        inner.setContentsMargins(35, 30, 35, 30)
        inner.setSpacing(20)

        lbl_tit = QLabel(titulo.upper())
        lbl_tit.setStyleSheet(
            f"color: {color}; font-family: 'Segoe UI'; font-weight: 900; font-size: 16px; letter-spacing: 1px;"
        )
        lbl_tit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        inner.addWidget(lbl_tit)

        lbl_msg = QLabel(mensaje)
        lbl_msg.setStyleSheet(
            "color: white; font-family: 'Segoe UI'; font-weight: 900; font-size: 14px;"
        )
        lbl_msg.setWordWrap(True)
        lbl_msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        inner.addWidget(lbl_msg)

        self.btn_lineal = QPushButton(tr("merma.stock_shelf", default="STOCK LINEAL"))
        self.btn_almacen = QPushButton(tr("merma.stock_warehouse", default="STOCK ALMACÉN"))
        self.btn_cancelar = QPushButton(tr("merma.cancel_op", default="CANCELAR OPERACIÓN"))

        for btn in [self.btn_lineal, self.btn_almacen]:
            btn.setStyleSheet(_NEON_BUTTON_INPUT_SS)
            btn.setMinimumWidth(200); btn.setMaximumWidth(340)  # responsive (P2)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            inner.addWidget(btn)

        self.btn_cancelar.setStyleSheet(
            "QPushButton { background-color: #8B949E; color: #0E1117; border-radius: 12px; padding: 12px; font-weight: 900; font-size: 13px; }"
            "QPushButton:hover { background-color: #A0A8B0; }"
        )
        self.btn_cancelar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cancelar.clicked.connect(self.reject)
        inner.addWidget(self.btn_cancelar)

        self.btn_lineal.clicked.connect(lambda: self._finalizar(_TIPO_LINEAL))
        self.btn_almacen.clicked.connect(lambda: self._finalizar(_TIPO_ALMACEN))

    def _finalizar(self, seleccion):
        self.resultado = seleccion
        self.accept()


# ---------------------------------------------------------------------------
# PÁGINAS
# ---------------------------------------------------------------------------


class _RegistrarMermaPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent  # ventana contenedora (para refrescar otras vistas)
        self.articulo_actual = None
        self.tipo_stock_seleccionado = None
        self.cantidad_a_mermar = 0

        # Permitir que el widget capture eventos de ratón para quitar el foco
        self.setFocusPolicy(Qt.FocusPolicy.ClickFocus)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(50, 50, 50, 50)
        layout.addStretch(1)

        # Fase 1: Buscador (Centrado)
        self.search_container = QFrame()
        search_ly = QVBoxLayout(self.search_container)

        # Icono de papelera centrado
        self.lbl_papelera = QLabel("🗑️")
        self.lbl_papelera.setStyleSheet(
            "font-size: 140px; margin-bottom: 20px; background: transparent; border: none;"
        )
        self.lbl_papelera.setAlignment(Qt.AlignmentFlag.AlignCenter)
        search_ly.addWidget(self.lbl_papelera)

        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText(
            tr("merma.search_ph", default="INTRODUCE NOMBRE O CÓDIGO DEL ARTÍCULO...")
        )
        self.search_bar.setStyleSheet(_NEON_INPUT_SS)
        # Reducción de tamaño horizontal de la barra de búsqueda
        self.search_bar.setMinimumWidth(280); self.search_bar.setMaximumWidth(560)  # responsive (P2)
        self.search_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.search_bar.returnPressed.connect(self._buscar)

        self._completer_model = QStringListModel()
        completer = QCompleter(self._completer_model, self.search_bar)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.search_bar.setCompleter(completer)
        self.search_bar.textChanged.connect(self._on_search_text_changed)

        search_ly.addWidget(self.search_bar, alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.search_container)

        # --- FASE 2: FLUJO DE REGISTRO (QStackedWidget para fases) ---
        self.flow_stacked_widget = QStackedWidget()
        self.flow_stacked_widget.setMinimumWidth(320); self.flow_stacked_widget.setMaximumWidth(600)  # responsive (P2)
        self.flow_stacked_widget.setStyleSheet("background: transparent; border: none;")
        self.flow_stacked_widget.setVisible(False)

        # --- Phase 1: Quantity and Stock Source ---
        self.phase1_widget = QFrame()
        self.phase1_widget.setStyleSheet("background: transparent; border: none;")
        phase1_ly = QVBoxLayout(self.phase1_widget)
        phase1_ly.setSpacing(20)
        phase1_ly.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.lbl_art_info = QLabel("-")
        self.lbl_art_info.setStyleSheet(
            f"color: {_CIAN}; font-size: 22px; font-weight: 900;"
        )
        phase1_ly.addWidget(self.lbl_art_info)

        self._lbl_qty = QLabel(
            tr("merma.qty_label", default="CANTIDAD A MERMAR:"),
            styleSheet="color: white; font-weight: 900;",
        )
        phase1_ly.addWidget(self._lbl_qty)
        self.input_cantidad = QSpinBox()
        self.input_cantidad.setRange(1, 9999)
        self.input_cantidad.setFixedWidth(200)
        self.input_cantidad.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.input_cantidad.setStyleSheet(_NEON_INPUT_SS)
        phase1_ly.addWidget(self.input_cantidad)

        self._lbl_source = QLabel(
            tr("merma.source_q", default="¿DE DÓNDE DESCONTAR EL STOCK?"),
            styleSheet="color: white; font-weight: bold; font-size: 14px; margin-top: 15px;",
        )
        phase1_ly.addWidget(self._lbl_source)

        self.btn_lineal = QPushButton(tr("merma.stock_shelf", default="STOCK LINEAL"))
        self.btn_almacen = QPushButton(tr("merma.stock_warehouse", default="STOCK ALMACÉN"))
        for btn in (self.btn_lineal, self.btn_almacen):
            btn.setMinimumWidth(240); btn.setMaximumWidth(440)  # responsive (P2)
            btn.setStyleSheet(_NEON_BUTTON_INPUT_SS)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
        phase1_ly.addWidget(self.btn_lineal)
        phase1_ly.addWidget(self.btn_almacen)

        self.btn_cancel_phase1 = QPushButton(tr("merma.cancel", default="CANCELAR"))
        self.btn_cancel_phase1.setStyleSheet(
            "QPushButton { background-color: #8B949E; color: #0E1117; border-radius: 12px; padding: 12px; font-weight: 900; font-size: 13px; }"
            "QPushButton:hover { background-color: #A0A8B0; color: #0E1117; }"
        )
        self.btn_cancel_phase1.setFixedWidth(200)
        self.btn_cancel_phase1.clicked.connect(self._reset_flow)
        phase1_ly.addWidget(self.btn_cancel_phase1)

        self.flow_stacked_widget.addWidget(self.phase1_widget)

        # --- Phase 2: Motivo and Register ---
        self.phase2_widget = QFrame()
        self.phase2_widget.setStyleSheet("background: transparent; border: none;")
        phase2_ly = QVBoxLayout(self.phase2_widget)
        phase2_ly.setSpacing(20)
        phase2_ly.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self._lbl_reason = QLabel(
            tr("merma.reason_label", default="MOTIVO DE LA MERMA:"),
            styleSheet="color: white; font-weight: bold; font-size: 14px;",
        )
        phase2_ly.addWidget(self._lbl_reason)
        self.combo_motivo = QComboBox()
        self.combo_motivo.setEditable(True)
        self.combo_motivo.addItems(self._motivos_traducidos())
        self.combo_motivo.setMinimumWidth(240); self.combo_motivo.setMaximumWidth(440)  # responsive (P2)
        self.combo_motivo.setStyleSheet(_NEON_INPUT_SS)
        phase2_ly.addWidget(self.combo_motivo)

        self.btn_save = QPushButton(tr("merma.save", default="REGISTRAR MERMA"))
        self.btn_save.setStyleSheet(_BTN_CIAN_SS)
        self.btn_save.setMinimumWidth(200); self.btn_save.setMaximumWidth(340)  # responsive (P2)
        self.btn_save.clicked.connect(self._registrar_final)
        phase2_ly.addWidget(self.btn_save)

        self.btn_back_phase2 = QPushButton(tr("merma.back", default="VOLVER"))
        self.btn_back_phase2.setStyleSheet(
            "QPushButton { background-color: #8B949E; color: #0E1117; border-radius: 12px; padding: 12px; font-weight: 900; font-size: 13px; }"
            "QPushButton:hover { background-color: #A0A8B0; color: #0E1117; }"
        )
        self.btn_back_phase2.setFixedWidth(200)
        self.btn_back_phase2.clicked.connect(
            lambda: self.flow_stacked_widget.setCurrentIndex(0)
        )
        phase2_ly.addWidget(self.btn_back_phase2)

        self.flow_stacked_widget.addWidget(self.phase2_widget)

        layout.addWidget(
            self.flow_stacked_widget, alignment=Qt.AlignmentFlag.AlignCenter
        )
        layout.addStretch(1)

        # Connections
        self.btn_lineal.clicked.connect(
            lambda: self._seleccionar_origen(_TIPO_LINEAL)
        )
        self.btn_almacen.clicked.connect(
            lambda: self._seleccionar_origen(_TIPO_ALMACEN)
        )
        self._cargar_completer()

    @staticmethod
    def _motivos_traducidos():
        return [
            tr("merma.reason_expiry", default="Caducidad"),
            tr("merma.reason_breakage", default="Rotura"),
            tr("merma.reason_theft", default="Robo"),
            tr("merma.reason_sample", default="Muestra"),
            tr("merma.reason_inventory_error", default="Error de Inventario"),
            tr("merma.reason_other", default="Otro"),
        ]

    def _retraducir(self):
        self.search_bar.setPlaceholderText(
            tr("merma.search_ph", default="INTRODUCE NOMBRE O CÓDIGO DEL ARTÍCULO...")
        )
        self._lbl_qty.setText(tr("merma.qty_label", default="CANTIDAD A MERMAR:"))
        self._lbl_source.setText(tr("merma.source_q", default="¿DE DÓNDE DESCONTAR EL STOCK?"))
        self.btn_lineal.setText(tr("merma.stock_shelf", default="STOCK LINEAL"))
        self.btn_almacen.setText(tr("merma.stock_warehouse", default="STOCK ALMACÉN"))
        self.btn_cancel_phase1.setText(tr("merma.cancel", default="CANCELAR"))
        self._lbl_reason.setText(tr("merma.reason_label", default="MOTIVO DE LA MERMA:"))
        actual = self.combo_motivo.currentText()
        self.combo_motivo.clear()
        self.combo_motivo.addItems(self._motivos_traducidos())
        self.combo_motivo.setEditText(actual)
        self.btn_save.setText(tr("merma.save", default="REGISTRAR MERMA"))
        self.btn_back_phase2.setText(tr("merma.back", default="VOLVER"))

    def _cargar_completer(self):
        # Assuming _get_todos_articulos_para_completer is available in src.db.conexion
        from src.db.conexion import _get_todos_articulos_para_completer

        articulos = _get_todos_articulos_para_completer()
        etiquetas = [f"{a[0]} – {a[1]}" for a in articulos]
        self._completer_model.setStringList(etiquetas)

    def _on_search_text_changed(self, text):
        if not text.strip():
            self._reset_flow()

    def _buscar(self):  # Renamed from _buscar_articulo for clarity
        q = self.search_bar.text().strip()
        if not q:
            return

        # Extract code from "CODE – NAME" format if completer is used
        codigo = q.split("–")[0].strip() if "–" in q else q
        art = obtener_articulo(codigo)

        if art:
            self.articulo_actual = art
            self.lbl_art_info.setText(f"{art['nombre'].upper()} ({art['codigo']})")
            self.search_container.setVisible(False)
            self.flow_stacked_widget.setCurrentIndex(0)  # Show Phase 1
            self.flow_stacked_widget.setVisible(True)
            self.input_cantidad.setValue(1)
            # Set max value for spinbox based on available stock
            max_stock = max(
                int(art.get("Stock_tienda", 0)), int(art.get("Stock_total", 0))
            )
            self.input_cantidad.setMaximum(max(1, max_stock))  # Ensure min 1
        else:
            mostrar_mensaje(
                self,
                tr("merma.not_found_title", default="No Encontrado"),
                tr("merma.not_found_msg", default="No se encontró el artículo: {q}", q=q),
                nivel="warning",
            )

    def _seleccionar_origen(self, tipo):
        self.tipo_stock_seleccionado = tipo
        cant = self.input_cantidad.value()
        codigo = self.articulo_actual["codigo"]

        col = "Stock_tienda" if tipo == _TIPO_LINEAL else "Stock_total"
        disponible = int(self.articulo_actual.get(col, 0))

        if cant > disponible:
            mostrar_mensaje(
                self,
                tr("merma.blocked_title", default="Operación Bloqueada"),
                tr("merma.insufficient_msg",
                   default="STOCK INSUFICIENTE EN {tipo}.\nDisponible: {disp}\nSolicitado: {sol}",
                   tipo=_tipo_stock_label(tipo), disp=disponible, sol=cant),
                nivel="error",
            )
            return

        self.cantidad_a_mermar = cant
        self.flow_stacked_widget.setCurrentIndex(1)  # Show Phase 2 (Motivo)

    def _registrar_final(self):
        cant = self.cantidad_a_mermar
        motivo = self.combo_motivo.currentText().strip()
        codigo = self.articulo_actual["codigo"]
        tipo = self.tipo_stock_seleccionado

        if not motivo:
            mostrar_mensaje(
                self,
                tr("merma.error_title", default="Error"),
                tr("merma.empty_reason", default="El motivo de la merma no puede estar vacío."),
                nivel="warning",
            )
            return

        # A2.3: registro de merma + descuento de stock ATÓMICOS (una transacción).
        _col_merma = "Stock_tienda" if tipo == _TIPO_LINEAL else "Stock_total"
        if registrar_merma(codigo, cant, self.combo_motivo.currentText(), columna_stock=_col_merma):
            mostrar_mensaje(
                self,
                tr("merma.success_title", default="Éxito"),
                tr("merma.registered_msg",
                   default="Merma registrada y stock actualizado correctamente."),
                nivel="success",
            )
            self._reset_flow()
            # Optionally refresh other pages if they exist and are visible
            if (
                hasattr(self.main_window, "_page_modificar_eliminar")
                and self.main_window._page_modificar_eliminar.isVisible()
            ):
                self.main_window._page_modificar_eliminar.cargar_datos()

    def _reset_flow(self):
        self.articulo_actual = None
        self.tipo_stock_seleccionado = None
        self.cantidad_a_mermar = 0
        self.flow_stacked_widget.setVisible(False)
        self.search_container.setVisible(True)
        self.search_bar.clear()
        self.search_bar.setFocus()
        self.flow_stacked_widget.setCurrentIndex(0)  # Reset to first phase

    def mousePressEvent(self, event):
        """Quita la selección de la barra de búsqueda al hacer clic fuera."""
        self.search_bar.clearFocus()
        super().mousePressEvent(event)

    def cargar_datos(self):
        """Refresca el autocompletado al entrar en la pestaña."""
        self._cargar_completer()


class _ModificarEliminarMermaPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 30)
        layout.setSpacing(20)

        self._title = title = QLabel(tr("merma.list_title", default="REGISTRO DE MERMAS DEL MES"))
        title.setStyleSheet(
            f"color: {_CIAN}; font-size: 20px; font-weight: 900; letter-spacing: 1px;"
        )
        layout.addWidget(title)

        self.container_tabla, self.tabla = construir_tabla_estilizada(self)
        if self.container_tabla.layout():
            self.container_tabla.layout().setContentsMargins(2, 2, 2, 2)

        # Estética de redondeo superior para no cortar el neón
        self.tabla.horizontalHeader().setStyleSheet(f"""
            QHeaderView {{
                background-color: transparent;
                border: none;
            }}
            QHeaderView::section {{
                background-color: #1A1D23;
                color: {_CIAN};
                border: none;
                padding: 10px;
                font-weight: bold;
            }}
            QHeaderView::section:hover {{
                background-color: {_CIAN};
                color: #0E1117;
            }}
            QHeaderView::section:first {{ border-top-left-radius: 18px; }}
            QHeaderView::section:last {{ border-top-right-radius: 18px; }}
        """)

        self.tabla.setColumnCount(5)
        self.tabla.setHorizontalHeaderLabels(self._headers())
        self.tabla.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.tabla.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Fixed
        )
        self.tabla.setColumnWidth(4, 260)
        layout.addWidget(self.container_tabla)

    @staticmethod
    def _headers():
        return [
            tr("merma.col_code", default="CÓDIGO"),
            tr("merma.col_qty", default="CANTIDAD"),
            tr("merma.col_reason", default="MOTIVO"),
            tr("merma.col_date", default="FECHA"),
            tr("merma.col_actions", default="ACCIONES"),
        ]

    def _retraducir(self):
        self._title.setText(tr("merma.list_title", default="REGISTRO DE MERMAS DEL MES"))
        self.tabla.setHorizontalHeaderLabels(self._headers())
        self.cargar_datos()

    def cargar_datos(self):
        self.tabla.setRowCount(0)
        mes_actual = datetime.now().strftime("%Y-%m")
        mermas = obtener_mermas(mes=mes_actual)

        for i, m in enumerate(mermas):
            self.tabla.insertRow(i)
            id_m, cod, cant, mot, fecha = m

            for j, val in enumerate([cod, cant, mot, str(fecha)[:16]]):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tabla.setItem(i, j, item)

            btn_box = QWidget()
            btn_lyt = QHBoxLayout(btn_box)
            btn_lyt.setContentsMargins(5, 2, 5, 2)

            btn_edit = QPushButton(tr("merma.edit_btn", default="MODIFICAR"))
            btn_edit.setStyleSheet(
                _BTN_CIAN_SS
                + "QPushButton { padding: 5px 10px; font-size: 10px; height: 30px; }"
            )
            btn_edit.clicked.connect(lambda _, mid=id_m, r=i: self._modificar(mid, r))

            btn_del = QPushButton(tr("merma.del_btn", default="ELIMINAR"))
            btn_del.setStyleSheet(
                "QPushButton { background-color: #0D1117; color: #F85149; font-weight: 900; border-radius: 8px; padding: 5px 10px; font-size: 10px; height: 30px; border: 2px solid #F85149; }"
                "QPushButton:hover { background-color: #F85149; color: #0E1117; border: 2px solid #F85149; }"
            )
            btn_del.clicked.connect(lambda _, mid=id_m, r=i: self._eliminar(mid, r))

            btn_lyt.addWidget(btn_edit)
            btn_lyt.addWidget(btn_del)
            self.tabla.setCellWidget(i, 4, btn_box)

    def _modificar(self, id_merma, row):
        codigo = self.tabla.item(row, 0).text()
        cant_antigua = int(self.tabla.item(row, 1).text())
        nueva_cant, ok = QInputDialog.getInt(
            self,
            tr("merma.edit_title", default="Modificar Merma"),
            tr("merma.new_qty_prompt", default="Nueva cantidad para {codigo}:", codigo=codigo),
            cant_antigua,
            1,
            9999,
        )
        if not ok or nueva_cant == cant_antigua:
            return

        art = obtener_articulo(codigo)
        diff = nueva_cant - cant_antigua

        if diff > 0:
            dlg = _StockSourceDialog(
                self,
                tr("merma.increase_title", default="Aumentar Merma"),
                tr("merma.increase_msg",
                   default="Va a añadir {diff} uds. ¿De dónde descontar el stock?", diff=diff),
                _CIAN,
            )
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            tipo = dlg.resultado
            if diff > int(
                art.get("Stock_tienda" if tipo == _TIPO_LINEAL else "Stock_total", 0)
            ):
                mostrar_mensaje(
                    self,
                    tr("merma.blocked2_title", default="Bloqueado"),
                    tr("merma.db_insufficient", default="Stock insuficiente en la base de datos."),
                    nivel="error",
                )
                return
            adj = -diff
        else:
            dlg = _StockSourceDialog(
                self,
                tr("merma.reduce_title", default="Reducir Merma"),
                tr("merma.reduce_msg",
                   default="Va a reducir la merma. ¿A dónde sumar el stock devuelto?"),
                _CIAN,
            )
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            tipo, adj = dlg.resultado, abs(diff)

        if modificar_merma(id_merma, nueva_cant):
            _ajustar_stock_articulo_por_tipo(codigo, adj, tipo)
            mostrar_mensaje(
                self,
                tr("merma.success_title", default="Éxito"),
                tr("merma.updated_msg", default="Merma actualizada correctamente."),
                nivel="success",
            )
            self.cargar_datos()

    def _eliminar(self, id_merma, row):
        codigo = self.tabla.item(row, 0).text()
        cantidad = int(self.tabla.item(row, 1).text())
        dlg = _StockSourceDialog(
            self,
            tr("merma.delete_title", default="Eliminar Registro"),
            tr("merma.delete_q", default="¿A dónde desea devolver el stock total?"),
            "#F85149",
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        if eliminar_merma(id_merma):
            _ajustar_stock_articulo_por_tipo(codigo, cantidad, dlg.resultado)
            mostrar_mensaje(
                self,
                tr("merma.success_title", default="Éxito"),
                tr("merma.deleted_msg", default="Registro eliminado y stock repuesto."),
                nivel="success",
            )
            self.cargar_datos()


class _ExportarExcelPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 30)
        layout.setSpacing(20)

        self._title = title = QLabel(tr("merma.reports_title", default="REPORTES MENSUALES"))
        title.setStyleSheet(
            f"color: {_CIAN}; font-size: 20px; font-weight: 900; letter-spacing: 1px;"
        )
        layout.addWidget(title)

        self.container_tabla, self.tabla = construir_tabla_estilizada(self)
        if self.container_tabla.layout():
            self.container_tabla.layout().setContentsMargins(2, 2, 2, 2)

        # Estética de redondeo superior para no cortar el neón
        self.tabla.horizontalHeader().setStyleSheet(f"""
            QHeaderView {{
                background-color: transparent;
                border: none;
            }}
            QHeaderView::section {{
                background-color: #1A1D23;
                color: {_CIAN};
                border: none;
                padding: 10px;
                font-weight: bold;
            }}
            QHeaderView::section:hover {{
                background-color: {_CIAN};
                color: #0E1117;
            }}
            QHeaderView::section:first {{ border-top-left-radius: 18px; }}
            QHeaderView::section:last {{ border-top-right-radius: 18px; }}
        """)

        self.tabla.setColumnCount(3)
        self.tabla.setHorizontalHeaderLabels(self._headers())
        self.tabla.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.container_tabla)

    @staticmethod
    def _headers():
        return [
            tr("merma.col_month", default="MES / AÑO"),
            tr("merma.col_file", default="ARCHIVO"),
            tr("merma.col_action", default="ACCIÓN"),
        ]

    def _retraducir(self):
        self._title.setText(tr("merma.reports_title", default="REPORTES MENSUALES"))
        self.tabla.setHorizontalHeaderLabels(self._headers())
        self.cargar_datos()

    def cargar_datos(self):
        self.tabla.setRowCount(0)
        mermas = obtener_mermas()
        meses = sorted(list(set(str(m[4])[:7] for m in mermas)), reverse=True)

        for i, mes_str in enumerate(meses):
            self.tabla.insertRow(i)
            dt = datetime.strptime(mes_str, "%Y-%m")
            fecha_disp = dt.strftime("%m/%Y")
            file_name = f"Mermas_{dt.strftime('%m_%Y')}.xlsx"

            for j, val in enumerate([fecha_disp, file_name]):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tabla.setItem(i, j, item)

            btn = QPushButton(tr("merma.export_btn", default="EXPORTAR"))
            btn.setStyleSheet(
                _BTN_CIAN_SS
                + "QPushButton { padding: 5px; font-size: 11px; height: 32px; }"
            )
            btn.clicked.connect(lambda _, m=mes_str, f=file_name: self._exportar(m, f))
            self.tabla.setCellWidget(i, 2, btn)

    def _exportar(self, mes_f, name):
        try:
            mermas = obtener_mermas(mes=mes_f)
            _h = (
                tr("merma.col_code", default="CÓDIGO"),
                tr("merma.col_qty", default="CANTIDAD"),
                tr("merma.col_reason", default="MOTIVO"),
                tr("merma.col_date", default="FECHA"),
            )
            df = pd.DataFrame(
                [
                    {
                        _h[0]: m[1],
                        _h[1]: m[2],
                        _h[2]: m[3],
                        _h[3]: str(m[4])[:16],
                    }
                    for m in mermas
                ]
            )
            ruta_dir = os.path.join(os.getcwd(), "documentos", "mermas")
            os.makedirs(ruta_dir, exist_ok=True)
            path = os.path.join(ruta_dir, name)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Mermas")
                ws = writer.sheets["Mermas"]
                from openpyxl.styles import Alignment, Font, PatternFill

                fill = PatternFill(
                    start_color="00FFC6", end_color="00FFC6", fill_type="solid"
                )
                hdr_font = Font(name="Segoe UI", bold=True)
                for cell in ws[1]:
                    cell.fill, cell.font, cell.alignment = (
                        fill,
                        hdr_font,
                        Alignment(horizontal="center"),
                    )
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font, cell.alignment = Font(name="Segoe UI"), Alignment(
                            horizontal="center"
                        )

            mostrar_mensaje(
                self,
                tr("merma.export_title", default="Exportación"),
                tr("merma.export_ok", default="Archivo generado:\n{path}", path=path),
                nivel="success",
            )
            from src.utils import plataforma
            plataforma.abrir_carpeta(ruta_dir)
        except Exception as e:
            mostrar_mensaje(
                self,
                tr("merma.error_title", default="Error"),
                tr("merma.export_err", default="Fallo al exportar Excel: {e}", e=e),
                nivel="error",
            )


class GestionMermasWindow(QWidget):
    def __init__(self, callback_vuelta=None, usuario=None, **kwargs):
        super().__init__()

        self.callback_vuelta = callback_vuelta
        self.usuario_actual = usuario
        self.setWindowTitle(tr("merma.window_title", default="SMART MANAGER - GESTIÓN DE MERMAS"))
        self.resize(1100, 750)
        self.setStyleSheet(f"background-color: {_FONDO};")

        self.setup_ui()
        i18n.conectar_retraduccion(self, self._retraducir)

        # P3 (UX-TPV-01): sidebar colapsable con persistencia por usuario.
        try:
            from src.gui.sidebar_colapsable import instalar_sidebar_colapsable
            if getattr(self, "sidebar", None) is not None:
                instalar_sidebar_colapsable(self, self.sidebar, usuario=self.usuario_actual, clave="mermas")
        except Exception:
            pass

    def setup_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # --- SIDEBAR ---
        sidebar = QFrame()
        self.sidebar = sidebar  # P3: referencia para el toggle colapsable
        sidebar.setFixedWidth(280)
        sidebar.setStyleSheet(
            f"background-color: {_PANEL_BG}; border-right: 1px solid {_BORDE};"
        )
        side_ly = QVBoxLayout(sidebar)
        side_ly.setContentsMargins(0, 40, 0, 20)

        lbl_m = QLabel(tr("merma.smart_losses", default="SMART LOSSES"))
        lbl_m.setStyleSheet(  # Sidebar title
            "color: white; font-size: 16px; font-weight: 900; margin-left: 30px; margin-bottom: 35px; letter-spacing: 2px;"
        )
        side_ly.addWidget(lbl_m)

        self._tab_keys = ["merma.tab_register", "merma.tab_edit", "merma.tab_export"]
        _tab_def = ["REGISTRAR MERMA", "MODIFICAR / ELIMINAR", "EXPORTAR EXCEL"]
        self.nav_btns = []
        for idx, key in enumerate(self._tab_keys):
            btn = _SidebarBtn(tr(key, default=_tab_def[idx]))
            btn.clicked.connect(lambda _, i=idx: self._ir_a(i))
            side_ly.addWidget(btn)
            self.nav_btns.append(btn)

        side_ly.addStretch()
        self._btn_exit = btn_exit = _SidebarBtn(tr("merma.exit", default="SALIR AL MENÚ"))
        btn_exit.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #F85149;
                border: none;
                border-left: 4px solid transparent;
                font-size: 12px;
                font-family: 'Segoe UI';
                font-weight: 900;
                text-align: left;
                padding-left: 28px;
            }
            QPushButton:hover { /* Hover para el botón de salir */
                background-color: #F85149;
                color: #0E1117;
                border-radius: 0px;
            }
            """)
        btn_exit.clicked.connect(self.volver_menu_principal)
        side_ly.addWidget(btn_exit)
        root.addWidget(sidebar)

        # --- CONTENIDO ---
        self.vistas = QStackedWidget()
        self._page_registrar = _RegistrarMermaPage(self)
        self._page_modificar_eliminar = _ModificarEliminarMermaPage(self)
        self._page_exportar = _ExportarExcelPage(self)
        for p in (self._page_registrar, self._page_modificar_eliminar, self._page_exportar):
            self.vistas.addWidget(p)
        root.addWidget(self.vistas)

        self._ir_a(0)

    def _retraducir(self):
        self.setWindowTitle(tr("merma.window_title", default="SMART MANAGER - GESTIÓN DE MERMAS"))
        _tab_def = ["REGISTRAR MERMA", "MODIFICAR / ELIMINAR", "EXPORTAR EXCEL"]
        for i, btn in enumerate(self.nav_btns):
            btn.setText(tr(self._tab_keys[i], default=_tab_def[i]))
        self._btn_exit.setText(tr("merma.exit", default="SALIR AL MENÚ"))
        for page in (self._page_registrar, self._page_modificar_eliminar, self._page_exportar):
            if hasattr(page, "_retraducir"):
                page._retraducir()

    def _ir_a(self, index):
        self.vistas.setCurrentIndex(index)
        for i, btn in enumerate(self.nav_btns):
            btn.setChecked(i == index)
        # Refrescar datos de la vista
        widget = self.vistas.widget(index)
        if hasattr(widget, "cargar_datos"):
            widget.cargar_datos()

    def volver_menu_principal(self):
        if self.callback_vuelta:
            self.callback_vuelta()
        self.close()
