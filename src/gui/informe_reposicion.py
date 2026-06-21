# src/gui/informe_reposicion.py
import os
from datetime import datetime

import pandas as pd
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QApplication,
    QDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QStackedWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from assets.estilo_global import (
    construir_tabla_estilizada,
    repolish_widget,
)
from src.db.conexion import (
    obtener_articulo,
    obtener_conexion,
    set_stock_esperado,
)
from src.db.conexion import stock_signals as global_stock_signals
from src.utils import i18n
from src.utils.i18n import tr


class InformeReposicionWindow(QWidget):
    def __init__(
        self, callback_vuelta=None, usuario=None, stock_signals=None, **kwargs
    ):
        super().__init__()

        self.callback_vuelta = callback_vuelta
        self.usuario_actual = usuario

        if isinstance(usuario, dict):
            self.perfil = usuario.get("perfil", "OPERARIO")
        else:
            self.perfil = getattr(usuario, "perfil", "OPERARIO")

        self.signals = (
            stock_signals or global_stock_signals
        )  # Usar la instancia global si no se provee una
        self.signals.stock_actualizado.connect(self._on_stock_actualizado)

        self.setWindowTitle(tr("repo.window_title", default="Smart Manager - Informe de Reposición"))
        self.resize(950, 600)

        self.crear_campo_repuesto()
        self.setup_ui()
        i18n.conectar_retraduccion(self, self._retraducir)

    # ============================================================
    # BLOQUE INICIALIZACIÓN DE BASE DE DATOS
    # ============================================================

    def crear_campo_repuesto(self):
        """Verifica y añade la columna repuesto usando sintaxis MariaDB."""
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute("SHOW COLUMNS FROM articulos LIKE 'repuesto'")
                if not cur.fetchone():
                    cur.execute(
                        "ALTER TABLE articulos ADD COLUMN repuesto INTEGER DEFAULT 0"
                    )
                    conn.commit()
        except Exception as e:
            print(f"Error al verificar/crear columna repuesto: {e}")

    # ---------------------------------------------------------------------------
    # CONSTANTES Y ESTILOS
    # ---------------------------------------------------------------------------
    _CIAN = "#00FFC6"
    _FONDO = "#0E1117"
    _PANEL_BG = "#161B22"
    _BORDE = "#30363D"

    _NEON_INPUT_SS = f"""
    QLineEdit {{
        background-color: #161B22;
        color: #FFFFFF;
        border: 2px solid {_CIAN};
        border-radius: 12px;
        padding: 12px 20px;
        font-size: 16px;
        font-family: 'Segoe UI';
        font-weight: bold;
    }}
    QLineEdit:focus {{
        border: 2px solid {_CIAN};
        background-color: #1A2230;
    }}
    """

    _BTN_ROJO_SS = """
    QPushButton {
        background-color: #0E1117;
        color: #FF4B4B;
        font-weight: bold;
        border-radius: 14px;
        padding: 12px 24px;
        font-size: 13px;
        font-family: 'Segoe UI';
        border: 2px solid #FF4B4B;
    }
    QPushButton:hover {
        background-color: #FF4B4B;
        color: #0E1117;
        border: 2px solid #FF4B4B;
    }
    """

    _BTN_CIAN_SS = f"""
    QPushButton {{
        background-color: #0E1117;
        color: {_CIAN};
        font-weight: bold;
        border-radius: 14px;
        padding: 12px 24px;
        font-size: 13px;
        font-family: 'Segoe UI';
        border: 2px solid {_CIAN};
    }}
    QPushButton:hover {{
        background-color: {_CIAN};
        color: #0E1117;
        border: 2px solid {_CIAN};
    }}
    """

    # ---------------------------------------------------------------------------
    # COMPONENTES DE INTERFAZ
    # ---------------------------------------------------------------------------

    class _SidebarBtn(QPushButton):
        def __init__(self, text, parent=None):
            super().__init__(text, parent)
            self.setObjectName("btn_sidebar")
            self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)
            self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            self.setMouseTracking(True)
            self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            self.setFixedHeight(55)
            self.setCursor(Qt.CursorShape.PointingHandCursor)
            self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            self.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    color: #FFFFFF;
                    border: none;
                    border-left: 4px solid transparent;
                    border-radius: 0px;
                    font-size: 12px;
                    font-family: 'Segoe UI';
                    font-weight: 900;
                    text-align: left;
                    padding-left: 28px;
                }}
                QPushButton:hover {{
                    background-color: #FFFFFF;
                    color: #0E1117;
                }}
                QPushButton:checked {{
                    background-color: #1A2230;
                    border-left: 4px solid {InformeReposicionWindow._CIAN};
                    color: {InformeReposicionWindow._CIAN};
                }}
            """)

        def enterEvent(self, event):
            super().enterEvent(event)
            repolish_widget(self)

        def leaveEvent(self, event):
            super().leaveEvent(event)
            repolish_widget(self)

    def _sombra_cian(self, widget):
        fx = QGraphicsDropShadowEffect()
        fx.setBlurRadius(22)
        fx.setColor(QColor(self._CIAN))
        fx.setOffset(0)
        widget.setGraphicsEffect(fx)

    # ---------------------------------------------------------------------------
    # PÁGINAS DE CONTENIDO (Placeholders iniciales)
    # ---------------------------------------------------------------------------

    class _EstadoReposicionPage(QWidget):
        def __init__(self, parent=None):
            super().__init__(parent)
            layout = QVBoxLayout(self)
            layout.setContentsMargins(30, 30, 30, 30)
            layout.setSpacing(20)

            self._title = title = QLabel(tr("repo.status_title", default="ESTADO DE REPOSICIÓN"))
            title.setStyleSheet(
                f"color: {InformeReposicionWindow._CIAN}; font-size: 24px; font-weight: bold;"
            )
            layout.addWidget(title)

            # Placeholder para la tabla
            self.container_tabla, self.tabla = construir_tabla_estilizada(self)
            # Margen de seguridad para no tapar el contorno neón del contenedor y respetar el redondeo
            if self.container_tabla.layout():
                self.container_tabla.layout().setContentsMargins(2, 2, 2, 2)

            self.tabla.horizontalHeader().setStyleSheet(f"""
                QHeaderView {{
                    background-color: transparent;
                    border: none;
                }}
                QHeaderView::section {{
                    background-color: #1A1D23;
                    color: {InformeReposicionWindow._CIAN};
                    border: none;
                    padding: 10px;
                    font-weight: bold;
                }}
                QHeaderView::section:hover {{
                    background-color: {InformeReposicionWindow._CIAN};
                    color: #0E1117;
                }}
                QHeaderView::section:first {{ border-top-left-radius: 18px; }}
                QHeaderView::section:last {{ border-top-right-radius: 18px; }}
            """)
            self.tabla.setStyleSheet(
                "QTableWidget { background-color: transparent; border: none; }"
            )

            self.tabla.setColumnCount(5)
            self.tabla.setHorizontalHeaderLabels(self._cols())
            self.tabla.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.Stretch
            )
            layout.addWidget(self.container_tabla)

            self.btn_refresh = QPushButton("🔄 " + tr("repo.refresh", default="ACTUALIZAR TABLA"))
            self.btn_refresh.setStyleSheet(InformeReposicionWindow._BTN_CIAN_SS)
            self.btn_refresh.setFixedWidth(250)
            self.btn_refresh.clicked.connect(self._ejecutar_actualizacion)
            InformeReposicionWindow._sombra_cian(parent, self.btn_refresh)
            layout.addWidget(self.btn_refresh, alignment=Qt.AlignmentFlag.AlignRight)

            layout.addStretch(1)

        @staticmethod
        def _cols():
            return [
                tr("repo.col_code", default="CÓDIGO"),
                tr("repo.col_name", default="NOMBRE"),
                tr("repo.col_wh", default="STOCK ALMACÉN"),
                tr("repo.col_shelf", default="STOCK LINEAL"),
                tr("repo.col_expected", default="STOCK ESPERADO"),
            ]

        def _retraducir(self):
            self._title.setText(tr("repo.status_title", default="ESTADO DE REPOSICIÓN"))
            self.tabla.setHorizontalHeaderLabels(self._cols())
            self.btn_refresh.setText("🔄 " + tr("repo.refresh", default="ACTUALIZAR TABLA"))

        def _ejecutar_actualizacion(self):
            self.btn_refresh.setText("⌛ " + tr("repo.refreshing", default="ACTUALIZANDO..."))
            self.btn_refresh.setEnabled(False)
            QApplication.processEvents()
            self.cargar_datos()
            # Pequeño delay de cortesía para que se vea la animación antes de restaurar
            QTimer.singleShot(800, self._finalizar_actualizacion)

        def _finalizar_actualizacion(self):
            self.btn_refresh.setText("🔄 " + tr("repo.refresh", default="ACTUALIZAR TABLA"))
            self.btn_refresh.setEnabled(True)

        def cargar_datos(self):
            self.tabla.setRowCount(0)
            try:
                with obtener_conexion() as conn:
                    cur = conn.cursor()
                    # COALESCE para manejar posibles valores NULL en la base de datos
                    cur.execute(
                        "SELECT codigo, nombre, COALESCE(Stock_total, 0), COALESCE(Stock_tienda, 0), COALESCE(Stock_esperado, 0) FROM articulos"
                    )

                    for row in cur.fetchall():
                        codigo, nombre, s_alm, s_lin, s_esp = row

                        # Asegurarse de que s_esp sea un número para el cálculo
                        s_esp = int(s_esp) if s_esp is not None else 0
                        s_lin = int(s_lin) if s_lin is not None else 0

                        # Lógica de umbral: < 70% del stock esperado
                        umbral = s_esp * 0.7
                        if s_lin < umbral:
                            r = self.tabla.rowCount()
                            self.tabla.insertRow(r)
                            items = [
                                str(codigo),
                                str(nombre),
                                str(s_alm),
                                str(s_lin),
                                str(s_esp),
                            ]
                            for c, val in enumerate(items):
                                item = QTableWidgetItem(val)
                                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                                self.tabla.setItem(r, c, item)
                self.tabla.resizeRowsToContents()
            except Exception as e:
                print(f"Error cargando reposición: {e}")

    class _EditarStockEsperadoDialog(QDialog):
        def __init__(self, parent, codigo, nombre, stock_esperado_actual):
            super().__init__(parent)
            self.setWindowTitle(tr("repo.edit_title", default="Editar Stock Esperado"))
            self.setFixedSize(450, 300)
            self.setStyleSheet(f"""
                QDialog {{
                    background-color: {InformeReposicionWindow._PANEL_BG};
                    border: 2px solid {InformeReposicionWindow._CIAN};
                    border-radius: 15px;
                }}
                QLabel {{
                    color: #FFFFFF;
                    font-family: 'Segoe UI';
                    font-weight: bold;
                    font-size: 14px;
                    border: none;
                    background: transparent;
                }}
            """)

            layout = QVBoxLayout(self)
            layout.setContentsMargins(30, 30, 30, 30)
            layout.setSpacing(15)

            lbl_articulo = QLabel(tr("repo.item_label", default="Artículo: {nombre} ({codigo})", nombre=nombre, codigo=codigo))
            layout.addWidget(lbl_articulo)

            lbl_actual = QLabel(tr("repo.current_label", default="Stock Esperado Actual: {valor}", valor=stock_esperado_actual))
            layout.addWidget(lbl_actual)

            layout.addSpacing(10)

            lbl_nuevo = QLabel(tr("repo.new_label", default="Nuevo Stock Esperado:"))
            layout.addWidget(lbl_nuevo)

            self.input_nuevo_stock = QLineEdit(str(stock_esperado_actual))
            self.input_nuevo_stock.setStyleSheet(InformeReposicionWindow._NEON_INPUT_SS)
            self.input_nuevo_stock.setFixedWidth(200)
            layout.addWidget(
                self.input_nuevo_stock, alignment=Qt.AlignmentFlag.AlignCenter
            )

            layout.addStretch(1)

            btn_row = QHBoxLayout()
            btn_aceptar = QPushButton(tr("repo.accept", default="ACEPTAR"))
            btn_aceptar.setStyleSheet(InformeReposicionWindow._BTN_CIAN_SS)
            btn_aceptar.clicked.connect(self.accept)
            InformeReposicionWindow._sombra_cian(parent, btn_aceptar)

            btn_cancelar = QPushButton(tr("repo.cancel", default="CANCELAR"))
            btn_cancelar.setStyleSheet(InformeReposicionWindow._BTN_ROJO_SS)
            btn_cancelar.clicked.connect(self.reject)
            InformeReposicionWindow._sombra_cian(parent, btn_cancelar)

            btn_row.addWidget(btn_aceptar)
            btn_row.addWidget(btn_cancelar)
            layout.addLayout(btn_row)

        def get_nuevo_stock(self):
            try:
                return int(self.input_nuevo_stock.text())
            except ValueError:
                return None

    class _EditarStockEsperadoPage(QWidget):
        def __init__(self, parent=None):
            super().__init__(parent)
            layout = QVBoxLayout(self)
            layout.setSpacing(30)
            layout.addStretch(1)
            self.main_window = (
                parent  # Reference to InformeReposicionWindow for signals and constants
            )

            lbl_icon = QLabel("✏️")
            lbl_icon.setStyleSheet("font-size: 160px;")
            lbl_icon.setFixedHeight(220)
            layout.addWidget(lbl_icon, alignment=Qt.AlignmentFlag.AlignCenter)
            layout.addSpacing(20)

            self._lbl_text = lbl_text = QLabel(tr("repo.edit_page_title", default="EDITAR STOCK ESPERADO"))
            lbl_text.setStyleSheet(
                f"color: {InformeReposicionWindow._CIAN}; font-size: 24px; font-weight: bold;"
            )
            layout.addWidget(lbl_text, alignment=Qt.AlignmentFlag.AlignCenter)

            self.search_bar = QLineEdit()
            self.search_bar.setPlaceholderText(
                tr("repo.search_ph", default="Introduce código o nombre para editar stock esperado...")
            )
            self.search_bar.setStyleSheet(InformeReposicionWindow._NEON_INPUT_SS)
            self.search_bar.setFixedWidth(500)
            self.search_bar.returnPressed.connect(self._buscar)
            layout.addWidget(self.search_bar, alignment=Qt.AlignmentFlag.AlignCenter)

            self._btn_buscar = btn_buscar = QPushButton(tr("repo.search_btn", default="BUSCAR ARTÍCULO"))
            btn_buscar.setStyleSheet(InformeReposicionWindow._BTN_CIAN_SS)
            btn_buscar.setFixedWidth(250)
            btn_buscar.clicked.connect(self._buscar)
            InformeReposicionWindow._sombra_cian(parent, btn_buscar)
            layout.addWidget(btn_buscar, alignment=Qt.AlignmentFlag.AlignCenter)

            layout.addStretch(1)

        def _retraducir(self):
            self._lbl_text.setText(tr("repo.edit_page_title", default="EDITAR STOCK ESPERADO"))
            self.search_bar.setPlaceholderText(
                tr("repo.search_ph", default="Introduce código o nombre para editar stock esperado...")
            )
            self._btn_buscar.setText(tr("repo.search_btn", default="BUSCAR ARTÍCULO"))

        def _buscar(self):
            termino = self.search_bar.text().strip()
            if not termino:
                QMessageBox.warning(
                    self,
                    tr("repo.search_warn_title", default="Búsqueda"),
                    tr("repo.search_warn_msg", default="Por favor, introduce un código o nombre de artículo."),
                )
                return

            articulo = obtener_articulo(termino)  # This now returns a dict
            if not articulo:
                # Try searching by name if not found by code
                with obtener_conexion() as conn:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT codigo, nombre, Stock_esperado FROM articulos WHERE nombre LIKE %s",
                        (f"%{termino}%",),
                    )
                    res_tuple = cur.fetchone()
                    if res_tuple:
                        # Convert tuple to dict manually for consistency
                        articulo = {
                            "codigo": res_tuple[0],
                            "nombre": res_tuple[1],
                            "Stock_esperado": res_tuple[2],
                        }
                if not articulo:
                    QMessageBox.warning(
                        self,
                        tr("repo.not_found_title", default="No encontrado"),
                        tr("repo.not_found_msg", default="No se encontró el artículo: {termino}", termino=termino),
                    )
                    return

            codigo = articulo.get("codigo")
            nombre = articulo.get("nombre")
            stock_esperado_actual = int(articulo.get("Stock_esperado", 0))

            dialog = InformeReposicionWindow._EditarStockEsperadoDialog(
                self, codigo, nombre, stock_esperado_actual
            )
            if dialog.exec() == QDialog.DialogCode.Accepted:
                nuevo_stock = dialog.get_nuevo_stock()
                if nuevo_stock is not None and nuevo_stock >= 0:
                    if set_stock_esperado(codigo, nuevo_stock):
                        QMessageBox.information(
                            self,
                            tr("repo.success_title", default="Éxito"),
                            tr("repo.updated_msg",
                               default="Stock esperado de '{nombre}' actualizado a {valor}.",
                               nombre=nombre, valor=nuevo_stock),
                        )
                        self.main_window.signals.stock_actualizado.emit(
                            codigo
                        )  # Emit signal to update tables
                        # Trigger update for the Estado Reposicion page if it's the current view
                        if (
                            self.main_window._vistas.currentWidget()
                            == self.main_window._page_estado
                        ):
                            self.main_window._page_estado.cargar_datos()
                    else:
                        QMessageBox.critical(
                            self,
                            tr("repo.error_title", default="Error"),
                            tr("repo.update_err_msg",
                               default="No se pudo actualizar el stock esperado de '{nombre}'.", nombre=nombre),
                        )
                else:
                    QMessageBox.warning(
                        self,
                        tr("repo.invalid_title", default="Entrada inválida"),
                        tr("repo.invalid_msg", default="Por favor, introduce un número válido para el stock esperado."),
                    )
            self.search_bar.clear()

    class _ExportarInformePage(QWidget):
        def __init__(self, parent=None):
            super().__init__(parent)
            layout = QVBoxLayout(self)
            layout.setSpacing(30)
            layout.addStretch(1)
            self.main_window = parent

            lbl_icon = QLabel("📥")
            lbl_icon.setStyleSheet("font-size: 160px;")
            lbl_icon.setFixedHeight(200)
            layout.addWidget(lbl_icon, alignment=Qt.AlignmentFlag.AlignCenter)
            layout.addSpacing(10)

            self._lbl_text = lbl_text = QLabel(tr("repo.export_page_title", default="EXPORTAR INFORME DE REPOSICIÓN"))
            lbl_text.setStyleSheet(
                f"color: {InformeReposicionWindow._CIAN}; font-size: 22px; font-weight: bold;"
            )
            layout.addWidget(lbl_text, alignment=Qt.AlignmentFlag.AlignCenter)

            self._btn_export = btn_export = QPushButton(tr("repo.export_btn", default="EXPORTAR Y MARCAR COMO REPUESTO"))
            btn_export.setStyleSheet(InformeReposicionWindow._BTN_CIAN_SS)
            btn_export.setFixedSize(380, 60)
            btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_export.clicked.connect(self._exportar_flujo)
            InformeReposicionWindow._sombra_cian(parent, btn_export)
            layout.addWidget(btn_export, alignment=Qt.AlignmentFlag.AlignHCenter)

            layout.addStretch(1)

        def _retraducir(self):
            self._lbl_text.setText(tr("repo.export_page_title", default="EXPORTAR INFORME DE REPOSICIÓN"))
            self._btn_export.setText(tr("repo.export_btn", default="EXPORTAR Y MARCAR COMO REPUESTO"))

        def _exportar_flujo(self):
            # 1. Obtener artículos bajo el umbral
            articulos_a_reponer = []
            try:
                with obtener_conexion() as conn:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT codigo, nombre, COALESCE(Stock_total, 0), COALESCE(Stock_tienda, 0), COALESCE(Stock_esperado, 0) FROM articulos"
                    )
                    for row in cur.fetchall():
                        codigo, nombre, s_alm, s_lin, s_esp = row
                        if s_lin < (s_esp * 0.7):
                            uds_reponer = s_esp - s_lin
                            if uds_reponer > 0:
                                articulos_a_reponer.append(
                                    {
                                        "Código": codigo,
                                        "Nombre": nombre,
                                        "Unidades a reponer": uds_reponer,
                                        "s_esp": s_esp,
                                        "s_lin": s_lin,
                                        "s_alm": s_alm,
                                    }
                                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    tr("repo.error_title", default="Error"),
                    tr("repo.db_query_err", default="Error al consultar base de datos: {e}", e=e),
                )
                return

            if not articulos_a_reponer:
                QMessageBox.information(
                    self,
                    tr("repo.nodata_title", default="Sin datos"),
                    tr("repo.nodata_msg", default="No hay artículos que necesiten reposición en este momento."),
                )
                return

            # 2. Confirmación
            res = QMessageBox.question(
                self,
                tr("repo.confirm_title", default="Confirmar Exportación"),
                tr("repo.confirm_msg",
                   default="Se va a generar un informe con {n} artículos.\n¿Desea continuar y actualizar el stock en el sistema?",
                   n=len(articulos_a_reponer)),
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if res != QMessageBox.StandardButton.Yes:
                return

            # 3. Generar Excel
            try:
                # Carpeta de destino
                ruta_base = os.path.abspath(os.getcwd())
                folder = os.path.join(ruta_base, "documentos", "informes_reposicion")
                os.makedirs(folder, exist_ok=True)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"informe_reposicion_{timestamp}.xlsx"
                ruta_excel = os.path.join(folder, filename)

                # Crear DataFrame solo con las 3 columnas solicitadas (cabeceras traducidas)
                df = pd.DataFrame(articulos_a_reponer)[
                    ["Código", "Nombre", "Unidades a reponer"]
                ]
                df.columns = [
                    tr("repo.exp_code", default="Código"),
                    tr("repo.exp_name", default="Nombre"),
                    tr("repo.exp_units", default="Unidades a reponer"),
                ]
                _hoja = tr("repo.exp_sheet", default="Reposición")

                with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name=_hoja)
                    workbook = writer.book
                    worksheet = writer.sheets[_hoja]

                    # Formato: Segoe UI Bold y Cabeceras Turquesas
                    from openpyxl.styles import Alignment, Font, PatternFill

                    header_font = Font(name="Segoe UI", bold=True, color="000000")
                    header_fill = PatternFill(
                        start_color="00FFC6", end_color="00FFC6", fill_type="solid"
                    )
                    data_font = Font(name="Segoe UI", bold=True)

                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.font = data_font
                            cell.alignment = Alignment(horizontal="center")

                    # Ajustar ancho de columnas
                    worksheet.column_dimensions["A"].width = 15
                    worksheet.column_dimensions["B"].width = 40
                    worksheet.column_dimensions["C"].width = 20

                # 4. Actualizar Base de Datos (Marcar como repuesto)
                with obtener_conexion() as conn:
                    cur = conn.cursor()
                    for art in articulos_a_reponer:
                        codigo = art["Código"]
                        reponer = art["Unidades a reponer"]
                        nuevo_lineal = art["s_esp"]
                        # Restar del almacén lo que se mueve al lineal
                        nuevo_almacen = max(0, art["s_alm"] - reponer)

                        cur.execute(
                            "UPDATE articulos SET Stock_tienda = %s, Stock_total = %s WHERE codigo = %s",
                            (nuevo_lineal, nuevo_almacen, codigo),
                        )
                        # Emitir señal global para otros módulos
                        self.main_window.signals.stock_actualizado.emit(str(codigo))
                    conn.commit()

                # H3: integra la reposición (movimiento almacén→lineal) en kárdex y
                # sincroniza el ledger stock_almacen (evita divergencia caché↔ledger).
                try:
                    from src.db import kardex
                    from src.db import stock_almacen as _SA
                    for art in articulos_a_reponer:
                        cod = art["Código"]; mov = int(art.get("Unidades a reponer") or 0)
                        if cod and mov > 0:
                            kardex.registrar_movimiento(cod, "TRASPASO", mov, origen="ALMACEN",
                                                        destino="LINEAL",
                                                        observaciones="Reposición a lineal")
                        if cod and _SA.esta_gestionado(cod):
                            _SA.reseed_articulo(cod)
                except Exception:
                    pass

                # 5. Limpiar pestaña de Estado Reposición y notificar
                if hasattr(self.main_window, "_page_estado"):
                    self.main_window._page_estado.cargar_datos()

                QMessageBox.information(
                    self,
                    tr("repo.success_title", default="Éxito"),
                    tr("repo.export_ok_msg",
                       default="Informe exportado correctamente en:\n{ruta}\n\nStock actualizado en el sistema.",
                       ruta=ruta_excel),
                )
                # Abrir la carpeta
                os.startfile(folder)

            except Exception as e:
                QMessageBox.critical(
                    self,
                    tr("repo.error_title", default="Error"),
                    tr("repo.export_fail_msg", default="No se pudo completar la operación: {e}", e=e),
                )

    # ---------------------------------------------------------------------------
    # VENTANA PRINCIPAL
    # ---------------------------------------------------------------------------

    def setup_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ---- SIDEBAR ----
        sidebar = QFrame()
        sidebar.setFixedWidth(280)
        sidebar.setStyleSheet(
            f"background-color: {self._PANEL_BG}; border-right: 1px solid {self._BORDE};"
        )

        side_ly = QVBoxLayout(sidebar)
        side_ly.setContentsMargins(0, 40, 0, 20)
        side_ly.setSpacing(0)

        lbl_m = QLabel(tr("repo.smart_reposicion", default="SMART REPOSICIÓN"))
        lbl_m.setStyleSheet(
            "color: #ffffff; font-size: 16px; font-weight: 900; margin-left: 30px; "
            "margin-bottom: 35px; letter-spacing: 2px; border: none; background: transparent;"
        )
        side_ly.addWidget(lbl_m)

        self._tab_keys = ["repo.tab_status", "repo.tab_edit", "repo.tab_export"]
        _tab_def = ["ESTADO REPOSICIÓN", "EDITAR STOCK ESPERADO", "EXPORTAR INFORME"]

        self._nav_btns = []
        for idx, key in enumerate(self._tab_keys):
            btn = self._SidebarBtn(tr(key, default=_tab_def[idx]))
            btn.setCheckable(True)
            btn.setAutoExclusive(True)
            btn.clicked.connect(lambda _, i=idx: self._ir_a(i))
            side_ly.addWidget(btn)
            self._nav_btns.append(btn)

        side_ly.addStretch()

        self._btn_exit = btn_exit = self._SidebarBtn(tr("repo.exit", default="SALIR AL MENÚ"))
        btn_exit.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #F85149;
                border: none;
                border-left: 4px solid transparent;
                border-radius: 0px;
                font-size: 12px;
                font-family: 'Segoe UI';
                font-weight: 900;
                text-align: left;
                padding-left: 28px;
            }
            QPushButton:hover {
                background-color: #F85149;
                color: #0E1117;
            }
        """)
        btn_exit.clicked.connect(self.volver_menu_principal)
        side_ly.addWidget(btn_exit)
        root.addWidget(sidebar)

        # ---- CONTENT AREA ----
        self._vistas = QStackedWidget()
        self._page_estado = self._EstadoReposicionPage(self)
        self._page_editar = self._EditarStockEsperadoPage(self)
        self._page_exportar = self._ExportarInformePage(self)

        self._vistas.addWidget(self._page_estado)
        self._vistas.addWidget(self._page_editar)
        self._vistas.addWidget(self._page_exportar)

        root.addWidget(self._vistas)
        self._ir_a(0)

        self.setLayout(root)
        self.setStyleSheet(f"background-color: {self._FONDO};")

    # ============================================================
    # BLOQUE ESTILO DE BOTONES
    # ============================================================

    def estilo_boton(self, btn, rojo=False):
        if rojo:
            base, hover, text_color, padding = "#FF4B4B", "#FF2222", "#FFFFFF", "12px"
        else:
            base, hover, text_color, padding = "#00FFC6", "#00DDAA", "#0E1117", "20px"

        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {base};
                color: {text_color};
                font-weight: bold;
                border-radius: 15px;
                padding: {padding};
            }}
            QPushButton:hover {{
                background-color: {hover};
            }}
        """)
        btn.setFont(QFont("Segoe UI", 11))
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(25)
        shadow.setColor(QColor(base))
        shadow.setOffset(0)
        btn.setGraphicsEffect(shadow)

    def _convert_to_int(self, valor):
        try:
            return int(valor)
        except (ValueError, TypeError):
            return 0

    # ============================================================
    # BLOQUE CONSULTA DE DATOS DE REPOSICIÓN
    # ============================================================

    def obtener_articulos_reposicion(self):
        """Obtiene artículos bajo el umbral usando el context manager."""
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT codigo, nombre, Stock_total, Stock_tienda, 
                           COALESCE(stock_esperado, 0), capacidad_lineal
                    FROM articulos
                    ORDER BY nombre ASC
                    """)
                resultados = []
                for (
                    codigo,
                    nombre,
                    stock_total,
                    stock_tienda,
                    stock_esperado,
                    capacidad_lineal,
                ) in cur.fetchall():

                    stock_total = self._convert_to_int(stock_total)
                    stock_tienda = self._convert_to_int(stock_tienda)
                    stock_esperado = self._convert_to_int(stock_esperado)

                    if stock_esperado == 0 and capacidad_lineal:
                        stock_esperado = capacidad_lineal // 2

                    if stock_tienda < stock_esperado:
                        resultados.append(
                            (codigo, nombre, stock_total, stock_tienda, stock_esperado)
                        )
                return resultados
        except Exception as e:
            print(f"[ERROR] No se pudieron obtener los artículos: {e}")
            return []

    # ============================================================
    # BLOQUE CARGA Y ACTUALIZACIÓN DE TABLA (Lógica para Parte 2)
    # ============================================================

    def actualizar_lista(self, show_message=False):
        try:
            articulos = self.obtener_articulos_reposicion()
            self.tabla.setRowCount(0)

            if not articulos:
                if show_message:
                    QMessageBox.information(
                        self,
                        tr("repo.no_results_title", default="Sin resultados"),
                        tr("repo.no_results_msg", default="No hay artículos bajo el umbral para reposición."),
                    )
                return

            for codigo, nombre, stock_total, stock_tienda, stock_esperado in articulos:
                row_pos = self.tabla.rowCount()
                self.tabla.insertRow(row_pos)

                if stock_tienda < stock_esperado:
                    color = QColor(80, 0, 0)
                elif stock_tienda > stock_esperado:
                    color = QColor(0, 80, 0)
                else:
                    color = QColor(26, 29, 35)

                for col, value in enumerate(
                    [codigo, nombre, stock_total, stock_tienda, stock_esperado]
                ):
                    item = QTableWidgetItem(str(value))
                    item.setBackground(color)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.tabla.setItem(row_pos, col, item)

            self.tabla.resizeColumnsToContents()
            self.tabla.resizeRowsToContents()
            self.tabla.horizontalHeader().setStretchLastSection(True)

            if show_message:
                QMessageBox.information(
                    self,
                    tr("repo.updated_title", default="Actualizado"),
                    tr("repo.list_updated_msg", default="Lista de reposición actualizada correctamente."),
                )

        except Exception as e:
            QMessageBox.critical(
                self,
                tr("repo.error_title", default="Error"),
                tr("repo.list_load_err", default="No se pudo cargar la lista:\n{e}", e=e),
            )

    def _on_stock_actualizado(self, codigo):
        try:
            articulo = obtener_articulo(codigo)
            if not articulo:
                return
            stock_total, stock_tienda, capacidad_lineal, stock_esperado = (
                self._convert_to_int(articulo[2]),
                self._convert_to_int(articulo[3]),
                articulo[5],
                self._convert_to_int(articulo[7]),
            )
            if stock_esperado == 0 and capacidad_lineal:
                stock_esperado = capacidad_lineal // 2
            nombre = articulo[1]

            for r in range(self.tabla.rowCount()):
                if self.tabla.item(r, 0).text() == codigo:
                    if stock_tienda < stock_esperado:
                        color = QColor(80, 0, 0)
                    elif stock_tienda > stock_esperado:
                        color = QColor(0, 80, 0)
                    else:
                        color = QColor(26, 29, 35)

                    for col, value in enumerate(
                        [codigo, nombre, stock_total, stock_tienda, stock_esperado]
                    ):
                        if not self.tabla.item(r, col):
                            self.tabla.setItem(r, col, QTableWidgetItem(str(value)))
                        else:
                            self.tabla.item(r, col).setText(str(value))
                        self.tabla.item(r, col).setBackground(color)
                        self.tabla.item(r, col).setTextAlignment(
                            Qt.AlignmentFlag.AlignCenter
                        )
                    break
        except Exception as e:
            print(f"Error actualizando artículo {codigo}: {e}")

    # ============================================================
    # BLOQUE EDICIÓN DE STOCK ESPERADO (Lógica para Parte 3)
    # ============================================================

    def editar_stock_esperado(self):
        try:
            codigo, ok = QInputDialog.getText(
                self,
                tr("repo.leg_edit_code_title", default="Editar Stock_esperado"),
                tr("repo.leg_edit_code_prompt", default="Introduce el código del artículo:"),
            )
            if not ok or not codigo:
                return
            nuevo_stock, ok = QInputDialog.getInt(
                self,
                tr("repo.leg_edit_val_title", default="Nuevo Stock_esperado"),
                tr("repo.leg_edit_val_prompt", default="Introduce el nuevo valor:"),
                0, 0, 99999,
            )
            if not ok:
                return

            if not set_stock_esperado(codigo, nuevo_stock):
                QMessageBox.critical(
                    self,
                    tr("repo.error_title", default="Error"),
                    tr("repo.leg_db_err_msg", default="No se pudo actualizar el Stock_esperado en la base de datos."),
                )
                return

            self.signals.stock_actualizado.emit(codigo)
            QMessageBox.information(
                self,
                tr("repo.updated_title", default="Actualizado"),
                tr("repo.leg_updated_msg", default="Stock_esperado de {codigo} actualizado a {valor}.",
                   codigo=codigo, valor=nuevo_stock),
            )
            self.actualizar_lista()
        except Exception as e:
            QMessageBox.critical(
                self,
                tr("repo.error_title", default="Error"),
                tr("repo.leg_update_err", default="No se pudo actualizar el Stock_esperado:\n{e}", e=e),
            )

    # ============================================================
    # BLOQUE EXPORTACIÓN DE INFORME (Lógica para Parte 4)
    # ============================================================

    def exportar_y_marcar(self):
        try:
            filas_exportar = []

            for r in range(self.tabla.rowCount()):
                stock_tienda = self._convert_to_int(self.tabla.item(r, 3).text())
                stock_esperado = self._convert_to_int(self.tabla.item(r, 4).text())
                if stock_tienda < stock_esperado:
                    filas_exportar.append(
                        [self.tabla.item(r, c).text() for c in range(5)]
                    )

            if not filas_exportar:
                QMessageBox.warning(
                    self,
                    tr("repo.warn_title", default="Aviso"),
                    tr("repo.no_export_msg", default="No hay artículos bajo el umbral para exportar."),
                )
                return

            base_dir = os.path.abspath(
                os.path.join(os.path.dirname(__file__), "../../")
            )
            export_dir = os.path.join(base_dir, "documentos")
            os.makedirs(export_dir, exist_ok=True)

            nombre_archivo = (
                f"informe_reposicion_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
            )
            ruta_export = os.path.join(export_dir, nombre_archivo)

            df = pd.DataFrame(
                filas_exportar,
                columns=[
                    "Código",
                    "Nombre",
                    "Stock_total (almacén)",
                    "Stock_tienda",
                    "Stock_esperado",
                ],
            )
            df.to_excel(ruta_export, index=False)

            try:
                with obtener_conexion() as conn:
                    cur = conn.cursor()
                    hoy = datetime.now().strftime("%Y-%m-%d")

                    for fila in filas_exportar:
                        codigo = fila[0]
                        stock_esperado = self._convert_to_int(fila[4])

                        cur.execute(
                            """
                            UPDATE articulos 
                            SET repuesto = 1, Stock_tienda = ?, ultima_recepcion = ? 
                            WHERE codigo = ?
                            """,
                            (stock_esperado, hoy, codigo),
                        )
                    conn.commit()

                for fila in filas_exportar:
                    self.signals.stock_actualizado.emit(fila[0])

                QMessageBox.information(
                    self,
                    tr("repo.success_title", default="Éxito"),
                    tr("repo.leg_export_ok_msg",
                       default="Informe exportado y stock actualizado correctamente:\n{ruta}", ruta=ruta_export),
                )
                self.actualizar_lista()

            except Exception as db_error:
                QMessageBox.critical(
                    self,
                    tr("repo.db_error_title", default="Error de Base de Datos"),
                    tr("repo.db_update_err", default="No se pudo actualizar la BD: {db_error}", db_error=db_error),
                )

        except Exception as e:
            QMessageBox.critical(
                self,
                tr("repo.error_title", default="Error"),
                tr("repo.export_report_err", default="No se pudo exportar el informe:\n{e}", e=e),
            )

    # ============================================================
    # BLOQUE NAVEGACIÓN DE PESTAÑAS
    # ============================================================

    def _retraducir(self):
        self.setWindowTitle(tr("repo.window_title", default="Smart Manager - Informe de Reposición"))
        _tab_def = ["ESTADO REPOSICIÓN", "EDITAR STOCK ESPERADO", "EXPORTAR INFORME"]
        for i, btn in enumerate(self._nav_btns):
            btn.setText(tr(self._tab_keys[i], default=_tab_def[i]))
        self._btn_exit.setText(tr("repo.exit", default="SALIR AL MENÚ"))
        for page in (self._page_estado, self._page_editar, self._page_exportar):
            if hasattr(page, "_retraducir"):
                page._retraducir()

    def _ir_a(self, index):
        self._vistas.setCurrentIndex(index)
        for i, btn in enumerate(self._nav_btns):
            btn.setChecked(i == index)
            repolish_widget(btn)

    def volver_menu_principal(self):
        """Regresa al menú principal cerrando la ventana actual."""
        if self.callback_vuelta:
            self.callback_vuelta()
            self.close()
        else:
            self.close()
