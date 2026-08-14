"""
Smart Manager - Módulo de Stock
Vista lateral (sidebar) con 6 pestañas:
  0 · Stock Tienda
  1 · Stock Almacén Central
  2 · Editar Stock
  3 · Importar Stock
  4 · Exportar Stock
  5 · Inventario
"""

import os
from datetime import datetime

import pandas as pd
from PyQt6.QtCore import QObject, Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QStackedWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from assets.estilo_global import (
    aplicar_estilo_widget,
    construir_tabla_estilizada,
    repolish_widget,
)
from src.db.conexion import (
    ensure_schema,
    modificar_stock_completo,
)
from src.utils import i18n
from src.utils.i18n import tr

# ---------------------------------------------------------------------------
# Module-level signals (backward-compat: imported by src.db.conexion)
# ---------------------------------------------------------------------------


class StockSignals(QObject):
    stock_actualizado = pyqtSignal(str)


stock_signals = StockSignals()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
_CIAN = "#00FFC6"
_FONDO = "#0E1117"
_PANEL_BG = "#161B22"
_GRIS_PANEL = "#1A1D23"
_BORDE = "#30363D"

# Neon input
_NEON_INPUT_SS = f"""
QLineEdit {{
    background-color: #161B22;
    color: #FFFFFF;
    border: 2px solid {_CIAN};
    border-radius: 12px;
    padding: 8px 14px;
    font-size: 13px;
    font-family: 'Segoe UI';
}}
QLineEdit:focus {{
    border: 2px solid #00E6B2;
    background-color: #1A2230;
    outline: none;
}}
"""

# Cyan action button (no focus rect, no shadow — add shadow manually only on main page buttons)
_BTN_CIAN_SS = f"""
QPushButton {{
    background-color: #0E1117;
    color: {_CIAN};
    font-weight: bold;
    border-radius: 14px;
    padding: 12px 24px;
    font-size: 13px;
    font-family: 'Segoe UI';
    border: 2px solid {_CIAN};
    outline: none;
}}
QPushButton:hover {{
    background-color: {_CIAN};
    color: #0E1117;
    border: 2px solid {_CIAN};
}}
QPushButton:pressed {{
    background-color: #00C79A;
    color: #0E1117;
}}
QPushButton:focus {{
    outline: none;
}}
"""

# Green save button
_BTN_VERDE_SS = """
QPushButton {
    background-color: #2EA043;
    color: #000000;
    font-weight: bold;
    border-radius: 12px;
    padding: 10px 28px;
    font-size: 13px;
    font-family: 'Segoe UI';
    border: none;
    outline: none;
}
QPushButton:hover {
    background-color: #FFFFFF;
    color: #000000;
}
QPushButton:pressed {
    background-color: #238636;
    color: #000000;
}
QPushButton:focus {
    outline: none;
}
"""

# Red cancel button (no focus rect)
_BTN_ROJO_SS = """
QPushButton {
    background-color: #0E1117;
    color: #FF4B4B;
    font-weight: bold;
    border-radius: 14px;
    padding: 10px 20px;
    font-size: 12px;
    font-family: 'Segoe UI';
    border: 2px solid #FF4B4B;
    outline: none;
}
QPushButton:hover {
    background-color: #FF4B4B;
    color: #0E1117;
    border: 2px solid #FF4B4B;
}
QPushButton:focus {
    outline: none;
}
"""

# Shared description panel background
_DESC_SS = (
    f"color: #8B949E; font-size: 13px; "
    f"background-color: {_PANEL_BG}; border-radius: 12px; padding: 14px; "
    f"border: 1px solid {_BORDE};"
)

# Textos descriptivos por defecto (fallback si falta la clave i18n).
_IMPORT_DESC_DEFAULT = (
    "Importa artículos y niveles de stock desde un fichero Excel (.xlsx) o CSV/TXT. "
    "Las columnas del fichero deben coincidir con los campos de la base de datos. "
    "Columnas requeridas:\n"
    "  • codigo          → Código del artículo (obligatorio, clave única)\n"
    "  • nombre          → Nombre del artículo\n"
    "  • Stock_tienda    → Stock lineal (expuesto en tienda)\n"
    "  • Stock_total     → Stock almacén tienda\n"
    "  • Stock_central   → Stock almacén central\n"
    "  • Stock_esperado  → Stock mínimo esperado en el lineal\n\n"
    "Columnas opcionales: descripcion, categoria, seccion, precio"
)
_EXPORT_DESC_DEFAULT = (
    "Genera un informe Excel con todos los niveles de stock actuales. "
    "El fichero se guarda automáticamente en documentos/stocks/. "
    "El informe incluye: Código, Nombre, Stock Lineal, Stock Almacén,\n"
    "Stock Almacén Central y Stock Esperado de cada artículo."
)
_INVENTORY_DESC_DEFAULT = (
    "Accede a la carpeta compartida de inventario en Google Drive "
    "para ver y gestionar los documentos de inventario. "
    "Se abrirá tu navegador predeterminado con la carpeta de\n"
    "Google Drive configurada para este negocio."
)


# ---------------------------------------------------------------------------
# Sidebar button: calls aplicar_estilo_widget like SidebarButton in recepcion_pale.py
# ---------------------------------------------------------------------------
class _SidebarBtn(QPushButton):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setObjectName("btn_sidebar")
        self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setMouseTracking(True)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                border: none;
                border-left: 4px solid transparent;
                border-radius: 0px;
                font-size: 12px;
                font-family: 'Segoe UI';
                font-weight: 900;
                text-align: left;
                padding-left: 28px;
                color: #FFFFFF;
            }}
            QPushButton:hover {{
                background-color: #FFFFFF;
                color: #0E1117;
            }}
            QPushButton:checked {{
                background-color: #1A2230;
                border-left: 4px solid {_CIAN};
                color: {_CIAN};
            }}
        """)
        try:
            aplicar_estilo_widget(self)
        except Exception:
            pass

    def enterEvent(self, event):
        super().enterEvent(event)
        repolish_widget(self)

    def leaveEvent(self, event):
        super().leaveEvent(event)
        repolish_widget(self)


# ---------------------------------------------------------------------------
# Helper: shadow effect
# ---------------------------------------------------------------------------
def _sombra_cian(widget):
    fx = QGraphicsDropShadowEffect()
    fx.setBlurRadius(22)
    fx.setColor(QColor(_CIAN))
    fx.setOffset(0)
    widget.setGraphicsEffect(fx)


# ---------------------------------------------------------------------------
# Helper: build styled table with top-margin fix for corner cuts
# ---------------------------------------------------------------------------
def _crear_tabla(parent, cols):
    contenedor, tabla = construir_tabla_estilizada(parent)
    tabla.setStyleSheet(f"""
        QTableWidget {{
            border: none;
            background-color: transparent;
            outline: none;
        }}
        QHeaderView {{
            background-color: transparent;
            border: none;
        }}
        QHeaderView::section {{
            background-color: #1A1D23;
            color: {_CIAN};
            border: none;
        }}
        QHeaderView::section:hover {{
            background-color: {_CIAN};
            color: #0E1117;
        }}
        QHeaderView::section:first {{
            border-top-left-radius: 18px;
        }}
        QHeaderView::section:last {{
            border-top-right-radius: 18px;
        }}
    """)
    contenedor.layout().setContentsMargins(2, 2, 2, 2)
    tabla.setColumnCount(len(cols))
    tabla.setHorizontalHeaderLabels(cols)
    tabla.horizontalHeader().setStretchLastSection(True)
    tabla.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
    tabla.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
    tabla.setFocusPolicy(Qt.FocusPolicy.NoFocus)
    return contenedor, tabla


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def _buscar_articulos(query: str, id_familia=None):
    """Búsqueda por texto y, opcionalmente, por FAMILIA (id_familia=0 → artículos sin familia)."""
    # Cliente fino (Fase 3): búsqueda de stock en la capa de datos.
    from src.db.articulos import buscar_stock
    return buscar_stock(query, id_familia)


def _get_todos_articulos():
    from src.db.articulos import listar_codigo_nombre
    return listar_codigo_nombre()


def _get_articulo_stock(codigo: str):
    from src.db.articulos import obtener_stock
    return obtener_stock(codigo)


# ---------------------------------------------------------------------------
# Tab pages
# ---------------------------------------------------------------------------


class _StockTiendaPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 16)
        layout.setSpacing(14)

        self._lbl = lbl = QLabel(tr("stock.title_store", default="Stock Tienda"))
        lbl.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        lbl.setStyleSheet(f"color: {_CIAN};")
        layout.addWidget(lbl)

        fila_busq = QHBoxLayout(); fila_busq.setSpacing(10)
        self.buscador = QLineEdit()
        self.buscador.setPlaceholderText(tr("stock.search_ph", default="Buscar por código o nombre…"))
        self.buscador.setStyleSheet(_NEON_INPUT_SS)
        self.buscador.setFixedHeight(44)
        self.buscador.textChanged.connect(self._filtrar)
        fila_busq.addWidget(self.buscador, 1)

        # Filtro por FAMILIA de producto (reutiliza db/familias; "Todas" = sin filtro).
        self.cmb_familia = QComboBox()
        self.cmb_familia.setFixedHeight(44)
        self.cmb_familia.setMinimumWidth(210)
        self.cmb_familia.setStyleSheet(
            f"QComboBox{{background:{_PANEL_BG};color:#FFFFFF;border:2px solid {_CIAN};border-radius:12px;"
            f"padding:6px 14px;font-size:14px;font-weight:bold;}}"
            f"QComboBox::drop-down{{border:none;width:22px;}}"
            f"QComboBox QAbstractItemView{{background:{_PANEL_BG};color:#FFFFFF;border:1px solid {_CIAN};"
            f"selection-background-color:{_CIAN};selection-color:#0E1117;}}")
        self.cmb_familia.setCursor(Qt.CursorShape.PointingHandCursor)
        self._recargar_familias()
        self.cmb_familia.currentIndexChanged.connect(self._filtrar)
        fila_busq.addWidget(self.cmb_familia)
        layout.addLayout(fila_busq)

        # Fase 8 (IA predictiva VISIBLE en Smart Stock): tarjeta de previsión de demanda + riesgo de rotura,
        # reutilizando gui/prediccion_card ← services/prediccion (motor real). Degradable: si no hay datos
        # suficientes o el motor no está disponible, no se pinta (nunca inventa cifras/riesgo).
        self._cargar_ia_predictiva(layout)

        contenedor, self.tabla = _crear_tabla(self, self._cols())
        hh = self.tabla.horizontalHeader()
        hh.setStretchLastSection(False)
        for c in range(4):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(contenedor)
        self.cargar()

    def _cargar_ia_predictiva(self, layout):
        """Pinta la previsión de demanda y el riesgo de rotura (motor real, tenant actual). Silenciosa."""
        from PyQt6.QtWidgets import QHBoxLayout
        try:
            from src.db.empresa import empresa_actual_id
            from src.services.prediccion import consulta, riesgo_rotura
            from src.gui.prediccion_card import tarjeta_prevision, tarjeta_riesgo
            id_emp = empresa_actual_id()
            fila = QHBoxLayout(); fila.setSpacing(12)
            r = consulta.responder("previsión de demanda", id_emp, horizonte=30)
            if r.get("aplicable") and r.get("suficiente"):
                fila.addWidget(tarjeta_prevision(consulta.resumen_ui(r["detalle"])))
                # Riesgo de rotura agregado de empresa (demanda media prevista vs stock objetivo global).
                det = r["detalle"]; pred = det.get("prediccion") or []
                dem = (sum(pred) / len(pred)) if pred else 0
                rr = riesgo_rotura.evaluar(stock_actual=0, demanda_diaria=dem) if dem else {"nivel": "INSUFICIENTE",
                     "recomendacion": "No hay datos suficientes para calcular el riesgo."}
                fila.addWidget(tarjeta_riesgo(rr))
            fila.addStretch(1)
            if fila.count() > 1:            # sólo si hay al menos una tarjeta
                layout.addLayout(fila)
        except Exception:
            pass                           # la IA nunca rompe la pantalla de stock

    @staticmethod
    def _cols():
        return [
            tr("stock.col_ean", default="EAN"),
            tr("stock.col_item", default="Artículo"),
            tr("stock.col_shelf", default="Stock Lineal"),
            tr("stock.col_warehouse", default="Stock Almacén"),
        ]

    def _retraducir(self):
        self._lbl.setText(tr("stock.title_store", default="Stock Tienda"))
        self.buscador.setPlaceholderText(tr("stock.search_ph", default="Buscar por código o nombre…"))
        self.tabla.setHorizontalHeaderLabels(self._cols())

    def _recargar_familias(self):
        self.cmb_familia.blockSignals(True)
        self.cmb_familia.clear()
        self.cmb_familia.addItem(tr("stock.fam_all", default="Todas las familias"), None)
        self.cmb_familia.addItem(tr("stock.fam_none", default="— Sin familia —"), 0)
        try:
            from src.db.familias import listar_familias
            for f in listar_familias():
                self.cmb_familia.addItem(f["nombre"], f["id"])
        except Exception:
            pass
        self.cmb_familia.blockSignals(False)

    def cargar(self):
        self._aplicar()

    def _filtrar(self, *_):
        self._aplicar()

    def _aplicar(self):
        self._poblar(_buscar_articulos(self.buscador.text().strip(), self.cmb_familia.currentData()))

    def _poblar(self, rows):
        self.tabla.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, v in enumerate([str(row[0]), str(row[1]), str(row[2]), str(row[3])]):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.tabla.setItem(r, c, item)


class _EditChoiceDialog(QDialog):
    CANCELAR = 0
    LINEAL_ALMACEN = 1
    SOLO_LINEAL = 2
    SOLO_ALMACEN = 3

    def __init__(self, nombre_art: str, parent=None):
        super().__init__(parent)
        self.choice = self.CANCELAR
        self.setModal(True)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setMinimumWidth(400)

        _outer = QVBoxLayout(self)
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.setSpacing(0)
        _box = QFrame()
        _box.setObjectName("_editchoicebox")
        _box.setStyleSheet(f"""
            QFrame#_editchoicebox {{
                background-color: #161B22;
                border: 2px solid {_CIAN};
                border-radius: 16px;
            }}
            QLabel {{
                background: transparent;
                border: none;
            }}
        """)
        _outer.addWidget(_box)

        layout = QVBoxLayout(_box)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(12)

        lbl_art = QLabel(nombre_art)
        lbl_art.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        lbl_art.setStyleSheet("color: #FFFFFF;")
        lbl_art.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_art.setWordWrap(True)
        layout.addWidget(lbl_art)

        lbl_q = QLabel(tr("stock.choice_q", default="¿Qué stock deseas editar?"))
        lbl_q.setStyleSheet("color: #8B949E; font-size: 13px;")
        lbl_q.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_q)

        layout.addSpacing(6)

        _BTN_CHOICE_SS = f"""
            QPushButton {{
                background-color: #0E1117;
                color: {_CIAN};
                font-weight: bold;
                border-radius: 12px;
                padding: 12px 20px;
                font-size: 13px;
                font-family: 'Segoe UI';
                border: 2px solid {_CIAN};
                outline: none;
            }}
            QPushButton:hover {{
                background-color: {_CIAN};
                color: #0E1117;
            }}
            QPushButton:focus {{ outline: none; }}
        """
        _BTN_CANCEL_SS = f"""
            QPushButton {{
                background-color: #0E1117;
                color: {_CIAN};
                font-weight: bold;
                border-radius: 12px;
                padding: 10px 20px;
                font-size: 13px;
                font-family: 'Segoe UI';
                border: 2px solid {_CIAN};
                outline: none;
            }}
            QPushButton:hover {{
                background-color: {_CIAN};
                color: #0E1117;
            }}
            QPushButton:focus {{ outline: none; }}
        """

        for label, choice_val in [
            (tr("stock.choice_both", default="LINEAL Y ALMACÉN"), self.LINEAL_ALMACEN),
            (tr("stock.choice_shelf", default="SOLO LINEAL"), self.SOLO_LINEAL),
            (tr("stock.choice_warehouse", default="SOLO ALMACÉN"), self.SOLO_ALMACEN),
        ]:
            btn = QPushButton(label)
            btn.setStyleSheet(_BTN_CHOICE_SS)
            btn.setFixedHeight(46)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.clicked.connect(lambda _, cv=choice_val: self._elegir(cv))
            layout.addWidget(btn)

        btn_cancel = QPushButton(tr("stock.cancel", default="CANCELAR"))
        btn_cancel.setStyleSheet(_BTN_CANCEL_SS)
        btn_cancel.setFixedHeight(40)
        btn_cancel.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_cancel.clicked.connect(self.reject)
        layout.addWidget(btn_cancel)

    def _elegir(self, cv):
        self.choice = cv
        self.accept()


class _EditStockDialog(QDialog):
    def __init__(self, nombre_art: str, choice: int, lineal: int, almacen: int, parent=None):
        super().__init__(parent)
        self.new_lineal = None
        self.new_almacen = None

        self.setModal(True)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setMinimumWidth(400)

        _outer = QVBoxLayout(self)
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.setSpacing(0)
        _box = QFrame()
        _box.setObjectName("_editstockbox")
        _box.setStyleSheet(f"""
            QFrame#_editstockbox {{
                background-color: #161B22;
                border: 2px solid {_CIAN};
                border-radius: 16px;
            }}
            QLabel {{
                background: transparent;
                border: none;
            }}
        """)
        _outer.addWidget(_box)

        layout = QVBoxLayout(_box)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(12)

        lbl_art = QLabel(nombre_art)
        lbl_art.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        lbl_art.setStyleSheet("color: #FFFFFF;")
        lbl_art.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_art.setWordWrap(True)
        layout.addWidget(lbl_art)

        layout.addSpacing(4)

        show_lineal = choice in (_EditChoiceDialog.LINEAL_ALMACEN, _EditChoiceDialog.SOLO_LINEAL)
        show_almacen = choice in (_EditChoiceDialog.LINEAL_ALMACEN, _EditChoiceDialog.SOLO_ALMACEN)

        self._inp_lineal = None
        self._inp_almacen = None

        if show_lineal:
            row_ly = QHBoxLayout()
            lbl = QLabel(tr("stock.lbl_shelf", default="Stock Lineal:"))
            lbl.setStyleSheet("color: #8B949E; font-size: 13px;")
            lbl.setFixedWidth(130)
            self._inp_lineal = QLineEdit(str(lineal))
            self._inp_lineal.setStyleSheet(_NEON_INPUT_SS)
            self._inp_lineal.setFixedHeight(44)
            self._inp_lineal.setAlignment(Qt.AlignmentFlag.AlignCenter)
            row_ly.addWidget(lbl)
            row_ly.addWidget(self._inp_lineal)
            layout.addLayout(row_ly)

        if show_almacen:
            row_ly = QHBoxLayout()
            lbl = QLabel(tr("stock.lbl_warehouse", default="Stock Almacén:"))
            lbl.setStyleSheet("color: #8B949E; font-size: 13px;")
            lbl.setFixedWidth(130)
            self._inp_almacen = QLineEdit(str(almacen))
            self._inp_almacen.setStyleSheet(_NEON_INPUT_SS)
            self._inp_almacen.setFixedHeight(44)
            self._inp_almacen.setAlignment(Qt.AlignmentFlag.AlignCenter)
            row_ly.addWidget(lbl)
            row_ly.addWidget(self._inp_almacen)
            layout.addLayout(row_ly)

        layout.addSpacing(6)

        btn_guardar = QPushButton(tr("stock.save", default="GUARDAR"))
        btn_guardar.setStyleSheet(_BTN_VERDE_SS)
        btn_guardar.setFixedHeight(46)
        btn_guardar.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_guardar.clicked.connect(self._on_guardar)
        layout.addWidget(btn_guardar)

        _BTN_CANCEL_SS = f"""
            QPushButton {{
                background-color: #0E1117;
                color: {_CIAN};
                font-weight: bold;
                border-radius: 12px;
                padding: 10px 20px;
                font-size: 13px;
                font-family: 'Segoe UI';
                border: 2px solid {_CIAN};
                outline: none;
            }}
            QPushButton:hover {{
                background-color: {_CIAN};
                color: #0E1117;
            }}
            QPushButton:focus {{ outline: none; }}
        """
        btn_cancel = QPushButton(tr("stock.cancel", default="CANCELAR"))
        btn_cancel.setStyleSheet(_BTN_CANCEL_SS)
        btn_cancel.setFixedHeight(40)
        btn_cancel.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_cancel.clicked.connect(self.reject)
        layout.addWidget(btn_cancel)

    def _on_guardar(self):
        try:
            if self._inp_lineal is not None:
                self.new_lineal = int(self._inp_lineal.text() or "0")
            if self._inp_almacen is not None:
                self.new_almacen = int(self._inp_almacen.text() or "0")
        except ValueError:
            QMessageBox.warning(
                self,
                tr("stock.err_title", default="Error"),
                tr("stock.err_int_msg", default="Los valores deben ser números enteros."),
            )
            return
        self.accept()


class _EditarStockPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._todos = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 16)
        layout.setSpacing(14)

        self._lbl = lbl = QLabel(tr("stock.title_edit", default="Editar Stock"))
        lbl.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        lbl.setStyleSheet(f"color: {_CIAN};")
        layout.addWidget(lbl)

        self.buscador = QLineEdit()
        self.buscador.setPlaceholderText(tr("stock.search_ph_ean", default="Buscar por EAN o nombre…"))
        self.buscador.setStyleSheet(_NEON_INPUT_SS)
        self.buscador.setFixedHeight(44)
        self.buscador.textChanged.connect(self._filtrar)
        layout.addWidget(self.buscador)

        contenedor, self.tabla = _crear_tabla(self, self._cols())
        hh = self.tabla.horizontalHeader()
        hh.setStretchLastSection(False)
        for c in range(4):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        hh.resizeSection(4, 80)
        self.tabla.verticalHeader().setDefaultSectionSize(46)   # filas más altas: el lápiz no se corta
        self.tabla.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(contenedor)

        self.cargar()

    @staticmethod
    def _cols():
        return [
            tr("stock.col_ean", default="EAN"),
            tr("stock.col_item", default="Artículo"),
            tr("stock.col_shelf", default="Stock Lineal"),
            tr("stock.col_warehouse", default="Stock Almacén"),
            tr("stock.col_edit", default="Editar"),
        ]

    def _retraducir(self):
        self._lbl.setText(tr("stock.title_edit", default="Editar Stock"))
        self.buscador.setPlaceholderText(tr("stock.search_ph_ean", default="Buscar por EAN o nombre…"))
        self.tabla.setHorizontalHeaderLabels(self._cols())

    def cargar(self):
        self._todos = list(_buscar_articulos(""))
        self._poblar(self._todos)

    def _filtrar(self, texto):
        texto = texto.strip().lower()
        if not texto:
            self._poblar(self._todos)
        else:
            filtrados = [r for r in self._todos
                         if texto in str(r[0]).lower() or texto in str(r[1]).lower()]
            self._poblar(filtrados)

    def _poblar(self, rows):
        self.tabla.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for col_idx, val in enumerate([str(row[0]), str(row[1]), str(row[2]), str(row[3])]):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.tabla.setItem(r, col_idx, item)
            btn_edit = QPushButton("✏️")
            btn_edit.setFixedHeight(34)
            btn_edit.setStyleSheet("""
                QPushButton {
                    background: transparent;
                    border: none;
                    font-size: 16px;
                    padding: 0;
                }
                QPushButton:hover {
                    background-color: #1A2230;
                    border-radius: 6px;
                }
            """)
            btn_edit.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_edit.clicked.connect(lambda _, ri=r: self._on_edit_click(ri))
            self.tabla.setCellWidget(r, 4, btn_edit)

    def _on_edit_click(self, row_idx):
        item_ean = self.tabla.item(row_idx, 0)
        item_nom = self.tabla.item(row_idx, 1)
        item_lin = self.tabla.item(row_idx, 2)
        item_alm = self.tabla.item(row_idx, 3)
        if not item_ean or not item_nom:
            return

        nombre = f"{item_ean.text()} – {item_nom.text()}"
        try:
            lineal_actual = int(item_lin.text()) if item_lin else 0
        except ValueError:
            lineal_actual = 0
        try:
            almacen_actual = int(item_alm.text()) if item_alm else 0
        except ValueError:
            almacen_actual = 0

        dlg_choice = _EditChoiceDialog(nombre, self)
        if dlg_choice.exec() != QDialog.DialogCode.Accepted:
            return

        dlg_edit = _EditStockDialog(nombre, dlg_choice.choice, lineal_actual, almacen_actual, self)
        if dlg_edit.exec() != QDialog.DialogCode.Accepted:
            return

        codigo = item_ean.text()
        datos = _get_articulo_stock(codigo)
        if not datos:
            QMessageBox.critical(
                self,
                tr("stock.err_title", default="Error"),
                tr("stock.err_get_msg", default="No se pudo obtener el artículo."),
            )
            return

        new_lineal = dlg_edit.new_lineal if dlg_edit.new_lineal is not None else datos["lineal"]
        new_almacen = dlg_edit.new_almacen if dlg_edit.new_almacen is not None else datos["almacen"]

        ok = modificar_stock_completo(codigo, datos["central"], new_almacen, new_lineal)
        if ok:
            stock_signals.stock_actualizado.emit(codigo)
            self.tabla.clearSelection()
            if item_lin:
                item_lin.setText(str(new_lineal))
            if item_alm:
                item_alm.setText(str(new_almacen))
            for i, row in enumerate(self._todos):
                if str(row[0]) == codigo:
                    self._todos[i] = (row[0], row[1], new_lineal, new_almacen) + row[4:]
                    break
        else:
            QMessageBox.critical(
                self,
                tr("stock.err_title", default="Error"),
                tr("stock.err_upd_msg", default="No se pudo actualizar el stock."),
            )


class _ImportarHilo(QThread):
    """@deprecated (Strangler) — sustituido por el importador maestro `services.importacion` + la ventana
    `gui.migracion_gui.MigracionDatosWindow` (menú → "Migración de datos"), que hace upsert seguro por columnas
    mapeadas con `id_empresa`, familias y stock por kárdex, dry-run e idempotencia. Este hilo solo tocaba
    `articulos` y hacía `ALTER TABLE` por cada cabecera desconocida. Se conserva un ciclo y se eliminará cuando
    no queden referencias."""

    finalizado = pyqtSignal(str)

    def __init__(self, ruta):
        super().__init__()
        self.ruta = ruta

    def run(self):
        try:
            ext = os.path.splitext(self.ruta)[1].lower()
            if ext == ".xlsx":
                df = pd.read_excel(self.ruta)
            elif ext in (".csv", ".txt"):
                try:
                    df = pd.read_csv(self.ruta, sep="\t")
                except Exception:
                    df = pd.read_csv(self.ruta, sep=",")
            else:
                self.finalizado.emit(
                    tr("stock.fmt_incompatible",
                       default="Formato no compatible. Usa Excel (.xlsx) o CSV/TXT.")
                )
                return
            if df.empty:
                self.finalizado.emit(tr("stock.file_empty", default="El fichero está vacío."))
                return
            df.columns = [c.strip().lower() for c in df.columns]
            # Cliente fino (Fase 3): importación (ALTER dinámico + upsert) COMPARTIDA en la capa de datos.
            from src.db.articulos import importar_articulos_df
            importar_articulos_df(df)
            self.finalizado.emit(
                tr("stock.import_ok", default="Stock importado correctamente desde:\n{ruta}", ruta=self.ruta)
            )
        except Exception as e:
            self.finalizado.emit(tr("stock.import_err", default="Error al importar:\n{e}", e=e))


class _ImportarStockPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(16)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self._lbl_titulo = lbl_titulo = QLabel(tr("stock.title_import", default="Importar Stock"))
        lbl_titulo.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        lbl_titulo.setStyleSheet(f"color: {_CIAN};")
        layout.addWidget(lbl_titulo)

        self._lbl_desc = lbl_desc = QLabel(tr("stock.import_desc", default=_IMPORT_DESC_DEFAULT))
        lbl_desc.setStyleSheet(_DESC_SS)
        lbl_desc.setWordWrap(True)
        lbl_desc.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        layout.addWidget(lbl_desc)

        icon_lbl = QLabel("📦")
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_lbl.setStyleSheet(
            "font-size: 160px; background: transparent; border: none;"
        )
        icon_lbl.setFixedHeight(200)
        layout.addWidget(icon_lbl)

        self._btn = btn = QPushButton(tr("stock.import_btn", default="IMPORTAR STOCK"))
        btn.setStyleSheet(_BTN_CIAN_SS)
        btn.setMinimumSize(160, 62); btn.setMaximumWidth(320)  # responsive (P2): antes fijo 200x62
        btn.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn.clicked.connect(self._importar)
        _sombra_cian(btn)
        layout.addWidget(btn, alignment=Qt.AlignmentFlag.AlignHCenter)

        self.lbl_resultado = QLabel()
        self.lbl_resultado.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_resultado.setStyleSheet(f"color: {_CIAN}; font-size: 13px;")
        self.lbl_resultado.setWordWrap(True)
        self.lbl_resultado.setVisible(False)
        layout.addWidget(self.lbl_resultado)
        layout.addStretch()

    def _retraducir(self):
        self._lbl_titulo.setText(tr("stock.title_import", default="Importar Stock"))
        self._lbl_desc.setText(tr("stock.import_desc", default=_IMPORT_DESC_DEFAULT))
        self._btn.setText(tr("stock.import_btn", default="IMPORTAR STOCK"))

    def _importar(self):
        _all = tr("stock.all_files", default="Todos los archivos")
        ruta, _ = QFileDialog.getOpenFileName(
            self,
            tr("stock.file_dialog_title", default="Selecciona un fichero de stock"),
            "",
            f"Excel (*.xlsx);;CSV/TXT (*.csv *.txt);;{_all} (*)",
        )
        if not ruta:
            return
        self.lbl_resultado.setText(tr("stock.importing", default="Importando…"))
        self.lbl_resultado.setVisible(True)
        self._hilo = _ImportarHilo(ruta)
        self._hilo.finalizado.connect(self._on_finalizado)
        self._hilo.start()

    def _on_finalizado(self, mensaje):
        self.lbl_resultado.setText(mensaje)


class _ExportarStockPage(QWidget):
    def __init__(self, host=None, parent=None):
        super().__init__(parent)
        self._host = host   # MostrarStockWindow (para navegar a Documentos)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(32, 28, 32, 28)
        layout.setSpacing(16)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self._lbl_titulo = lbl_titulo = QLabel(tr("stock.title_export", default="Exportar Stock"))
        lbl_titulo.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        lbl_titulo.setStyleSheet(f"color: {_CIAN};")
        layout.addWidget(lbl_titulo)

        self._lbl_desc = lbl_desc = QLabel(tr("stock.export_desc", default=_EXPORT_DESC_DEFAULT))
        lbl_desc.setStyleSheet(_DESC_SS)
        lbl_desc.setWordWrap(True)
        lbl_desc.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        layout.addWidget(lbl_desc)

        icon_lbl = QLabel("🖨")
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_lbl.setStyleSheet(
            "font-size: 160px; background: transparent; border: none;"
        )
        icon_lbl.setFixedHeight(200)
        layout.addWidget(icon_lbl)

        self._btn = btn = QPushButton(tr("stock.export_btn", default="EXPORTAR STOCK"))
        btn.setStyleSheet(_BTN_CIAN_SS)
        btn.setMinimumSize(160, 62); btn.setMaximumWidth(320)  # responsive (P2): antes fijo 200x62
        btn.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn.clicked.connect(self._exportar)
        _sombra_cian(btn)
        layout.addWidget(btn, alignment=Qt.AlignmentFlag.AlignHCenter)

        # Mensaje + acción tras exportar (centrado en la pestaña).
        self.lbl_resultado = QLabel()
        self.lbl_resultado.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_resultado.setStyleSheet(f"color: {_CIAN}; font-size: 14px; font-weight: bold;")
        self.lbl_resultado.setWordWrap(True)
        self.lbl_resultado.setVisible(False)
        layout.addWidget(self.lbl_resultado)

        self._btn_ver = QPushButton(tr("stock.export_view_btn", default="📂 VER ARCHIVO"))
        self._btn_ver.setStyleSheet(_BTN_CIAN_SS)
        self._btn_ver.setMinimumSize(160, 52); self._btn_ver.setMaximumWidth(300)
        self._btn_ver.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        self._btn_ver.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_ver.clicked.connect(self._abrir_documentos_exportaciones)
        self._btn_ver.setVisible(False)
        layout.addWidget(self._btn_ver, alignment=Qt.AlignmentFlag.AlignHCenter)
        layout.addStretch()

    def _generar_excel(self):
        """Genera el informe Excel de stock y devuelve su ruta (o None si vacío/error,
        mostrando el motivo en pantalla). Reutilizable por export y envío por correo."""
        try:
            # Cliente fino (Fase 3): el DataFrame de stock para exportar viene de la capa de datos.
            from src.db.articulos import df_stock_export
            df = df_stock_export()

            if df.empty:
                self.lbl_resultado.setText(tr("stock.export_empty", default="No hay artículos para exportar."))
                self.lbl_resultado.setVisible(True)
                return None

            # Cabeceras del informe en el idioma activo.
            df.columns = [
                tr("stock.exp_code", default="Código"),
                tr("stock.exp_name", default="Nombre"),
                tr("stock.exp_shelf", default="Stock Lineal"),
                tr("stock.exp_warehouse", default="Stock Almacén"),
                tr("stock.exp_central", default="Stock Almacén Central"),
                tr("stock.exp_expected", default="Stock Esperado"),
            ]

            exports_dir = os.path.abspath(
                os.path.join(os.path.dirname(__file__), "../../documentos/stocks")
            )
            os.makedirs(exports_dir, exist_ok=True)
            ruta = os.path.join(
                exports_dir,
                f"Stock_Completo_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx",
            )

            with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Stock")
                ws = writer.sheets["Stock"]
                try:
                    from openpyxl.styles import Alignment, Font, PatternFill

                    fill = PatternFill("solid", fgColor="00FFC6")
                    font_header = Font(
                        name="Segoe UI", bold=True, color="0E1117", size=11
                    )
                    font_data = Font(name="Segoe UI", size=10)
                    align_center = Alignment(horizontal="center", vertical="center")
                    for cell in ws[1]:
                        cell.fill = fill
                        cell.font = font_header
                        cell.alignment = align_center
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.font = font_data
                            cell.alignment = align_center
                    for col in ws.columns:
                        max_len = max(len(str(c.value or "")) for c in col)
                        ws.column_dimensions[col[0].column_letter].width = max(
                            14, max_len + 4
                        )
                except Exception:
                    pass

            return ruta
        except Exception as e:
            self.lbl_resultado.setText(tr("stock.export_err", default="Error al exportar:\n{e}", e=e))
            self.lbl_resultado.setVisible(True)
            return None

    def _exportar(self):
        # 1) Generar el Excel con el stock actual.
        ruta = self._generar_excel()
        if not ruta:
            return
        # 2) Registrarlo en Documentos → categoría "Exportaciones Excel" (tipo=exportacion).
        registrado = False
        try:
            from src.db.documentos import registrar_documento
            registrado = bool(registrar_documento(
                ruta, tipo="exportacion",
                nombre=os.path.basename(ruta),
                referencia="STOCK",
            ))
        except Exception as e:
            self.lbl_resultado.setStyleSheet("color:#F85149;font-size:14px;font-weight:bold;")
            self.lbl_resultado.setText(tr(
                "stock.export_doc_err",
                default="Excel generado, pero no se pudo registrar en Documentos:\n{e}", e=e))
            self.lbl_resultado.setVisible(True)
            self._btn_ver.setVisible(False)
            return
        # 3) Mensaje centrado + botón "Ver archivo" hacia Exportaciones Excel.
        self.lbl_resultado.setStyleSheet(f"color:{_CIAN};font-size:14px;font-weight:bold;")
        self.lbl_resultado.setText(tr(
            "stock.export_done_msg",
            default=("El stock se ha exportado correctamente.\n"
                     "Puedes encontrar el archivo en la pestaña «Exportaciones Excel» "
                     "de la función Documentos, donde podrás compartirlo o enviarlo.")))
        self.lbl_resultado.setVisible(True)
        self._btn_ver.setVisible(registrado)

    def _abrir_documentos_exportaciones(self):
        """Abre la función Documentos posicionada en la categoría «Exportaciones Excel»."""
        from src.gui.centro_documental import CentroDocumentalWindow
        main = getattr(self._host, "main", None)
        try:
            if main is not None and hasattr(main, "manejar_apertura"):
                main.manejar_apertura(
                    "documentos", CentroDocumentalWindow,
                    callback_vuelta=getattr(main, "mostrar_menu_principal", None),
                    usuario=getattr(self._host, "usuario_actual", None),
                    categoria_inicial="exportacion")
            else:
                self._doc_win = CentroDocumentalWindow(
                    usuario=getattr(self._host, "usuario_actual", None),
                    categoria_inicial="exportacion")
                self._doc_win.showMaximized()
        except Exception as e:
            self.lbl_resultado.setStyleSheet("color:#F85149;font-size:14px;font-weight:bold;")
            self.lbl_resultado.setText(tr("stock.export_nav_err",
                                          default="No se pudo abrir Documentos:\n{e}", e=e))

    def _retraducir(self):
        self._lbl_titulo.setText(tr("stock.title_export", default="Exportar Stock"))
        self._lbl_desc.setText(tr("stock.export_desc", default=_EXPORT_DESC_DEFAULT))
        self._btn.setText(tr("stock.export_btn", default="EXPORTAR STOCK"))
        self._btn_ver.setText(tr("stock.export_view_btn", default="📂 VER ARCHIVO"))


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------


class MostrarStockWindow(QWidget):
    def __init__(
        self, callback_vuelta=None, usuario=None, stock_signals_instance=None, **kwargs
    ):
        super().__init__()

        self.callback_vuelta = callback_vuelta
        self.usuario_actual = usuario
        self.main = kwargs.get("main")   # menú principal (para navegar a otros módulos)
        if isinstance(usuario, dict):
            self.perfil = usuario.get("perfil", "OPERARIO")
        else:
            self.perfil = getattr(usuario, "perfil", "OPERARIO")

        ensure_schema()

        self.signals = stock_signals_instance or stock_signals
        self.signals.stock_actualizado.connect(self._on_stock_actualizado)

        self._setup_ui()
        self.setStyleSheet(f"background-color: {_FONDO}; color: white;")
        i18n.conectar_retraduccion(self, self._retraducir)

        # P3 (UX-TPV-01): sidebar colapsable con persistencia por usuario.
        try:
            from src.gui.sidebar_colapsable import instalar_sidebar_colapsable
            instalar_sidebar_colapsable(self, self.sidebar, usuario=self.usuario_actual, clave="stock")
        except Exception:
            pass

    def _setup_ui(self):
        self.setWindowTitle(tr("stock.smart_stock", default="Smart Stock"))
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ---- Sidebar (uses global QSS objectNames) ----
        sidebar = QFrame()
        self.sidebar = sidebar  # P3: referencia para el toggle colapsable
        sidebar.setObjectName("sidebar_logistica")
        sidebar.setFixedWidth(280)

        side_ly = QVBoxLayout(sidebar)
        side_ly.setContentsMargins(0, 40, 0, 20)
        side_ly.setSpacing(0)

        lbl_titulo = QLabel(tr("stock.smart_stock_2", default="SMART STOCK"))
        lbl_titulo.setObjectName("sidebar_title")
        side_ly.addWidget(lbl_titulo)

        self._tab_keys = [
            "stock.tab_store", "stock.tab_central", "stock.tab_edit",
            "stock.tab_import", "stock.tab_export", "stock.tab_inventory",
            "stock.tab_kardex", "stock.tab_lotes",
        ]
        _tab_def = ["STOCK TIENDA", "STOCK ALMACÉN", "EDITAR STOCK",
                    "IMPORTAR STOCK", "EXPORTAR STOCK", "INVENTARIO", "KÁRDEX", "LOTES"]

        self._nav_btns = []
        for idx, key in enumerate(self._tab_keys):
            btn = _SidebarBtn(tr(key, default=_tab_def[idx]))
            btn.setObjectName("btn_sidebar")
            btn.setCheckable(True)
            btn.setAutoExclusive(True)
            btn.setFixedHeight(55)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            btn.clicked.connect(lambda _, i=idx: self._ir_a(i))
            side_ly.addWidget(btn)
            self._nav_btns.append(btn)

        side_ly.addStretch()

        self._btn_exit = btn_exit = _SidebarBtn(tr("stock.exit", default="SALIR AL MENÚ"))
        btn_exit.setObjectName("btn_sidebar_exit")
        btn_exit.setFixedHeight(55)
        btn_exit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_exit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        btn_exit.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                border-left: 4px solid transparent;
                border-radius: 0px;
                font-size: 12px;
                font-family: 'Segoe UI';
                font-weight: 900;
                text-align: left;
                padding-left: 28px;
                color: #F85149;
            }
            QPushButton:hover {
                background-color: #F85149;
                color: #0E1117;
            }
        """)
        btn_exit.clicked.connect(self.volver_menu_principal)
        side_ly.addWidget(btn_exit)

        root.addWidget(sidebar)

        # ---- Content area ----
        self._vistas = QStackedWidget()
        self._vistas.setObjectName("contenido_logistica")
        self._vistas.setStyleSheet(f"background-color: {_FONDO};")

        self._page_tienda = _StockTiendaPage()
        # Reubicadas (embebidas, misma lógica): Stock por almacén sustituye a "central",
        # Inventario físico sustituye a "inventario", y se añade Kárdex.
        self._page_central = self._embed_window("stock_almacen_gui", "StockAlmacenWindow")
        self._page_editar = _EditarStockPage()
        self._page_importar = _ImportarStockPage()
        self._page_exportar = _ExportarStockPage(host=self)
        self._page_inventario = self._embed_window("inventario_fisico", "InventarioFisicoWindow")
        self._page_kardex = self._embed_window("kardex_visor", "KardexVisorWindow")
        # Lotes y caducidades migrado desde el menú principal (misma lógica, embebida).
        self._page_lotes = self._embed_window("lotes_caducidades", "LotesWindow")

        for page in (
            self._page_tienda,
            self._page_central,
            self._page_editar,
            self._page_importar,
            self._page_exportar,
            self._page_inventario,
            self._page_kardex,
            self._page_lotes,
        ):
            self._vistas.addWidget(page)

        # Responsive P2: el contenido (con ventanas embebidas anchas) se envuelve en un scroll para
        # que la ventana pueda encogerse en pantallas/resoluciones pequeñas (1366px, etc.) sin cortar
        # información — el contenido hace scroll en lugar de forzar un ancho mínimo grande. No altera
        # proporciones ni el diseño Enterprise.
        from PyQt6.QtWidgets import QScrollArea
        try:
            from src.gui.foundation import tokens as _T
            _sb = _T.qss_scrollbar()
        except Exception:
            _sb = ""
        _scroll_cont = QScrollArea()
        _scroll_cont.setWidgetResizable(True)
        _scroll_cont.setFrameShape(QScrollArea.Shape.NoFrame)
        _scroll_cont.setStyleSheet(f"QScrollArea{{background:{_FONDO};border:none;}}" + _sb)
        _scroll_cont.setWidget(self._vistas)
        root.addWidget(_scroll_cont)
        self._ir_a(0)

    def _embed_window(self, modulo, clase):
        """Instancia una ventana existente para EMBEBERLA como pestaña (sin tocar su lógica)."""
        mod = __import__(f"src.gui.{modulo}", fromlist=[clase])
        return getattr(mod, clase)(callback_vuelta=None, usuario=self.usuario_actual, main=self)

    def _retraducir(self):
        _tab_def = ["STOCK TIENDA", "STOCK ALMACÉN", "EDITAR STOCK",
                    "IMPORTAR STOCK", "EXPORTAR STOCK", "INVENTARIO", "KÁRDEX", "LOTES"]
        for i, btn in enumerate(self._nav_btns):
            btn.setText(tr(self._tab_keys[i], default=_tab_def[i]))
        self._btn_exit.setText(tr("stock.exit", default="SALIR AL MENÚ"))
        for page in (
            self._page_tienda, self._page_central, self._page_editar,
            self._page_importar, self._page_exportar, self._page_inventario,
            self._page_kardex, self._page_lotes,
        ):
            if hasattr(page, "_retraducir"):
                page._retraducir()

    def _ir_a(self, index: int):
        self._vistas.setCurrentIndex(index)
        for i, btn in enumerate(self._nav_btns):
            btn.setChecked(i == index)
            repolish_widget(btn)

    def _on_stock_actualizado(self, codigo: str):
        if hasattr(self._page_tienda, "cargar"):
            self._page_tienda.cargar()
        if hasattr(self._page_central, "cargar"):
            self._page_central.cargar()

    def volver_menu_principal(self):
        if self.callback_vuelta:
            self.callback_vuelta()
        self.close()

    # Backward-compat
    def cargar_stock(self):
        self._page_tienda.cargar()
        self._page_central.cargar()

    def actualizar_stock_articulo(self, codigo: str):
        self._on_stock_actualizado(codigo)
