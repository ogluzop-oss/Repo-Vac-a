# src/gui/ventas.py
import logging
import os
from datetime import datetime

from PyQt6.QtCore import QDate, QEvent, QObject, QPoint, QRectF, Qt, QThread, QTimer, pyqtSignal
from PyQt6.QtGui import QBitmap, QColor, QFont, QIcon, QPainter, QPainterPath, QPalette, QPen, QPolygon, QRegion
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QAbstractSpinBox,
    QCalendarWidget,
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QProgressDialog,
    QProxyStyle,
    QPushButton,
    QSpinBox,
    QStackedWidget,
    QStyleFactory,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from assets.estilo_global import mostrar_mensaje
from src.db.conexion import obtener_conexion
from src.utils import divisas, i18n
from src.utils.i18n import tr

logger = logging.getLogger(__name__)


def _meses_i18n(con_vacio=True):
    """Lista de nombres de mes traducidos al idioma activo.
    Si `con_vacio`, el índice 0 es "" (para indexar por nº de mes 1-12)."""
    nombres = [tr("common.mon_%d" % i) for i in range(1, 13)]
    return ([""] + nombres) if con_vacio else nombres

# ── Matplotlib (opcional) ─────────────────────────────────────────────────────
try:
    import matplotlib
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
    from matplotlib.figure import Figure
    matplotlib.use("QtAgg")
    MATPLOTLIB_OK = True
except Exception:
    MATPLOTLIB_OK = False

# ── Estilos globales ──────────────────────────────────────────────────────────
BG       = "#0E1117"
SIDEBAR  = "#111418"
CIAN     = "#00FFC6"
TEXTO    = "#E6EDF3"
BORDE    = "#21262D"
ROJO     = "#F85149"

_SS_BASE = f"""
QWidget {{ background-color: {BG}; color: {TEXTO}; font-family: 'Segoe UI'; }}
QFrame  {{ background-color: {BG}; }}
QLabel  {{ color: {TEXTO}; background: transparent; }}
"""

_SS_SIDEBAR_BTN = f"""
QPushButton {{
    background: transparent;
    color: #8B949E;
    text-align: left;
    padding: 6px 8px 6px 28px;
    border: none;
    border-radius: 0px;
    font-family: 'Segoe UI';
    font-size: 12px;
    font-weight: 900;
    letter-spacing: 0.5px;
    margin: 0px;
}}
QPushButton:hover {{
    background: #FFFFFF;
    color: {SIDEBAR};
}}
"""

_SS_SIDEBAR_BTN_ACTIVE = f"""
QPushButton {{
    background: {CIAN};
    color: {BG};
    text-align: left;
    padding: 6px 8px 6px 28px;
    border: none;
    border-radius: 0px;
    font-family: 'Segoe UI';
    font-size: 12px;
    font-weight: 900;
    letter-spacing: 0.5px;
    margin: 0px;
}}
"""

_SS_NEON_INPUT = f"""
QLineEdit, QDateEdit, QComboBox {{
    background: #161B22;
    color: {TEXTO};
    border: 1px solid {CIAN};
    border-radius: 6px;
    padding: 6px 10px;
    font-family: 'Segoe UI';
    font-size: 11px;
}}
QLineEdit:focus, QDateEdit:focus, QComboBox:focus {{
    border: 1.5px solid {CIAN};
}}
QDateEdit::drop-down {{ border: none; width: 18px; }}
"""

# Stylesheet exclusivo del widget QCalendarWidget (aplicado directamente al cal)
_CAL_SS = f"""
QCalendarWidget {{
    background: #0E1117;
    border: none;
    border-radius: 10px;
    min-width: 318px;
    min-height: 258px;
    padding: 0px 0px 4px 0px;
}}
QCalendarWidget QWidget {{
    background: #0E1117;
    alternate-background-color: #0E1117;
}}
QCalendarWidget QWidget#qt_calendar_navigationbar {{
    background: transparent;
    border: none;
    min-height: 42px;
}}
QCalendarWidget QToolButton {{
    color: {CIAN};
    background: #161B22;
    font-weight: 900;
    font-size: 24px;
    border: 2px solid {CIAN};
    border-radius: 12px;
    padding: 0px 8px 2px 8px;
    min-width: 34px;
    min-height: 34px;
}}
QCalendarWidget QToolButton:hover {{
    background: {CIAN};
    color: {BG};
}}
QCalendarWidget QMenu {{
    background: #161B22;
    color: {TEXTO};
    border: 2px solid {CIAN};
    border-radius: 12px;
}}
QCalendarWidget QSpinBox {{
    min-width: 0px;
    max-width: 0px;
    min-height: 0px;
    max-height: 0px;
    margin: 0px;
    padding: 0px;
    border: none;
    background: transparent;
    color: transparent;
}}
QCalendarWidget QSpinBox::up-button, QCalendarWidget QSpinBox::down-button {{
    width: 0px; height: 0px;
}}
QCalendarWidget QTableView {{
    gridline-color: transparent;
    border: none;
    background: #0E1117;
    border-radius: 10px;
}}
QCalendarWidget QHeaderView::section {{
    background: #0E1117;
    color: {CIAN};
    border: none;
    font-weight: 700;
    font-size: 10px;
    padding: 2px;
}}
QCalendarWidget QAbstractItemView {{
    background: #0E1117;
    color: {TEXTO};
    selection-background-color: {CIAN};
    selection-color: #0E1117;
    outline: none;
    gridline-color: transparent;
    border: none;
}}
QCalendarWidget QAbstractItemView:disabled {{
    color: #4B5563;
}}
"""

_MENU_CAL_SS = f"""
QMenu {{
    background: #161B22;
    color: {TEXTO};
    border: 1px solid {CIAN};
    border-radius: 6px;
    padding: 4px 0px;
}}
QMenu::item {{
    padding: 7px 22px;
    border-radius: 3px;
    margin: 1px 4px;
}}
QMenu::item:selected {{
    background: rgba(0,255,198,0.18);
    color: {CIAN};
}}
"""

_SS_BTN_CIAN = f"""
QPushButton {{
    background: transparent;
    color: {CIAN};
    border: 1.5px solid {CIAN};
    border-radius: 7px;
    padding: 8px 20px;
    font-family: 'Segoe UI';
    font-weight: 900;
    font-size: 11px;
}}
QPushButton:hover {{
    background: {CIAN};
    color: #0E1117;
}}
"""

_SS_BTN_VERDE = """
QPushButton {
    background: #0E1117;
    color: #3FB950;
    border: 2px solid #3FB950;
    border-radius: 8px;
    padding: 8px 20px;
    font-family: 'Segoe UI';
    font-weight: 900;
    font-size: 11px;
}
QPushButton:hover {
    background: #3FB950;
    color: #0E1117;
}
"""

_SS_BTN_EXIT = f"""
QPushButton {{
    background: transparent;
    color: {ROJO};
    border: none;
    border-radius: 0px;
    text-align: left;
    padding: 6px 8px 6px 28px;
    font-family: 'Segoe UI';
    font-weight: 900;
    font-size: 12px;
    letter-spacing: 0.5px;
}}
QPushButton:hover {{
    background: {ROJO};
    color: {BG};
}}
"""

_SS_TABLE = f"""
QTableWidget {{
    background: #161B22;
    color: {TEXTO};
    border: 1.5px solid {CIAN};
    border-radius: 8px;
    gridline-color: {BORDE};
    selection-background-color: rgba(0,255,198,0.18);
    selection-color: {CIAN};
    font-family: 'Segoe UI';
    font-size: 11px;
}}
QHeaderView::section {{
    background: {SIDEBAR};
    color: {CIAN};
    font-weight: 900;
    padding: 10px 8px;
    min-height: 40px;
    border: none;
    border-bottom: 1.5px solid {CIAN};
    font-family: 'Segoe UI';
    font-size: 14px;
    text-align: center;
}}
QHeaderView::section:first {{
    border-top-left-radius: 7px;
}}
QHeaderView::section:last {{
    border-top-right-radius: 7px;
}}
QHeaderView::section:hover {{
    background: {CIAN};
    color: {BG};
}}
QTableWidget::item {{ padding: 6px; }}
"""


# ─────────────────────────────────────────────────────────────────────────────
# Helper classes
# ─────────────────────────────────────────────────────────────────────────────

class _PopupBorderOverlay(QWidget):
    """Capa transparente sobre el popup del calendario para dibujar el borde neón.
    Fallback cuando Qt usa su QCalendarPopup interno en lugar de _NeonCalFrame."""

    def __init__(self, parent):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        self.setAutoFillBackground(False)
        self.setStyleSheet("background: transparent; border: none;")

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        r = QRectF(self.rect()).adjusted(1.5, 1.5, -1.5, -1.5)
        path = QPainterPath()
        path.addRoundedRect(r, 14, 14)
        p.setPen(QPen(QColor(CIAN), 2))
        p.setBrush(Qt.BrushStyle.NoBrush)
        p.drawPath(path)
        p.end()


class _NeonCalFrame(QFrame):
    """Ventana popup con fondo #11181D y borde neón.
    Usa setMask para esquinas redondeadas — más fiable que WA_TranslucentBackground en Windows."""

    def __init__(self):
        super().__init__(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setStyleSheet("background: #11181D; border: none;")

    def showEvent(self, event):
        super().showEvent(event)
        sz = self.size()
        if sz.width() > 0 and sz.height() > 0:
            bmp = QBitmap(sz)
            bmp.fill(Qt.GlobalColor.color0)
            p = QPainter(bmp)
            p.setBrush(Qt.GlobalColor.color1)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(0, 0, sz.width(), sz.height(), 14, 14)
            p.end()
            self.setMask(bmp)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.fillRect(self.rect(), QColor("#11181D"))
        r = QRectF(self.rect()).adjusted(1.5, 1.5, -1.5, -1.5)
        path = QPainterPath()
        path.addRoundedRect(r, 14, 14)
        p.setPen(QPen(QColor(CIAN), 2))
        p.setBrush(Qt.BrushStyle.NoBrush)
        p.drawPath(path)
        p.end()


class _NeonDateEdit(QDateEdit):
    """QDateEdit con popup de calendario propio (borde neón, esquinas redondeadas,
    posicionado bajo el campo en cualquier ventana). Dibuja su propio triángulo
    y abre el calendario al pulsar en su zona derecha."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._cal_popup = None
        # Sin popup nativo ni botones de spin: usamos showPopup() propio.
        self.setCalendarPopup(False)
        self.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)

    def paintEvent(self, event):
        super().paintEvent(event)
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        cx, cy = self.width() - 14, self.height() // 2
        tri = QPolygon([QPoint(cx - 5, cy - 3), QPoint(cx + 5, cy - 3), QPoint(cx, cy + 4)])
        p.setBrush(QColor(CIAN)); p.setPen(Qt.PenStyle.NoPen)
        p.drawPolygon(tri); p.end()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton and event.position().x() >= self.width() - 30:
            self.showPopup(); event.accept(); return
        super().mousePressEvent(event)

    def showPopup(self):
        if self._cal_popup is not None:
            try:
                if self._cal_popup.isVisible():
                    return
            except RuntimeError:
                pass
            self._cal_popup = None

        popup = _NeonCalFrame()
        popup.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)

        # 11px = 2px border + 9px inner gap; calendar never overlaps the border
        MARGIN = 11
        lay = QVBoxLayout(popup)
        lay.setContentsMargins(MARGIN, MARGIN, MARGIN, MARGIN)
        lay.setSpacing(0)

        cal = _VentasCalendarWidget(popup)
        cal.setSelectedDate(self.date())
        lay.addWidget(cal)

        def _on_date_clicked(qdate):
            self.setDate(qdate)
            try:
                popup.close()
            except RuntimeError:
                pass
            self._cal_popup = None

        cal.clicked.connect(_on_date_clicked)

        popup.setFixedSize(
            cal.minimumWidth() + 2 * MARGIN,
            cal.minimumHeight() + 2 * MARGIN,
        )
        popup.move(self.mapToGlobal(QPoint(0, self.height())))
        popup.show()
        popup.raise_()
        self._cal_popup = popup

        def _retry_nav(c=cal, retries=8):
            try:
                if c._ensure_custom_nav():
                    return
            except RuntimeError:
                return
            except Exception:
                pass
            if retries > 0:
                QTimer.singleShot(30, lambda: _retry_nav(c, retries - 1))

        QTimer.singleShot(0, _retry_nav)

    def hidePopup(self):
        if self._cal_popup is not None:
            try:
                self._cal_popup.close()
            except RuntimeError:
                pass
            self._cal_popup = None
        else:
            super().hidePopup()


class _VentasCalendarWidget(QCalendarWidget):
    """Calendario propio de ventas con navegación controlada por la app."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._month_btn = None
        self._year_btn = None
        self._month_popup = None
        self._year_popup = None
        self._nav_ready = False
        self.setGridVisible(False)
        self.setVerticalHeaderFormat(
            QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader
        )
        self.setHorizontalHeaderFormat(
            QCalendarWidget.HorizontalHeaderFormat.ShortDayNames
        )
        self.setStyleSheet(
            f"""
            QCalendarWidget {{
                background: #11181D;
                border: none;
            }}
            QCalendarWidget QWidget {{
                background: #11181D;
                alternate-background-color: #11181D;
                border: none;
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background: #11181D;
                border: none;
                min-height: 42px;
            }}
            QCalendarWidget QTableView {{
                background: #11181D;
                border: none;
                outline: none;
                gridline-color: transparent;
                selection-background-color: {CIAN};
                selection-color: {BG};
            }}
            QCalendarWidget QHeaderView::section {{
                background: #11181D;
                color: {CIAN};
                border: none;
                padding: 2px;
                font-family: 'Segoe UI';
                font-size: 10px;
                font-weight: 900;
            }}
            QCalendarWidget QAbstractItemView {{
                background: #11181D;
                color: {TEXTO};
                border: none;
                outline: none;
                selection-background-color: {CIAN};
                selection-color: {BG};
                font-family: 'Segoe UI';
                font-size: 11px;
                font-weight: 900;
            }}
            QCalendarWidget QAbstractItemView:disabled {{
                color: #4B5563;
            }}
            QCalendarWidget QToolButton#qt_calendar_yearbutton,
            QCalendarWidget QSpinBox#qt_calendar_yearedit,
            QCalendarWidget QWidget#qt_calendar_yearselector {{
                max-width: 0px; max-height: 0px;
                min-width: 0px; min-height: 0px;
                border: none; background: transparent;
                color: transparent; padding: 0px; margin: 0px;
            }}
            """
        )
        self.setMinimumSize(318, 258)
        self.currentPageChanged.connect(lambda _y, _m: self._sync_nav_texts())

    def showEvent(self, event):
        super().showEvent(event)
        self._ensure_custom_nav()
        QTimer.singleShot(0, self._style_popup)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._ensure_custom_nav()

    def _style_popup(self):
        """Aplica máscara redondeada y borde neón al popup que contiene el calendario.
        Maneja tanto _NeonCalFrame como el QCalendarPopup interno de Qt."""
        try:
            popup = self.window()
            if isinstance(popup, _NeonCalFrame) or popup is self:
                return
            sz = popup.size()
            if sz.width() <= 0 or sz.height() <= 0:
                return
            bmp = QBitmap(sz)
            bmp.fill(Qt.GlobalColor.color0)
            pa = QPainter(bmp)
            pa.setBrush(Qt.GlobalColor.color1)
            pa.setPen(Qt.PenStyle.NoPen)
            pa.drawRoundedRect(0, 0, sz.width(), sz.height(), 14, 14)
            pa.end()
            popup.setMask(bmp)
            overlay = getattr(self, '_popup_overlay', None)
            if overlay is None or not isinstance(overlay, _PopupBorderOverlay) or overlay.parent() is not popup:
                overlay = _PopupBorderOverlay(popup)
                overlay.setGeometry(popup.rect())
                self._popup_overlay = overlay
                overlay.show()
            overlay.raise_()
        except Exception:
            pass

    def _ensure_custom_nav(self):
        if self._nav_ready:
            return True
        nav = self.findChild(QWidget, "qt_calendar_navigationbar")
        prev_btn = self.findChild(QToolButton, "qt_calendar_prevmonth")
        next_btn = self.findChild(QToolButton, "qt_calendar_nextmonth")
        month_btn_orig = self.findChild(QToolButton, "qt_calendar_monthbutton")
        year_spin = self.findChild(QSpinBox, "qt_calendar_yearedit")
        if not nav or not prev_btn or not next_btn or not month_btn_orig:
            return False

        arrow_ss = f"""
            QToolButton {{
                color: {CIAN};
                background: transparent;
                font-family: 'Segoe UI';
                font-size: 20px;
                font-weight: 900;
                border: none;
                border-radius: 8px;
                padding: 0px 8px 4px 8px;
                min-width: 30px;
                min-height: 34px;
            }}
            QToolButton:hover {{
                background: rgba(0,255,198,0.15);
                color: {CIAN};
            }}
        """
        prev_btn.setText("←")
        next_btn.setText("→")
        for btn in (prev_btn, next_btn):
            btn.setIcon(QIcon())
            btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
            btn.setStyleSheet(arrow_ss)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)

        nav_ss = f"""
            QToolButton {{
                color: {CIAN};
                background: transparent;
                font-family: 'Segoe UI';
                font-size: 13px;
                font-weight: 900;
                border: none;
                border-radius: 8px;
                padding: 4px 10px;
                min-height: 34px;
            }}
            QToolButton:hover {{
                background: rgba(0,255,198,0.15);
                color: {CIAN};
            }}
            """

        # Ocultar widgets nativos
        month_btn_orig.setMaximumSize(0, 0)
        month_btn_orig.hide()
        if year_spin:
            year_spin.setMaximumSize(0, 0)
            year_spin.hide()

        # Centrar: prev ya esta en el layout nativo, insertar stretch + month + year + stretch
        # antes del next_btn que Qt pone al final
        layout = nav.layout()
        # Encontrar posicion de prev y next en el layout
        prev_idx = layout.indexOf(prev_btn)
        next_idx = layout.indexOf(next_btn)

        if self._month_btn is None:
            self._month_btn = QToolButton(nav)
            self._month_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            self._month_btn.setStyleSheet(nav_ss)
            self._month_btn.clicked.connect(self._open_month_popup)

        if self._year_btn is None:
            self._year_btn = QToolButton(nav)
            self._year_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            self._year_btn.setStyleSheet(nav_ss)
            self._year_btn.clicked.connect(self._open_year_popup)

        # Insertar: stretch, month, year, stretch entre prev y next
        insert_at = prev_idx + 1 if prev_idx >= 0 else 1
        layout.insertStretch(insert_at, 1)
        layout.insertWidget(insert_at + 1, self._month_btn)
        layout.insertWidget(insert_at + 2, self._year_btn)
        layout.insertStretch(insert_at + 3, 1)

        self._sync_nav_texts()
        self._nav_ready = True
        return True

    def _sync_nav_texts(self):
        meses = _meses_i18n()
        if self._month_btn is not None:
            self._month_btn.setText(f"{meses[self.monthShown()]} ▼")
        if self._year_btn is not None:
            self._year_btn.setText(f"{self.yearShown()} ▼")

    def _close_popup(self, attr_name):
        popup = getattr(self, attr_name, None)
        if popup is not None:
            try:
                popup.close()
            except Exception:
                pass
            setattr(self, attr_name, None)

    def _open_month_popup(self):
        self._close_popup("_year_popup")
        self._close_popup("_month_popup")
        if self._month_btn:
            self._month_btn.setAttribute(Qt.WidgetAttribute.WA_UnderMouse, False)
            self._month_btn.setDown(False)
            self._month_btn.repaint()

        popup = QFrame(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        popup.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        popup.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        popup.setStyleSheet("background: transparent; border: none;")

        inner = QFrame(popup)
        inner.setStyleSheet(
            f"""
            QFrame {{
                background: #161B22;
                border: 2px solid {CIAN};
                border-radius: 14px;
            }}
            QPushButton {{
                background: #161B22;
                color: {TEXTO};
                border: none;
                border-radius: 10px;
                padding: 8px 6px;
                font-family: 'Segoe UI';
                font-size: 13px;
                font-weight: 900;
                min-width: 84px;
                min-height: 34px;
            }}
            QPushButton:hover {{
                background: {CIAN};
                color: {BG};
            }}
            """
        )

        grid = QGridLayout(inner)
        grid.setContentsMargins(8, 8, 8, 8)
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(6)

        meses = _meses_i18n(con_vacio=False)
        for i, mes in enumerate(meses):
            row, col = divmod(i, 4)
            btn = QPushButton(mes)
            if self.monthShown() == i + 1:
                btn.setStyleSheet(
                    f"background: {CIAN}; color: {BG}; border: none; font-weight: 900;"
                )
            month_number = i + 1
            btn.clicked.connect(
                lambda _checked=False, mo=month_number: (
                    self.setCurrentPage(self.yearShown(), mo),
                    self._close_popup("_month_popup"),
                )
            )
            grid.addWidget(btn, row, col)

        outer = QVBoxLayout(popup)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)
        outer.addWidget(inner)
        popup.setFixedSize(inner.sizeHint().width(), inner.sizeHint().height())
        popup.move(self._month_btn.mapToGlobal(self._month_btn.rect().bottomLeft()))
        popup.show()
        popup.raise_()
        self._month_popup = popup

    def _open_year_popup(self):
        self._close_popup("_month_popup")
        self._close_popup("_year_popup")
        if self._year_btn:
            self._year_btn.setAttribute(Qt.WidgetAttribute.WA_UnderMouse, False)
            self._year_btn.setDown(False)
            self._year_btn.repaint()

        popup = QFrame(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        popup.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        popup.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        popup.setStyleSheet("background: transparent; border: none;")

        inner = QFrame(popup)
        inner.setStyleSheet(
            f"""
            QFrame {{
                background: #161B22;
                border: 2px solid {CIAN};
                border-radius: 12px;
            }}
            QListWidget {{
                background: #161B22;
                border: none;
                outline: none;
                font-family: 'Segoe UI';
                font-size: 12px;
                font-weight: 900;
            }}
            QListWidget::item {{
                color: {TEXTO};
                padding: 7px 10px;
                border-radius: 8px;
                min-height: 24px;
            }}
            QListWidget::item:hover {{
                background: {CIAN};
                color: {BG};
            }}
            QListWidget::item:selected {{
                background: {CIAN};
                color: {BG};
            }}
            """
        )

        list_w = QListWidget(inner)
        list_w.setUniformItemSizes(True)
        list_w.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        list_w.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        list_w.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        current_year = QDate.currentDate().year()
        years = list(range(current_year, 1924, -1))
        shown_year = self.yearShown()
        scroll_to_idx = 0
        bold_font = QFont("Segoe UI", 11, QFont.Weight.Bold)
        norm_font = QFont("Segoe UI", 11)

        for i, year in enumerate(years):
            item = QListWidgetItem(str(year))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if year == shown_year:
                item.setForeground(QColor(CIAN))
                item.setFont(bold_font)
                scroll_to_idx = i
            else:
                item.setFont(norm_font)
            list_w.addItem(item)

        list_w.itemClicked.connect(
            lambda item: (
                self.setCurrentPage(int(item.text()), self.monthShown()),
                self._close_popup("_year_popup"),
            )
        )
        list_w.setFixedWidth(112)
        list_w.setFixedHeight(7 * 34)

        inner_ly = QVBoxLayout(inner)
        inner_ly.setContentsMargins(6, 6, 6, 6)
        inner_ly.addWidget(list_w)

        outer_ly = QVBoxLayout(popup)
        outer_ly.setContentsMargins(0, 0, 0, 0)
        outer_ly.setSpacing(0)
        outer_ly.addWidget(inner)
        popup.setFixedSize(inner.sizeHint().width(), inner.sizeHint().height())
        popup.move(self._year_btn.mapToGlobal(self._year_btn.rect().bottomLeft()))
        popup.show()
        popup.raise_()
        self._year_popup = popup

        from PyQt6.QtCore import QTimer

        def _center_current():
            list_w.scrollToItem(
                list_w.item(scroll_to_idx),
                QAbstractItemView.ScrollHint.PositionAtCenter,
            )

        QTimer.singleShot(0, _center_current)


class _NoArrowComboStyle(QProxyStyle):
    """Fusion-based style that collapses the drop-down indicator to zero width."""
    def __init__(self):
        super().__init__(QStyleFactory.create("Fusion"))

    def pixelMetric(self, metric, opt=None, widget=None):
        from PyQt6.QtWidgets import QStyle
        if metric == QStyle.PixelMetric.PM_MenuButtonIndicator:
            return 0
        return super().pixelMetric(metric, opt, widget)

    def subControlRect(self, cc, opt, sc, widget=None):
        from PyQt6.QtCore import QRect
        from PyQt6.QtWidgets import QStyle
        if (cc == QStyle.ComplexControl.CC_ComboBox and
                sc == QStyle.SubControl.SC_ComboBoxArrow):
            return QRect()
        return super().subControlRect(cc, opt, sc, widget)

    def drawPrimitive(self, elem, opt, p, widget=None):
        from PyQt6.QtWidgets import QStyle
        if elem in (QStyle.PrimitiveElement.PE_IndicatorArrowDown,
                    QStyle.PrimitiveElement.PE_IndicatorButtonDropDown,
                    QStyle.PrimitiveElement.PE_FrameFocusRect):
            return
        super().drawPrimitive(elem, opt, p, widget)


_no_arrow_style: '_NoArrowComboStyle | None' = None


class _NeonComboBox(QComboBox):
    """QComboBox that pixel-paints over the platform drop-down indicator area."""
    _COVER_W = 26

    def paintEvent(self, event):
        super().paintEvent(event)
        p = QPainter(self)
        p.setPen(Qt.PenStyle.NoPen)
        # Clip to inner rounded rect so we don't paint over the neon border corners
        clip = QPainterPath()
        clip.addRoundedRect(QRectF(1, 1, self.width() - 2, self.height() - 2), 5, 5)
        p.setClipPath(clip)
        p.fillRect(
            self.width() - self._COVER_W, 1,
            self._COVER_W - 2, self.height() - 2,
            QColor("#161B22"),
        )
        p.end()


def _get_no_arrow_style() -> '_NoArrowComboStyle':
    global _no_arrow_style
    if _no_arrow_style is None:
        _no_arrow_style = _NoArrowComboStyle()
    return _no_arrow_style


class _ClipTableTopCorners(QObject):
    """Clips a QTableWidget's header top corners via bitmap mask on the header view."""
    def __init__(self, table, radius=7):
        super().__init__(table)
        self._r = radius
        table.horizontalHeader().installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() in (QEvent.Type.Resize, QEvent.Type.Show):
            from PyQt6.QtCore import QRect
            bmp = QBitmap(obj.size())
            bmp.fill(Qt.GlobalColor.color0)
            p = QPainter(bmp)
            p.setBrush(Qt.GlobalColor.color1)
            p.setPen(Qt.PenStyle.NoPen)
            # Extend rect below header so its bottom edge is never clipped
            extended = QRect(0, 0, obj.width(), obj.height() + self._r)
            p.drawRoundedRect(extended, self._r, self._r)
            p.end()
            obj.setMask(QRegion(bmp))
        return False


class _RoundTableCorners(QObject):
    """Redondea las 4 esquinas exteriores de un QTableWidget con una máscara, de
    forma que el contorno neón no se corte en ninguna esquina (ni la cabecera ni
    el cuerpo/scroll)."""
    def __init__(self, table, radius=8):
        super().__init__(table)
        self._r = radius
        table.installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() in (QEvent.Type.Resize, QEvent.Type.Show):
            from PyQt6.QtCore import QRect
            bmp = QBitmap(obj.size())
            bmp.fill(Qt.GlobalColor.color0)
            p = QPainter(bmp)
            p.setBrush(Qt.GlobalColor.color1)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(QRect(0, 0, obj.width(), obj.height()), self._r, self._r)
            p.end()
            obj.setMask(QRegion(bmp))
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _lbl(texto, bold=False, size=11, color=TEXTO):
    lbl = QLabel(texto)
    lbl.setStyleSheet(
        f"color: {color}; font-family: 'Segoe UI'; "
        f"font-size: {size}px; font-weight: {'700' if bold else '400'}; "
        f"background: transparent; border: none;"
    )
    return lbl


def _filter_lbl(texto):
    """Bold label at 13px for filter row headers."""
    return _lbl(texto, bold=True, size=13)


def _input_neon(placeholder=""):
    le = QLineEdit()
    le.setPlaceholderText(placeholder)
    le.setStyleSheet(_SS_NEON_INPUT)
    le.setFixedHeight(34)
    return le


def _show_month_grid(month_btn, cal):
    """Show a 4×3 month-picker popup anchored below the month button."""
    from PyQt6.QtWidgets import QFrame, QGridLayout, QPushButton, QVBoxLayout
    MESES = _meses_i18n(con_vacio=False)

    popup = QFrame(
        None,
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    popup.setStyleSheet(f"background: {BG};")

    inner = QFrame(popup)
    inner.setStyleSheet(f"""
        QFrame {{
            background: #161B22;
            border: 2px solid {CIAN};
            border-radius: 12px;
        }}
        QPushButton {{
            background: #161B22;
                color: {TEXTO};
                border: none;
                border-radius: 10px;
                padding: 8px 6px;
                font-family: 'Segoe UI';
                font-size: 13px;
                font-weight: 900;
                min-width: 84px;
                min-height: 34px;
        }}
        QPushButton:hover {{
            background: {CIAN};
            color: {BG};
        }}
    """)
    grid = QGridLayout(inner)
    grid.setContentsMargins(8, 8, 8, 8)
    grid.setHorizontalSpacing(6)
    grid.setVerticalSpacing(6)
    for i, mes in enumerate(MESES):
        row, col = i // 4, i % 4
        btn_m = QPushButton(mes)
        if cal.monthShown() == i + 1:
            btn_m.setStyleSheet(
                f"background: {CIAN}; color: {BG}; border: none; font-weight: 900;"
            )
        m = i + 1
        def _on_m(checked=False, mo=m, p=popup):
            cal.setCurrentPage(cal.yearShown(), mo)
            p.close()
        btn_m.clicked.connect(_on_m)
        grid.addWidget(btn_m, row, col)

    outer = QVBoxLayout(popup)
    outer.setContentsMargins(4, 4, 4, 4)
    outer.addWidget(inner)
    popup.setFixedSize(
        inner.sizeHint().width() + 8,
        inner.sizeHint().height() + 8,
    )

    pos = month_btn.mapToGlobal(month_btn.rect().bottomLeft())
    popup.move(pos)
    popup.show()


def _show_year_popup(year_btn, cal):
    """Scrollable year picker: current year down to 1925, 7 visible, QListWidget for efficiency."""
    from PyQt6.QtCore import QTimer
    from PyQt6.QtGui import QFont
    from PyQt6.QtWidgets import QAbstractItemView, QFrame, QListWidget, QListWidgetItem, QVBoxLayout

    cy = QDate.currentDate().year()
    years = list(range(cy, 1924, -1))
    cur_yr = cal.yearShown()

    popup = QFrame(
        None,
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    popup.setStyleSheet(f"background: {BG};")

    inner = QFrame(popup)
    inner.setStyleSheet(f"""
        QFrame {{
            background: #161B22;
            border: 2px solid {CIAN};
            border-radius: 12px;
        }}
        QListWidget {{
            background: #161B22;
            border: none;
            outline: none;
            font-family: 'Segoe UI';
            font-size: 12px;
            font-weight: 900;
        }}
        QListWidget::item {{
            color: {TEXTO};
            padding: 7px 10px;
            border-radius: 8px;
            min-height: 24px;
        }}
        QListWidget::item:hover {{
            background: {CIAN};
            color: {BG};
        }}
        QListWidget::item:selected {{
            background: {CIAN};
            color: {BG};
        }}
    """)

    list_w = QListWidget(inner)
    list_w.setUniformItemSizes(True)
    list_w.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
    list_w.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
    list_w.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

    bold_font = QFont("Segoe UI", 11, QFont.Weight.Bold)
    norm_font = QFont("Segoe UI", 11)

    scroll_to_idx = 0
    for i, y in enumerate(years):
        item = QListWidgetItem(str(y))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        if y == cur_yr:
            item.setForeground(QColor(CIAN))
            item.setFont(bold_font)
            scroll_to_idx = i
        else:
            item.setFont(norm_font)
        list_w.addItem(item)

    ITEM_H = 34
    list_w.setFixedWidth(110)
    list_w.setFixedHeight(7 * ITEM_H)

    inner_ly = QVBoxLayout(inner)
    inner_ly.setContentsMargins(6, 6, 6, 6)
    inner_ly.addWidget(list_w)

    outer_ly = QVBoxLayout(popup)
    outer_ly.setContentsMargins(4, 4, 4, 4)
    outer_ly.addWidget(inner)
    popup.setFixedSize(
        inner.sizeHint().width() + 8,
        inner.sizeHint().height() + 8,
    )

    def _on_item_click(item):
        yr = int(item.text())
        cal.setCurrentPage(yr, cal.monthShown())
        popup.close()

    list_w.itemClicked.connect(_on_item_click)

    pos = year_btn.mapToGlobal(year_btn.rect().bottomLeft())
    popup.move(pos)
    popup.show()

    def _center_current():
        list_w.scrollToItem(
            list_w.item(scroll_to_idx),
            QAbstractItemView.ScrollHint.PositionAtCenter,
        )

    QTimer.singleShot(0, _center_current)


def _do_calendar_nav(cal):
    """Customise calendar nav bar — called once on first popup show."""
    from PyQt6.QtGui import QIcon
    from PyQt6.QtWidgets import QSpinBox, QToolButton

    cal.setGridVisible(False)

    # ── Neon prev / next arrows ───────────────────────────────────────────────
    _ARROW_SS = f"""
        QToolButton {{
            color: {CIAN};
            background: transparent;
            font-weight: 900;
            font-size: 24px;
            border: none;
            padding: 0px 8px 2px 8px;
            min-width: 30px;
            min-height: 34px;
        }}
        QToolButton:hover {{
            color: #7CFFE1;
        }}
    """
    for name, sym in [("qt_calendar_prevmonth", "←"), ("qt_calendar_nextmonth", "→")]:
        btn = cal.findChild(QToolButton, name)
        if btn:
            btn.setIcon(QIcon())
            btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
            btn.setText("‹" if "prev" in name else "›")
            btn.setStyleSheet(_ARROW_SS)

    # ── Locate nav bar layout ─────────────────────────────────────────────────
    month_btn_orig = cal.findChild(QToolButton, "qt_calendar_monthbutton")
    year_spin      = cal.findChild(QSpinBox,    "qt_calendar_yearedit")
    nav = (year_spin.parentWidget()       if year_spin       else
           month_btn_orig.parentWidget()  if month_btn_orig  else None)
    lay = nav.layout() if nav else None
    if not lay:
        return

    # Locate positions using objectName — PyQt6 wrappers may differ by identity
    month_idx = year_idx = -1
    for i in range(lay.count()):
        item = lay.itemAt(i)
        w = item.widget() if item else None
        if not w:
            continue
        n = w.objectName()
        if n == "qt_calendar_monthbutton":
            month_idx = i
        elif n == "qt_calendar_yearedit":
            year_idx = i

    _NAV_SS = f"""
        QToolButton {{
            color: {CIAN};
            background: transparent;
            font-weight: 900;
            font-size: 13px;
            border: none;
            padding: 2px 8px;
            min-height: 28px;
        }}
        QToolButton:hover {{
            background: rgba(0,255,198,0.18);
            border-radius: 4px;
        }}
        QToolButton::menu-indicator {{ image: none; width: 0px; }}
    """

    # ── Replace original month button with our own ────────────────────────────
    # Qt's month button uses InstantPopup which cannot be reliably overridden.
    # We hide the original and insert a fresh QToolButton that emits clicked normally.
    our_month = QToolButton(nav)
    our_month.setCursor(Qt.CursorShape.PointingHandCursor)
    our_month.setStyleSheet(_NAV_SS)

    _MESES = _meses_i18n()

    def _sync_month():
        # Qt still updates the hidden original button's text with the locale name.
        raw = month_btn_orig.text() if month_btn_orig else ""
        if not raw:
            m = cal.monthShown()
            raw = _MESES[m] if 1 <= m <= 12 else "?"
        our_month.setText((raw[0].upper() + raw[1:] if raw else "?") + " ▼")

    cal.currentPageChanged.connect(lambda y, m: _sync_month())
    _sync_month()
    our_month.clicked.connect(lambda: _show_month_grid(our_month, cal))

    # Insert our button first (original shifts right), then collapse it
    if month_idx >= 0:
        lay.insertWidget(month_idx, our_month)
    else:
        lay.addWidget(our_month)
    our_month.show()

    # Collapse original in-place — Qt still updates its text internally
    if month_btn_orig:
        month_btn_orig.setMaximumSize(0, 0)
        month_btn_orig.hide()

    # ── Year button ───────────────────────────────────────────────────────────
    year_btn = QToolButton(nav)
    year_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    year_btn.setStyleSheet(_NAV_SS)

    def _update_year():
        year_btn.setText(f"{cal.yearShown()} ▼")

    cal.currentPageChanged.connect(lambda y, m: _update_year())
    _update_year()

    if year_idx >= 0:
        lay.insertWidget(year_idx, year_btn)
    else:
        lay.insertWidget(max(0, lay.count() - 1), year_btn)
    year_btn.show()

    def _open_year():
        try:
            _show_year_popup(year_btn, cal)
        except Exception as e:
            logger.warning(f"Year popup error: {e}", exc_info=True)

    year_btn.clicked.connect(_open_year)


def _date_neon(val: QDate = None):
    # _NeonDateEdit usa su propio popup (showPopup) → no se activa el nativo.
    de = _NeonDateEdit(val or QDate.currentDate())
    de.setDisplayFormat("dd/MM/yyyy")
    de.setStyleSheet(_SS_NEON_INPUT)
    de.setFixedHeight(34)
    return de


# Override fino del calendario de ventas:
# un único contorno exterior, selector de meses 4x3 y selector de años vertical.
_CAL_SS = f"""
QCalendarWidget {{
    background: #0E1117;
    border: none;
    border-radius: 10px;
    min-width: 318px;
    min-height: 258px;
    padding: 0px 0px 4px 0px;
}}
QCalendarWidget QWidget {{
    background: #0E1117;
    alternate-background-color: #0E1117;
}}
QCalendarWidget QWidget#qt_calendar_navigationbar {{
    background: transparent;
    border: none;
    min-height: 42px;
}}
QCalendarWidget QToolButton {{
    background: transparent;
    color: {CIAN};
    border: none;
    font-family: 'Segoe UI';
    font-weight: 900;
}}
QCalendarWidget QMenu {{
    background: #161B22;
    color: {TEXTO};
    border: 2px solid {CIAN};
    border-radius: 12px;
}}
QCalendarWidget QSpinBox {{
    min-width: 0px;
    max-width: 0px;
    min-height: 0px;
    max-height: 0px;
    margin: 0px;
    padding: 0px;
    border: none;
    background: transparent;
    color: transparent;
}}
QCalendarWidget QSpinBox::up-button, QCalendarWidget QSpinBox::down-button {{
    width: 0px; height: 0px;
}}
QCalendarWidget QTableView {{
    background: #0E1117;
    border: none;
    border-radius: 10px;
    gridline-color: transparent;
    outline: none;
}}
QCalendarWidget QHeaderView::section {{
    background: #0E1117;
    color: {CIAN};
    border: none;
    font-weight: 700;
    font-size: 10px;
    padding: 2px;
}}
QCalendarWidget QAbstractItemView {{
    background: #0E1117;
    color: {TEXTO};
    selection-background-color: {CIAN};
    selection-color: {BG};
    outline: none;
    gridline-color: transparent;
    border: none;
}}
QCalendarWidget QAbstractItemView:disabled {{
    color: #4B5563;
}}
"""


def _show_month_grid(month_btn, cal):
    from PyQt6.QtWidgets import QFrame, QGridLayout, QPushButton, QVBoxLayout

    meses = [m.lower() for m in _meses_i18n(con_vacio=False)]

    popup = QFrame(
        None,
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    popup.setStyleSheet(f"background: {BG};")

    inner = QFrame(popup)
    inner.setStyleSheet(
        f"""
        QFrame {{
            background: #161B22;
            border: 2px solid {CIAN};
            border-radius: 12px;
        }}
        QPushButton {{
            background: #161B22;
            color: {TEXTO};
            border: none;
            border-radius: 10px;
            padding: 8px 6px;
            font-family: 'Segoe UI';
            font-size: 11px;
            font-weight: 900;
            min-width: 84px;
            min-height: 34px;
        }}
        QPushButton:hover {{
            background: {CIAN};
            color: {BG};
        }}
        """
    )

    grid = QGridLayout(inner)
    grid.setContentsMargins(8, 8, 8, 8)
    grid.setHorizontalSpacing(6)
    grid.setVerticalSpacing(6)

    for i, mes in enumerate(meses):
        row, col = i // 4, i % 4
        btn = QPushButton(mes)
        if cal.monthShown() == i + 1:
            btn.setStyleSheet(
                f"background: {CIAN}; color: {BG}; border: none; font-weight: 900;"
            )
        mes_num = i + 1
        btn.clicked.connect(
            lambda _checked=False, mo=mes_num, p=popup: (
                cal.setCurrentPage(cal.yearShown(), mo),
                p.close(),
            )
        )
        grid.addWidget(btn, row, col)

    outer = QVBoxLayout(popup)
    outer.setContentsMargins(4, 4, 4, 4)
    outer.addWidget(inner)
    popup.setFixedSize(inner.sizeHint().width() + 8, inner.sizeHint().height() + 8)

    pos = month_btn.mapToGlobal(month_btn.rect().bottomLeft())
    popup.move(pos)
    popup.show()


def _show_year_popup(year_btn, cal):
    from PyQt6.QtCore import QTimer
    from PyQt6.QtGui import QFont
    from PyQt6.QtWidgets import (
        QAbstractItemView,
        QFrame,
        QListWidget,
        QListWidgetItem,
        QVBoxLayout,
    )

    current_year = QDate.currentDate().year()
    years = list(range(current_year, 1924, -1))
    shown_year = cal.yearShown()

    popup = QFrame(
        None,
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    popup.setStyleSheet(f"background: {BG};")

    inner = QFrame(popup)
    inner.setStyleSheet(
        f"""
        QFrame {{
            background: #161B22;
            border: 2px solid {CIAN};
            border-radius: 12px;
        }}
        QListWidget {{
            background: #161B22;
            border: none;
            outline: none;
            font-family: 'Segoe UI';
            font-size: 12px;
            font-weight: 900;
        }}
        QListWidget::item {{
            color: {TEXTO};
            padding: 7px 10px;
            border-radius: 8px;
            min-height: 24px;
        }}
        QListWidget::item:hover {{
            background: {CIAN};
            color: {BG};
        }}
        QListWidget::item:selected {{
            background: {CIAN};
            color: {BG};
        }}
        """
    )

    list_w = QListWidget(inner)
    list_w.setUniformItemSizes(True)
    list_w.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
    list_w.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
    list_w.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

    bold_font = QFont("Segoe UI", 11, QFont.Weight.Bold)
    norm_font = QFont("Segoe UI", 11)
    scroll_to_idx = 0

    for i, year in enumerate(years):
        item = QListWidgetItem(str(year))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        if year == shown_year:
            item.setForeground(QColor(CIAN))
            item.setFont(bold_font)
            scroll_to_idx = i
        else:
            item.setFont(norm_font)
        list_w.addItem(item)

    list_w.setFixedWidth(112)
    list_w.setFixedHeight(7 * 34)

    inner_ly = QVBoxLayout(inner)
    inner_ly.setContentsMargins(6, 6, 6, 6)
    inner_ly.addWidget(list_w)

    outer_ly = QVBoxLayout(popup)
    outer_ly.setContentsMargins(4, 4, 4, 4)
    outer_ly.addWidget(inner)
    popup.setFixedSize(inner.sizeHint().width() + 8, inner.sizeHint().height() + 8)

    list_w.itemClicked.connect(
        lambda item: (cal.setCurrentPage(int(item.text()), cal.monthShown()), popup.close())
    )

    pos = year_btn.mapToGlobal(year_btn.rect().bottomLeft())
    popup.move(pos)
    popup.show()

    def _center_current():
        list_w.scrollToItem(
            list_w.item(scroll_to_idx),
            QAbstractItemView.ScrollHint.PositionAtCenter,
        )

    QTimer.singleShot(0, _center_current)


def _do_calendar_nav(cal):
    from PyQt6.QtGui import QIcon
    from PyQt6.QtWidgets import QSpinBox, QToolButton

    cal.setGridVisible(False)

    arrow_ss = f"""
        QToolButton {{
            color: {CIAN};
            background: transparent;
            font-weight: 900;
            font-size: 24px;
            border: none;
            padding: 0px 8px 2px 8px;
            min-width: 30px;
            min-height: 34px;
        }}
        QToolButton:hover {{
            color: #7CFFE1;
        }}
    """

    for name in ("qt_calendar_prevmonth", "qt_calendar_nextmonth"):
        btn = cal.findChild(QToolButton, name)
        if btn:
            btn.setIcon(QIcon())
            btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
            btn.setText("\u27F5" if "prev" in name else "\u27F6")
            btn.setStyleSheet(arrow_ss)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)

    month_btn_orig = cal.findChild(QToolButton, "qt_calendar_monthbutton")
    year_spin = cal.findChild(QSpinBox, "qt_calendar_yearedit")
    nav = (
        year_spin.parentWidget()
        if year_spin
        else month_btn_orig.parentWidget()
        if month_btn_orig
        else None
    )
    lay = nav.layout() if nav else None
    if not lay:
        return

    month_idx = year_idx = -1
    for i in range(lay.count()):
        item = lay.itemAt(i)
        w = item.widget() if item else None
        if not w:
            continue
        n = w.objectName()
        if n == "qt_calendar_monthbutton":
            month_idx = i
        elif n == "qt_calendar_yearedit":
            year_idx = i

    nav_ss = f"""
        QToolButton {{
            color: {CIAN};
            background: transparent;
            font-weight: 900;
            font-size: 13px;
            border: none;
            padding: 4px 8px;
            min-height: 34px;
        }}
        QToolButton:hover {{
            color: #7CFFE1;
        }}
        QToolButton::menu-indicator {{ image: none; width: 0px; }}
    """

    our_month = getattr(cal, "_sm_month_btn", None)
    if our_month is None:
        our_month = QToolButton(nav)
        our_month.setCursor(Qt.CursorShape.PointingHandCursor)
        our_month.setStyleSheet(nav_ss)
        cal._sm_month_btn = our_month
        if month_idx >= 0:
            lay.insertWidget(month_idx, our_month)
        else:
            lay.addWidget(our_month)
        our_month.clicked.connect(lambda: _show_month_grid(our_month, cal))

    meses = _meses_i18n()

    def _sync_month():
        raw = month_btn_orig.text() if month_btn_orig else ""
        if not raw:
            m = cal.monthShown()
            raw = meses[m] if 1 <= m <= 12 else "?"
        our_month.setText((raw[0].upper() + raw[1:] if raw else "?") + " \u25BE")

    cal.currentPageChanged.connect(lambda y, m: _sync_month())
    _sync_month()
    our_month.show()

    if month_btn_orig:
        month_btn_orig.setMaximumSize(0, 0)
        month_btn_orig.hide()

    year_btn = getattr(cal, "_sm_year_btn", None)
    if year_btn is None:
        year_btn = QToolButton(nav)
        year_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        year_btn.setStyleSheet(nav_ss)
        cal._sm_year_btn = year_btn
        if year_idx >= 0:
            lay.insertWidget(year_idx, year_btn)
        else:
            lay.insertWidget(max(0, lay.count() - 1), year_btn)
        year_btn.clicked.connect(lambda: _show_year_popup(year_btn, cal))

    def _sync_year():
        year_btn.setText(f"{cal.yearShown()} \u25BE")

    cal.currentPageChanged.connect(lambda y, m: _sync_year())
    _sync_year()
    year_btn.show()

    if year_spin:
        year_spin.setMaximumSize(0, 0)
        year_spin.hide()


def _show_month_grid_v2(month_btn, cal):
    from PyQt6.QtWidgets import QFrame, QGridLayout, QPushButton, QVBoxLayout

    meses = [m.lower() for m in _meses_i18n(con_vacio=False)]

    popup = QFrame(
        None,
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    popup.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating, True)
    popup.setStyleSheet(f"background: {BG};")

    inner = QFrame(popup)
    inner.setStyleSheet(
        f"""
        QFrame {{
            background: #161B22;
            border: 2px solid {CIAN};
            border-radius: 12px;
        }}
        QPushButton {{
            background: #161B22;
            color: {TEXTO};
            border: none;
            border-radius: 10px;
            padding: 8px 6px;
            font-family: 'Segoe UI';
            font-size: 11px;
            font-weight: 900;
            min-width: 84px;
            min-height: 34px;
        }}
        QPushButton:hover {{
            background: {CIAN};
            color: {BG};
        }}
        """
    )

    grid = QGridLayout(inner)
    grid.setContentsMargins(8, 8, 8, 8)
    grid.setHorizontalSpacing(6)
    grid.setVerticalSpacing(6)

    for i, mes in enumerate(meses):
        row, col = divmod(i, 4)
        btn = QPushButton(mes)
        if cal.monthShown() == i + 1:
            btn.setStyleSheet(
                f"background: {CIAN}; color: {BG}; border: none; font-weight: 900;"
            )
        month_number = i + 1

        def _on_month(_checked=False, mo=month_number, p=popup):
            cal.setCurrentPage(cal.yearShown(), mo)
            p.close()

        btn.clicked.connect(_on_month)
        grid.addWidget(btn, row, col)

    outer = QVBoxLayout(popup)
    outer.setContentsMargins(4, 4, 4, 4)
    outer.addWidget(inner)
    popup.setFixedSize(inner.sizeHint().width() + 8, inner.sizeHint().height() + 8)
    popup.move(month_btn.mapToGlobal(month_btn.rect().bottomLeft()))
    popup.show()
    popup.raise_()


def _show_year_popup_v2(year_btn, cal):
    from PyQt6.QtCore import QTimer
    from PyQt6.QtGui import QFont
    from PyQt6.QtWidgets import (
        QAbstractItemView,
        QFrame,
        QListWidget,
        QListWidgetItem,
        QVBoxLayout,
    )

    current_year = QDate.currentDate().year()
    years = list(range(current_year, 1924, -1))
    shown_year = cal.yearShown()

    popup = QFrame(
        None,
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    popup.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating, True)
    popup.setStyleSheet(f"background: {BG};")

    inner = QFrame(popup)
    inner.setStyleSheet(
        f"""
        QFrame {{
            background: #161B22;
            border: 2px solid {CIAN};
            border-radius: 12px;
        }}
        QListWidget {{
            background: #161B22;
            border: none;
            outline: none;
            font-family: 'Segoe UI';
            font-size: 12px;
            font-weight: 900;
        }}
        QListWidget::item {{
            color: {TEXTO};
            padding: 7px 10px;
            border-radius: 8px;
            min-height: 24px;
        }}
        QListWidget::item:hover {{
            background: {CIAN};
            color: {BG};
        }}
        QListWidget::item:selected {{
            background: {CIAN};
            color: {BG};
        }}
        """
    )

    list_w = QListWidget(inner)
    list_w.setUniformItemSizes(True)
    list_w.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
    list_w.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
    list_w.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

    bold_font = QFont("Segoe UI", 11, QFont.Weight.Bold)
    norm_font = QFont("Segoe UI", 11)
    scroll_to_idx = 0

    for i, year in enumerate(years):
        item = QListWidgetItem(str(year))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        if year == shown_year:
            item.setForeground(QColor(CIAN))
            item.setFont(bold_font)
            scroll_to_idx = i
        else:
            item.setFont(norm_font)
        list_w.addItem(item)

    list_w.setFixedWidth(112)
    list_w.setFixedHeight(7 * 34)

    inner_ly = QVBoxLayout(inner)
    inner_ly.setContentsMargins(6, 6, 6, 6)
    inner_ly.addWidget(list_w)

    outer_ly = QVBoxLayout(popup)
    outer_ly.setContentsMargins(4, 4, 4, 4)
    outer_ly.addWidget(inner)
    popup.setFixedSize(inner.sizeHint().width() + 8, inner.sizeHint().height() + 8)

    def _on_year(item):
        cal.setCurrentPage(int(item.text()), cal.monthShown())
        popup.close()

    list_w.itemClicked.connect(_on_year)

    popup.move(year_btn.mapToGlobal(year_btn.rect().bottomLeft()))
    popup.show()
    popup.raise_()

    def _center_current():
        list_w.scrollToItem(
            list_w.item(scroll_to_idx),
            QAbstractItemView.ScrollHint.PositionAtCenter,
        )

    QTimer.singleShot(0, _center_current)


def _do_calendar_nav_v2(cal):
    from PyQt6.QtCore import QTimer
    from PyQt6.QtGui import QIcon
    from PyQt6.QtWidgets import QSpinBox, QToolButton

    cal.setGridVisible(False)

    arrow_ss = f"""
        QToolButton {{
            color: {CIAN};
            background: transparent;
            font-weight: 900;
            font-size: 24px;
            border: none;
            padding: 0px 8px 2px 8px;
            min-width: 30px;
            min-height: 34px;
        }}
        QToolButton:hover {{
            color: #7CFFE1;
        }}
    """

    for name in ("qt_calendar_prevmonth", "qt_calendar_nextmonth"):
        btn = cal.findChild(QToolButton, name)
        if btn:
            btn.setIcon(QIcon())
            btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
            btn.setText("\u27F5" if "prev" in name else "\u27F6")
            btn.setStyleSheet(arrow_ss)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)

    month_btn_orig = cal.findChild(QToolButton, "qt_calendar_monthbutton")
    year_spin = cal.findChild(QSpinBox, "qt_calendar_yearedit")
    nav = (
        year_spin.parentWidget()
        if year_spin
        else month_btn_orig.parentWidget()
        if month_btn_orig
        else None
    )
    lay = nav.layout() if nav else None
    if not lay:
        return

    month_idx = year_idx = -1
    for i in range(lay.count()):
        item = lay.itemAt(i)
        w = item.widget() if item else None
        if not w:
            continue
        name = w.objectName()
        if name == "qt_calendar_monthbutton":
            month_idx = i
        elif name == "qt_calendar_yearedit":
            year_idx = i

    nav_ss = f"""
        QToolButton {{
            color: {CIAN};
            background: transparent;
            font-weight: 900;
            font-size: 13px;
            border: none;
            padding: 4px 8px;
            min-height: 34px;
        }}
        QToolButton:hover {{
            color: #7CFFE1;
        }}
        QToolButton::menu-indicator {{ image: none; width: 0px; }}
    """

    our_month = getattr(cal, "_sm_month_btn_v2", None)
    if our_month is None:
        our_month = QToolButton(nav)
        our_month.setCursor(Qt.CursorShape.PointingHandCursor)
        our_month.setStyleSheet(nav_ss)
        cal._sm_month_btn_v2 = our_month
        if month_idx >= 0:
            lay.insertWidget(month_idx, our_month)
        else:
            lay.addWidget(our_month)
        our_month.clicked.connect(
            lambda: QTimer.singleShot(0, lambda: _show_month_grid_v2(our_month, cal))
        )

    months = _meses_i18n()

    def _sync_month():
        raw = month_btn_orig.text() if month_btn_orig else ""
        if not raw:
            month = cal.monthShown()
            raw = months[month] if 1 <= month <= 12 else "?"
        our_month.setText((raw[0].upper() + raw[1:] if raw else "?") + " \u25BE")

    cal.currentPageChanged.connect(lambda y, m: _sync_month())
    _sync_month()
    our_month.show()

    if month_btn_orig:
        month_btn_orig.setMaximumSize(0, 0)
        month_btn_orig.hide()

    year_btn = getattr(cal, "_sm_year_btn_v2", None)
    if year_btn is None:
        year_btn = QToolButton(nav)
        year_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        year_btn.setStyleSheet(nav_ss)
        cal._sm_year_btn_v2 = year_btn
        if year_idx >= 0:
            lay.insertWidget(year_idx, year_btn)
        else:
            lay.insertWidget(max(0, lay.count() - 1), year_btn)
        year_btn.clicked.connect(
            lambda: QTimer.singleShot(0, lambda: _show_year_popup_v2(year_btn, cal))
        )

    def _sync_year():
        year_btn.setText(f"{cal.yearShown()} \u25BE")

    cal.currentPageChanged.connect(lambda y, m: _sync_year())
    _sync_year()
    year_btn.show()

    if year_spin:
        year_spin.setMaximumSize(0, 0)
        year_spin.hide()


_SS_COMBO_VIEW = f"""
QListView {{
    background: #161B22;
    color: {TEXTO};
    border: none;
    outline: none;
    padding: 2px;
}}
QListView::item {{
    background: #161B22;
    color: {TEXTO};
    padding: 5px 10px;
    border-radius: 2px;
}}
QListView::item:hover {{
    background: rgba(0,255,198,0.12);
    color: {CIAN};
}}
QListView::item:selected {{
    background: rgba(0,255,198,0.22);
    color: {CIAN};
}}
QListView QScrollBar:vertical {{
    background: transparent; width: 12px; margin: 2px 0px;
}}
QListView QScrollBar::handle:vertical {{
    background: {CIAN}; min-height: 24px; border-radius: 6px;
}}
QListView QScrollBar::add-line:vertical, QListView QScrollBar::sub-line:vertical {{
    border: none; background: none; width: 0px; height: 0px;
}}
QListView QScrollBar::add-page:vertical, QListView QScrollBar::sub-page:vertical {{
    background: transparent;
}}
"""

def _separador():
    sep = QFrame()
    sep.setFrameShape(QFrame.Shape.HLine)
    sep.setStyleSheet(f"color: {BORDE}; background: {BORDE};")
    sep.setFixedHeight(1)
    return sep


# ─────────────────────────────────────────────────────────────────────────────
# Hilo de importación de histórico de ventas
# ─────────────────────────────────────────────────────────────────────────────

class _ImportarHistoricoHilo(QThread):
    finalizado = pyqtSignal(str)

    def __init__(self, ruta):
        super().__init__()
        self.ruta = ruta

    def run(self):
        try:
            import pandas as pd
            ext = os.path.splitext(self.ruta)[1].lower()
            if ext == ".xlsx":
                df = pd.read_excel(self.ruta)
            elif ext == ".csv":
                df = pd.read_csv(self.ruta)
            else:
                self.finalizado.emit(tr("vta.fmt_unsupported", default="Formato no soportado. Usa Excel (.xlsx) o CSV."))
                return

            df.columns = [c.strip().lower() for c in df.columns]
            col_fecha  = next((c for c in df.columns if "fecha" in c or "date" in c), None)
            col_total  = next((c for c in df.columns if "total" in c or "importe" in c or "factur" in c), None)

            if not col_fecha or not col_total:
                self.finalizado.emit(
                    tr("vta.need_columns", default="El archivo debe tener columnas 'fecha' y 'total' (o variantes similares).")
                )
                return

            insertados = 0
            with obtener_conexion() as conn:
                cur = conn.cursor()
                for _, row in df.iterrows():
                    try:
                        fecha_val  = pd.to_datetime(row[col_fecha]).date()
                        total_val  = float(row[col_total])
                        dia_semana = fecha_val.weekday()
                        cur.execute(
                            "INSERT INTO prevision_historico "
                            "(fecha, total_facturado, fuente, dia_semana) "
                            "VALUES (%s, %s, 'IMPORTADO', %s) "
                            "ON DUPLICATE KEY UPDATE "
                            "total_facturado = VALUES(total_facturado), "
                            "dia_semana = VALUES(dia_semana)",
                            (fecha_val, total_val, dia_semana),
                        )
                        insertados += 1
                    except Exception:
                        continue
                conn.commit()

            self.finalizado.emit(
                tr("vta.hist_import_ok", default="Histórico importado: {n} días cargados desde\n{ruta}", n=insertados, ruta=self.ruta)
            )
        except Exception as e:
            self.finalizado.emit(tr("vta.import_err", default="Error al importar: {e}", e=e))


# ─────────────────────────────────────────────────────────────────────────────
# Ventana principal
# ─────────────────────────────────────────────────────────────────────────────

class VentasAnaliticaWindow(QWidget):
    def __init__(self, callback_vuelta=None, usuario=None, **kwargs):
        super().__init__()
        self.callback_vuelta = callback_vuelta
        self.usuario = usuario
        self._venta_id_sel  = None
        self._sidebar_btns  = []
        self._busqueda_hilo = None

        self.setStyleSheet(_SS_BASE)
        self._setup_ui()
        i18n.conectar_retraduccion(self, self._retraducir)

    def _retraducir(self):
        _tab_def = ["RESUMEN DE VENTAS", "HISTÓRICO DE VENTAS", "RENDIMIENTO"]
        for i, btn in enumerate(self._sidebar_btns):
            _d = _tab_def[i] if i < len(_tab_def) else ""
            btn.setText(tr(self._tab_keys[i], default=_d))
        if hasattr(self, "_btn_exit"):
            self._btn_exit.setText(tr("vta.exit", default="SALIR AL MENÚ"))

    # ── Layout principal ──────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_sidebar())

        self.stack = QStackedWidget()
        self.stack.setStyleSheet(f"background: {BG};")
        # BUSCAR VENTAS se ha migrado a la ventana de BÚSQUEDA/REIMPRESIÓN del TPV.
        self.stack.addWidget(self._panel_resumen_ventas())
        self.stack.addWidget(self._panel_prevision())
        self.stack.addWidget(self._panel_rendimiento())
        root.addWidget(self.stack)

        self._cambiar_tab(0)

    # ── Sidebar ───────────────────────────────────────────────────────────────

    def _build_sidebar(self):
        sidebar = QFrame()
        sidebar.setFixedWidth(280)
        sidebar.setStyleSheet(
            f"QFrame {{ background: {SIDEBAR}; border-right: 1px solid {BORDE}; }}"
        )

        ly = QVBoxLayout(sidebar)
        ly.setContentsMargins(0, 40, 0, 40)
        ly.setSpacing(5)

        titulo = QLabel(tr("ventas.smart_ventas", default="Smart VENTAS"))
        titulo.setStyleSheet(
            "color: #ffffff; font-size: 16px; font-weight: 900; margin-left: 30px; "
            "margin-bottom: 35px; letter-spacing: 2px; border: none; background: transparent;"
        )
        ly.addWidget(titulo)

        self._tab_keys = ["vta.tab_summary", "vta.tab_forecast", "vta.tab_rendimiento"]
        _tab_def = ["RESUMEN DE VENTAS", "HISTÓRICO DE VENTAS", "RENDIMIENTO"]
        for i, texto in enumerate(_tab_def):
            btn = QPushButton(tr(self._tab_keys[i], default=texto))
            btn.setObjectName("btn_sidebar")
            btn.setStyleSheet(_SS_SIDEBAR_BTN)
            btn.setFixedHeight(55)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            idx = i
            btn.clicked.connect(lambda _, x=idx: self._cambiar_tab(x))
            self._sidebar_btns.append(btn)
            ly.addWidget(btn)

        ly.addStretch()

        self._btn_exit = btn_exit = QPushButton(tr("vta.exit", default="SALIR AL MENÚ"))
        btn_exit.setObjectName("btn_sidebar_exit")
        btn_exit.setStyleSheet(_SS_BTN_EXIT)
        btn_exit.setFixedHeight(55)
        btn_exit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_exit.clicked.connect(self._volver)
        ly.addWidget(btn_exit)

        return sidebar

    def _cambiar_tab(self, idx: int):
        self.stack.setCurrentIndex(idx)
        for i, btn in enumerate(self._sidebar_btns):
            btn.setStyleSheet(_SS_SIDEBAR_BTN_ACTIVE if i == idx else _SS_SIDEBAR_BTN)
        # Pestaña RENDIMIENTO (ahora idx 2): refresca al mes actual.
        if idx == 2 and hasattr(self, "tbl_rend"):
            import datetime as _now_dt
            _hoy = _now_dt.date.today()
            if (getattr(self, "_rend_anio", None), getattr(self, "_rend_mes", None)) != (_hoy.year, _hoy.month):
                self._rend_anio, self._rend_mes = _hoy.year, _hoy.month
            self._cargar_rendimiento()

    def _volver(self):
        self.hide()
        if self.callback_vuelta:
            self.callback_vuelta()
        self.close()

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — RESUMEN DE VENTAS
    # ══════════════════════════════════════════════════════════════════════════

    def _panel_resumen_ventas(self):
        page = QWidget()
        page.setStyleSheet(f"background: {BG};")
        root = QVBoxLayout(page)
        root.setContentsMargins(24, 20, 24, 16)
        root.setSpacing(14)

        root.addWidget(_lbl(tr("vta.summary_title", default="RESUMEN DE VENTAS"), bold=True, size=15, color=CIAN))
        root.addWidget(_separador())

        # Filtros
        filtros_frame = QFrame()
        filtros_frame.setObjectName("filtros_res")
        filtros_frame.setStyleSheet(
            f"QFrame#filtros_res {{ background: #0B0F14; border: 1px solid {BORDE}; border-radius: 10px; }}"
        )
        fly = QVBoxLayout(filtros_frame)
        fly.setSpacing(10)
        fly.setContentsMargins(14, 14, 14, 14)

        hoy = QDate.currentDate()
        row_d = QHBoxLayout(); row_d.setSpacing(6)
        self.res_fecha_desde = _date_neon(hoy.addDays(-30))
        self.res_fecha_hasta = _date_neon(hoy)
        for txt, w in ((tr("vta.lbl_date_from", default="Fecha desde"), self.res_fecha_desde), (tr("vta.lbl_date_to", default="Fecha hasta"), self.res_fecha_hasta)):
            lbl = _filter_lbl(txt)
            lbl.setFixedWidth(85)
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            row_d.addWidget(lbl); row_d.addWidget(w)
            row_d.addSpacing(10)
        row_d.addStretch()
        fly.addLayout(row_d)

        row_f = QHBoxLayout(); row_f.setSpacing(6)
        self.res_articulo = _input_neon(tr("vta.ph_article_opt", default="Nombre o código de artículo (opcional)"))
        self.res_seccion  = _input_neon(tr("vta.ph_section_opt", default="Sección (opcional)"))
        for txt, w in ((tr("vta.lbl_article", default="Artículo"), self.res_articulo), (tr("vta.lbl_section", default="Sección"), self.res_seccion)):
            lbl = _filter_lbl(txt)
            lbl.setFixedWidth(65)
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            row_f.addWidget(lbl); row_f.addWidget(w)
            row_f.addSpacing(10)
        fly.addLayout(row_f)

        btn_res = QPushButton(tr("vta.btn_summary", default="GENERAR RESUMEN"))
        btn_res.setStyleSheet(_SS_BTN_CIAN)
        btn_res.setFixedHeight(36)
        btn_res.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_res.clicked.connect(self._generar_resumen)
        fly.addWidget(btn_res, alignment=Qt.AlignmentFlag.AlignRight)
        root.addWidget(filtros_frame)

        # Área de gráfica
        self.chart_frame = QFrame()
        self.chart_frame.setStyleSheet(
            f"QFrame {{ background: #161B22; border: 1px solid {BORDE}; border-radius: 10px; }}"
        )
        self._chart_ly = QVBoxLayout(self.chart_frame)
        self._chart_ly.setContentsMargins(10, 10, 10, 10)
        self._chart_placeholder = _lbl(
            tr("vta.chart_placeholder", default="Selecciona un período y pulsa «Generar Resumen»."),
            size=12, color="#8B949E"
        )
        self._chart_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._chart_ly.addWidget(self._chart_placeholder)
        root.addWidget(self.chart_frame, stretch=1)

        return page

    def _generar_resumen(self):
        desde  = self.res_fecha_desde.date().toString("yyyy-MM-dd")
        hasta  = self.res_fecha_hasta.date().toString("yyyy-MM-dd")
        art    = self.res_articulo.text().strip() or None
        secc   = self.res_seccion.text().strip() or None

        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                if art:
                    cur.execute(
                        "SELECT DATE(v.fecha) AS dia, SUM(vi.subtotal) AS total "
                        "FROM ventas v JOIN venta_items vi ON vi.venta_id = v.id "
                        "WHERE DATE(v.fecha) BETWEEN %s AND %s "
                        "AND (vi.codigo_articulo = %s OR vi.nombre LIKE %s) "
                        "GROUP BY dia ORDER BY dia",
                        (desde, hasta, art, f"%{art}%"),
                    )
                elif secc:
                    cur.execute(
                        "SELECT DATE(v.fecha) AS dia, SUM(vi.subtotal) AS total "
                        "FROM ventas v JOIN venta_items vi ON vi.venta_id = v.id "
                        "WHERE DATE(v.fecha) BETWEEN %s AND %s AND vi.seccion = %s "
                        "GROUP BY dia ORDER BY dia",
                        (desde, hasta, secc),
                    )
                else:
                    cur.execute(
                        "SELECT DATE(v.fecha) AS dia, SUM(v.total) AS total "
                        "FROM ventas v WHERE DATE(v.fecha) BETWEEN %s AND %s "
                        "GROUP BY dia ORDER BY dia",
                        (desde, hasta),
                    )
                datos = cur.fetchall()

                # Top 10
                cur.execute(
                    "SELECT vi.codigo_articulo, vi.nombre, SUM(vi.cantidad) AS uds "
                    "FROM ventas v JOIN venta_items vi ON vi.venta_id = v.id "
                    "WHERE DATE(v.fecha) BETWEEN %s AND %s "
                    + ("AND vi.seccion = %s " if secc else "")
                    + "GROUP BY vi.codigo_articulo, vi.nombre ORDER BY uds DESC LIMIT 10",
                    (desde, hasta, secc) if secc else (desde, hasta),
                )
                top10 = cur.fetchall()

            self._dibujar_grafica(datos, top10, desde, hasta)
        except Exception as e:
            self._chart_placeholder.setText(tr("vta.summary_err", default="Error al generar resumen: {e}", e=e))

    def _dibujar_grafica(self, datos, top10, desde, hasta):
        # Clear old chart
        while self._chart_ly.count():
            item = self._chart_ly.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if not datos:
            lbl = _lbl(tr("vta.no_data_period", default="Sin datos para el período seleccionado."), size=12, color="#8B949E")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._chart_ly.addWidget(lbl)
            return

        if not MATPLOTLIB_OK:
            lbl = _lbl(
                tr("vta.matplotlib_missing", default="Instala matplotlib para ver gráficas:\npip install matplotlib"),
                size=11, color="#8B949E"
            )
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._chart_ly.addWidget(lbl)
            return

        fechas = [str(r[0]) for r in datos]
        totales = [float(r[1] or 0) for r in datos]
        n_dias  = (QDate.fromString(hasta, "yyyy-MM-dd").toJulianDay()
                   - QDate.fromString(desde, "yyyy-MM-dd").toJulianDay()) + 1

        fig = Figure(figsize=(8, 3.5), facecolor="#161B22", tight_layout=True)
        ax  = fig.add_subplot(111)
        ax.set_facecolor("#0D1117")
        ax.tick_params(colors="#8B949E", labelsize=8)
        ax.spines[:].set_color("#21262D")
        ax.xaxis.label.set_color("#8B949E")
        ax.yaxis.label.set_color("#8B949E")

        if n_dias == 1:
            ax.plot(fechas, totales, color=CIAN, linewidth=2, marker="o", markersize=5)
            ax.fill_between(range(len(fechas)), totales, alpha=0.15, color=CIAN)
        else:
            bars = ax.bar(fechas, totales, color=CIAN, alpha=0.85)
            ax.bar_label(bars, fmt=f"%.0f{divisas.simbolo()}", fontsize=7, color="#8B949E")

        ax.set_title(tr("vta.chart_title", default="Facturación  {desde} → {hasta}", desde=desde, hasta=hasta), color=CIAN, fontsize=10, pad=8)
        ax.set_ylabel(divisas.simbolo(), color="#8B949E", fontsize=9)
        step = max(1, len(fechas) // 8)
        ax.set_xticks(range(0, len(fechas), step))
        ax.set_xticklabels(fechas[::step], rotation=30, ha="right", fontsize=7, color="#8B949E")

        canvas = FigureCanvasQTAgg(fig)
        canvas.setStyleSheet("background: #161B22;")
        self._chart_ly.addWidget(canvas)

        if top10:
            lbl_top = _lbl(tr("vta.top10_title", default="TOP 10 MÁS VENDIDOS"), bold=True, size=10, color=CIAN)
            lbl_top.setContentsMargins(0, 6, 0, 2)
            self._chart_ly.addWidget(lbl_top)
            _u = tr("vta.unit_uds_low", default="uds")
            top_txt = "  |  ".join(f"{nom or cod} ({int(uds)} {_u})" for cod, nom, uds in top10)
            self._chart_ly.addWidget(_lbl(top_txt, size=9, color="#8B949E"))

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — PREVISIÓN FACTURACIÓN
    # ══════════════════════════════════════════════════════════════════════════

    def _panel_prevision(self):
        page = QWidget()
        page.setStyleSheet(f"background: {BG};")
        root = QVBoxLayout(page)
        root.setContentsMargins(24, 20, 24, 16)
        root.setSpacing(14)

        _center = Qt.AlignmentFlag.AlignCenter
        root.addWidget(_lbl(tr("vta.forecast_title", default="HISTÓRICO DE VENTAS"), bold=True, size=15, color=CIAN))
        root.addWidget(_separador())

        # ── Recuadro explicativo (arriba) ──────────────────────────────────────
        expl = QFrame(); expl.setObjectName("expl_hist")
        expl.setStyleSheet(f"QFrame#expl_hist {{ background: #161B22; border: 1px solid {BORDE}; border-radius: 10px; }}")
        ely = QVBoxLayout(expl); ely.setContentsMargins(18, 12, 18, 12); ely.setSpacing(4)
        ely.addWidget(_lbl(tr("vta.hist_help_title", default="¿QUÉ ES ESTA PESTAÑA?"), bold=True, size=13, color=CIAN))
        _help = _lbl(tr(
            "vta.hist_help_body",
            default=(
                "Sube tu facturación de años anteriores para alimentar la previsión inteligente. "
                "La tabla de abajo resume el histórico importado, columna por columna:\n"
                "•  Año — ejercicio al que pertenecen los datos.\n"
                "•  Días — número de días con datos registrados ese año.\n"
                "•  Total facturado — suma de la facturación de todo el año.\n"
                "•  Fuente — origen de los datos (archivo importado o ventas reales del TPV).\n\n"
                "Archivos compatibles: .xlsx o .csv con DOS columnas → FECHA (DD/MM/AAAA) "
                "y TOTAL (importe facturado ese día). Cada fila = un día."
            )), bold=True, size=14, color="#C9D1D9")
        _help.setWordWrap(True); ely.addWidget(_help)
        root.addWidget(expl)

        root.addStretch()

        # Icono + subtítulo + botón (centrados, más abajo)
        lbl_icon = QLabel("📈")
        lbl_icon.setAlignment(_center)
        lbl_icon.setStyleSheet("font-size: 60px; background: transparent; color: white;")
        root.addWidget(lbl_icon)

        lbl_sub = _lbl(tr("vta.forecast_subtitle", default="Sistema de previsión inteligente basado en histórico de ventas y modelos IA"),
                       size=11, color="#8B949E")
        lbl_sub.setAlignment(_center)
        lbl_sub.setWordWrap(True)
        root.addWidget(lbl_sub)

        root.addSpacing(16)

        btn_row = QHBoxLayout(); btn_row.addStretch()
        self.btn_subir_hist = QPushButton(tr("vta.btn_upload_hist", default="SUBIR VENTAS PASADAS"))
        self.btn_subir_hist.setStyleSheet(_SS_BTN_CIAN)
        self.btn_subir_hist.setFixedHeight(48); self.btn_subir_hist.setFixedWidth(260)
        self.btn_subir_hist.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_subir_hist.clicked.connect(self._subir_ventas_pasadas)
        btn_row.addWidget(self.btn_subir_hist); btn_row.addStretch()
        root.addLayout(btn_row)

        # Estado importación
        self.lbl_prev_estado = _lbl("", size=10, color="#8B949E")
        self.lbl_prev_estado.setAlignment(_center)
        root.addWidget(self.lbl_prev_estado)

        root.addStretch()

        # Info tabla de objetivos
        info_frame = QFrame()
        info_frame.setObjectName("info_historico")
        info_frame.setStyleSheet(
            f"QFrame#info_historico {{ background: #161B22; border: 1px solid {BORDE}; border-radius: 10px; }}"
        )
        ifly = QVBoxLayout(info_frame)
        ifly.setContentsMargins(16, 12, 16, 12)
        ifly.setSpacing(6)
        ifly.addWidget(_lbl(tr("vta.hist_imported", default="HISTÓRICO IMPORTADO"), bold=True, size=13, color=CIAN))
        self.tbl_historico = QTableWidget(0, 4)
        self.tbl_historico.setHorizontalHeaderLabels([
            tr("vta.col_year", default="Año"),
            tr("vta.col_days", default="Días"),
            tr("vta.col_total_billed", default="Total facturado"),
            tr("vta.col_source", default="Fuente"),
        ])
        self.tbl_historico.setStyleSheet(_SS_TABLE)
        self.tbl_historico.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl_historico.setMaximumHeight(180)
        self.tbl_historico.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl_historico.verticalHeader().setVisible(False)
        _ClipTableTopCorners(self.tbl_historico)
        ifly.addWidget(self.tbl_historico)
        root.addWidget(info_frame)

        self._cargar_resumen_historico()
        return page

    def _cargar_resumen_historico(self):
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT YEAR(fecha) AS anio, COUNT(*) AS dias, "
                    "SUM(total_facturado) AS total, fuente "
                    "FROM prevision_historico "
                    "GROUP BY anio, fuente ORDER BY anio DESC, fuente"
                )
                filas = cur.fetchall()
            self.tbl_historico.setRowCount(len(filas))
            for r, (anio, dias, total, fuente) in enumerate(filas):
                for c, val in enumerate([str(anio), str(dias), f"{divisas.formatear(float(total or 0))}", fuente]):
                    item = QTableWidgetItem(val)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.tbl_historico.setItem(r, c, item)
        except Exception:
            pass

    def _subir_ventas_pasadas(self):
        ruta, _ = QFileDialog.getOpenFileName(
            self, tr("vta.file_dialog_hist", default="Seleccionar archivo de ventas históricas"),
            "", "Archivos (*.xlsx *.csv)"
        )
        if not ruta:
            return

        prog = QProgressDialog(tr("vta.importing_hist", default="Importando histórico…"), None, 0, 0, self)
        prog.setWindowModality(Qt.WindowModality.ApplicationModal)
        prog.setWindowTitle(tr("vta.import_title", default="Importación"))
        prog.setCancelButton(None)
        prog.show()

        self._import_hilo = _ImportarHistoricoHilo(ruta)

        def _fin(msg):
            prog.close()
            self.lbl_prev_estado.setText(msg)
            self._cargar_resumen_historico()

        self._import_hilo.finalizado.connect(_fin)
        self._import_hilo.start()

    def _ver_prevision(self):
        try:
            anio_actual = datetime.now().year
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT COUNT(*) FROM prevision_historico"
                )
                cnt = cur.fetchone()[0]

            if cnt == 0:
                mostrar_mensaje(
                    self, tr("vta.no_data_title", default="Sin datos"),
                    tr("vta.forecast_need_data",
                       default="Primero sube al menos un archivo de ventas pasadas\npara que el sistema pueda calcular previsiones.")
                )
                return

            self._generar_prevision_excel(anio_actual)
        except Exception as e:
            mostrar_mensaje(self, tr("vta.error_title", default="Error"), str(e), "error")

    def _generar_prevision_excel(self, anio: int):
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT fecha, total_facturado FROM prevision_historico ORDER BY fecha"
                )
                hist = cur.fetchall()

            if not hist:
                return

            df = pd.DataFrame(hist, columns=["fecha", "total"])
            df["fecha"] = pd.to_datetime(df["fecha"])
            total_anual = df["total"].sum()

            if total_anual == 0:
                mostrar_mensaje(self, tr("vta.no_data_title", default="Aviso"), tr("vta.not_enough_data", default="El histórico no tiene datos suficientes."), "warning")
                return

            df["peso"] = df["total"] / total_anual

            # Generate forecast for target year using weights
            fechas_anio = pd.date_range(f"{anio}-01-01", f"{anio}-12-31")
            objetivo = total_anual  # default: same as historical total
            try:
                with obtener_conexion() as conn:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT objetivo_anual FROM prevision_objetivos WHERE anio = %s", (anio,)
                    )
                    row = cur.fetchone()
                    if row and row[0]:
                        objetivo = float(row[0])
            except Exception:
                pass

            peso_medio = df["peso"].mean() if len(df) else 1 / 365

            wb = Workbook()
            ws = wb.active
            ws.title = tr("vta.report_sheet", default="Previsión {anio}", anio=anio)

            amarillo = PatternFill("solid", fgColor="FFC000")
            turquesa = PatternFill("solid", fgColor="00BCD4")
            hdr_font = Font(name="Segoe UI", bold=True, size=10)
            cab_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
            cen = Alignment(horizontal="center", vertical="center")

            cabeceras = [
                tr("vta.col_day", default="Día"),
                tr("vta.col_prev_year", default="Año anterior"),
                tr("vta.col_year_n", default="Año {anio}", anio=anio),
                tr("vta.col_forecast", default="Previsión ventas"),
                tr("vta.col_diff", default="Diferencia"),
            ]
            for c, h in enumerate(cabeceras, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = amarillo
                cell.font = hdr_font
                cell.alignment = cen

            fila = 2
            for fecha in fechas_anio:
                mes = fecha.month
                dia = fecha.day
                hist_match = df[df["fecha"].dt.month == mes][df["fecha"].dt.day == dia]
                anio_ant  = round(hist_match["total"].mean(), 2) if not hist_match.empty else 0
                prevision = round(objetivo * peso_medio, 2)

                ws.cell(row=fila, column=1, value=fecha.strftime("%d/%m/%Y")).alignment = cen
                ws.cell(row=fila, column=2, value=anio_ant).alignment = cen
                ws.cell(row=fila, column=3, value="").alignment = cen  # introduce manual
                ws.cell(row=fila, column=4, value=prevision).alignment = cen
                f = ws.cell(row=fila, column=5, value=f"=C{fila}-D{fila}")
                f.alignment = cen

                if dia == 1 or fecha == fechas_anio[-1]:
                    mes_lbl = fecha.strftime("%B %Y").upper()
                    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
                    cell = ws.cell(row=fila, column=1, value=mes_lbl)
                    cell.fill = turquesa
                    cell.font = cab_font
                    cell.alignment = cen
                    fila += 1
                    ws.cell(row=fila, column=1, value=fecha.strftime("%d/%m/%Y")).alignment = cen
                    ws.cell(row=fila, column=2, value=anio_ant).alignment = cen
                    ws.cell(row=fila, column=3, value="").alignment = cen
                    ws.cell(row=fila, column=4, value=prevision).alignment = cen
                    ws.cell(row=fila, column=5, value=f"=C{fila}-D{fila}").alignment = cen

                fila += 1

            carpeta = os.path.abspath(
                os.path.join(os.path.dirname(__file__), "../../documentos/informes")
            )
            os.makedirs(carpeta, exist_ok=True)
            ruta = os.path.join(carpeta, f"Prevision_{anio}.xlsx")
            wb.save(ruta)

            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "INSERT INTO prevision_objetivos (anio, objetivo_anual, excel_generado, ruta_excel_drive) "
                    "VALUES (%s, %s, 1, %s) ON DUPLICATE KEY UPDATE excel_generado=1, ruta_excel_drive=%s",
                    (anio, objetivo, ruta, ruta),
                )
                conn.commit()

            mostrar_mensaje(self, tr("vta.forecast_generated_title", default="Previsión generada"), tr("vta.excel_saved", default="Excel guardado en:\n{ruta}", ruta=ruta))
            try:
                import platform
                import subprocess
                if platform.system() == "Windows":
                    os.startfile(ruta)
                else:
                    subprocess.Popen(["xdg-open", ruta])
            except Exception:
                pass

        except Exception as e:
            mostrar_mensaje(self, tr("vta.error_title", default="Error"), tr("vta.excel_gen_err", default="No se pudo generar el Excel:\n{e}", e=e), "error")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 4 — RENDIMIENTO (facturación diaria + productividad por tienda)
    # ══════════════════════════════════════════════════════════════════════════
    _MESES_RND = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                  "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    def _panel_rendimiento(self):
        page = QWidget(); page.setStyleSheet(f"background: {BG};")
        root = QVBoxLayout(page); root.setContentsMargins(24, 20, 24, 16); root.setSpacing(12)
        _now = datetime.now()
        self._rend_anio, self._rend_mes = _now.year, _now.month
        self._rend_updating = False

        root.addWidget(_lbl(tr("vta.perf_title", default="RENDIMIENTO"), bold=True, size=15, color=CIAN))
        root.addWidget(_separador())
        self._rend_mes_lbl = _lbl(f"{self._MESES_RND[self._rend_mes]} {self._rend_anio}", bold=True, size=14, color="#C9D1D9")
        root.addWidget(self._rend_mes_lbl)

        # Tabla única: rendimiento + previsión IA (última columna, a la derecha).
        self._rend_cols = [
            tr("vta.perf_c_dia", default="Día"),
            tr("vta.perf_c_fact", default="Fact. día"),
            tr("vta.perf_c_factac", default="Fact. acum."),
            tr("vta.perf_c_cli", default="Nº clientes"),
            tr("vta.perf_c_tm", default="Ticket medio"),
            tr("vta.perf_c_horas", default="Horas día"),
            tr("vta.perf_c_horasac", default="Horas acum."),
            tr("vta.perf_c_prod", default="Prod. día"),
            tr("vta.perf_c_prodac", default="Prod. acum."),
            tr("vta.perf_c_prev", default="Prev. Fact."),
        ]
        t = QTableWidget(0, len(self._rend_cols))
        t.setHorizontalHeaderLabels(self._rend_cols)
        t.setStyleSheet(_SS_TABLE)
        t.verticalHeader().setVisible(False)
        # Celdas editables (facturación / nº clientes / horas / previsión).
        t.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.SelectedClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        _hh = t.horizontalHeader()
        _hh.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)   # resto equitativo
        _hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # "Día" a la mitad
        t.setColumnWidth(0, 55)
        _RoundTableCorners(t)   # redondea las 4 esquinas (contorno neón continuo)
        self.tbl_rend = t
        root.addLayout(self._rend_acciones_row())   # imprimir / compartir
        root.addWidget(t, 1)

        br = QHBoxLayout(); br.addStretch()
        self.btn_rend_guardar = QPushButton(tr("vta.perf_save", default="GUARDAR CAMBIOS"))
        self.btn_rend_guardar.setStyleSheet(_SS_BTN_VERDE)
        self.btn_rend_guardar.setFixedHeight(44); self.btn_rend_guardar.setFixedWidth(220)
        self.btn_rend_guardar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_rend_guardar.clicked.connect(self._guardar_rendimiento)
        br.addWidget(self.btn_rend_guardar); root.addLayout(br)

        self._cargar_rendimiento()
        t.cellChanged.connect(self._rend_on_edit)
        return page

    # ── IA predictiva de facturación (alimentada por HISTÓRICO DE VENTAS) ──────
    def _serie_historica(self):
        """Serie diaria [(fecha, importe)] del histórico subido (prevision_historico)."""
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT fecha, SUM(total_facturado) FROM prevision_historico "
                    "GROUP BY fecha ORDER BY fecha")
                filas = cur.fetchall()
        except Exception:
            filas = []
        serie = []
        for f, t in filas:
            if f is None:
                continue
            fecha = f if hasattr(f, "weekday") else self._parse_fecha(f)
            if fecha is not None:
                serie.append((fecha, float(t or 0)))
        return serie

    def _serie_ventas_tpv(self):
        """Serie diaria [(fecha, importe)] de las ventas reales del TPV/autocobro."""
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT DATE(fecha), COALESCE(SUM(total),0) FROM ventas "
                    "GROUP BY DATE(fecha) ORDER BY DATE(fecha)")
                filas = cur.fetchall()
        except Exception:
            filas = []
        serie = []
        for f, t in filas:
            if f is None:
                continue
            fecha = f if hasattr(f, "weekday") else self._parse_fecha(f)
            if fecha is not None:
                serie.append((fecha, float(t or 0)))
        return serie

    def _prever_facturacion_mensual(self, anio, mes):
        """Previsión IA de facturación diaria para (anio, mes).

        Se alimenta del histórico de ventas pasadas subido en la pestaña
        HISTÓRICO DE VENTAS (tabla ``prevision_historico``) y, si no hay ningún
        archivo subido, recurre automáticamente a las ventas reales acumuladas
        por el TPV/autocobro (tabla ``ventas``). Tiene en cuenta eventos
        comerciales y festividades (Navidad, Black Friday, Halloween, Reyes,
        Rebajas, Semana Santa, etc.) además del día de la semana y la tendencia.
        Intenta primero un modelo de series temporales (Prophet, con las
        festividades como *holidays*) y, si no está disponible o hay pocos
        datos, usa una heurística estacional. Devuelve ``{dia: importe}``.
        """
        import calendar
        ndias = calendar.monthrange(anio, mes)[1]
        prevision = {d: 0.0 for d in range(1, ndias + 1)}
        serie = self._serie_historica()
        if not serie:
            # Respaldo: ventas reales acumuladas por el TPV/autocobro.
            serie = self._serie_ventas_tpv()
        if not serie:
            return prevision
        # 1) Modelo de series temporales (IA) si está disponible y hay datos.
        try:
            pred = self._prever_prophet(serie, anio, mes, ndias)
            if pred:
                return pred
        except Exception:
            pass
        # 2) Heurística estacional como respaldo.
        return self._prever_heuristico(serie, anio, mes, ndias)

    @staticmethod
    def _parse_fecha(valor):
        import datetime as _dt
        if isinstance(valor, _dt.datetime):
            return valor.date()
        if isinstance(valor, _dt.date):
            return valor
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
            try:
                return _dt.datetime.strptime(str(valor)[:19], fmt).date()
            except ValueError:
                continue
        return None

    # ── Calendario de eventos comerciales / estacionales (fuente única) ───────
    @staticmethod
    def _nth_weekday(anio, mes, weekday, n):
        """n-ésimo ``weekday`` (0=lunes) del mes. Ej.: 4º viernes de noviembre."""
        import datetime as _dt
        d = _dt.date(anio, mes, 1)
        offset = (weekday - d.weekday()) % 7
        return d + _dt.timedelta(days=offset + 7 * (n - 1))

    @staticmethod
    def _domingo_pascua(anio):
        """Domingo de Pascua (algoritmo de Gauss/Anonymous Gregorian)."""
        import datetime as _dt
        a = anio % 19
        b = anio // 100
        c = anio % 100
        d = b // 4
        e = b % 4
        f = (b + 8) // 25
        g = (b - f + 1) // 3
        h = (19 * a + b - d - g + 15) % 30
        i = c // 4
        k = c % 4
        ll = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * ll) // 451
        mes = (h + ll - 7 * m + 114) // 31
        dia = ((h + ll - 7 * m + 114) % 31) + 1
        return _dt.date(anio, mes, dia)

    def _fechas_eventos(self, anio):
        """Eventos comerciales del año: (nombre, fecha, lower_window,
        upper_window, factor_heurístico). Las ventanas son días relativos a la
        fecha; el factor es el multiplicador de facturación usado por la
        heurística (Prophet sólo necesita las fechas y aprende el efecto)."""
        import datetime as _dt
        D = _dt.date
        black_friday = self._nth_weekday(anio, 11, 4, 4)        # 4º viernes nov.
        cyber_monday = black_friday + _dt.timedelta(days=3)
        dia_madre = self._nth_weekday(anio, 5, 6, 1)            # 1er domingo mayo (ES)
        pascua = self._domingo_pascua(anio)
        jueves_santo = pascua - _dt.timedelta(days=3)
        return [
            ("Reyes",            D(anio, 1, 6),   -2, 0, 1.5),
            ("Rebajas invierno", D(anio, 1, 7),    0, 21, 1.25),
            ("San Valentín",     D(anio, 2, 14),  -3, 0, 1.2),
            ("Día del Padre",    D(anio, 3, 19),  -3, 0, 1.2),
            ("Semana Santa",     jueves_santo,     0, 3, 1.15),
            ("Día de la Madre",  dia_madre,       -3, 0, 1.3),
            ("Rebajas verano",   D(anio, 7, 1),    0, 30, 1.2),
            ("Vuelta al cole",   D(anio, 9, 1),    0, 14, 1.2),
            ("Halloween",        D(anio, 10, 31), -6, 0, 1.3),
            ("Black Friday",     black_friday,    -1, 1, 2.2),
            ("Cyber Monday",     cyber_monday,     0, 0, 1.8),
            ("Campaña Navidad",  D(anio, 12, 24), -9, 0, 1.6),
            ("Navidad",          D(anio, 12, 25),  0, 0, 0.3),
        ]

    @staticmethod
    def _festivos_nacionales(anio):
        """Festivos nacionales de España de baja/nula actividad comercial."""
        import datetime as _dt
        return [
            _dt.date(anio, 1, 1), _dt.date(anio, 1, 6), _dt.date(anio, 5, 1),
            _dt.date(anio, 8, 15), _dt.date(anio, 10, 12), _dt.date(anio, 11, 1),
            _dt.date(anio, 12, 6), _dt.date(anio, 12, 8), _dt.date(anio, 12, 25),
        ]

    def _factores_estacionales(self, anio):
        """{fecha: factor} multiplicativo por evento comercial/festivo."""
        import datetime as _dt
        fac = {}
        for _nombre, fecha, lw, uw, factor in self._fechas_eventos(anio):
            for off in range(lw, uw + 1):
                dia = fecha + _dt.timedelta(days=off)
                if dia.year != anio:
                    continue
                cur = fac.get(dia)
                if cur is None:
                    fac[dia] = factor
                elif factor < 1 or cur < 1:   # los días de cierre/baja mandan
                    fac[dia] = min(cur, factor)
                else:                          # si no, el mayor impulso
                    fac[dia] = max(cur, factor)
        for f in self._festivos_nacionales(anio):
            fac[f] = min(fac.get(f, 1.0), 0.3)
        return fac

    def _eventos_comerciales_prophet(self, anios):
        """DataFrame de holidays para Prophet con todos los eventos de cada año."""
        import pandas as pd
        filas = []
        for a in anios:
            for nombre, fecha, lw, uw, _factor in self._fechas_eventos(a):
                filas.append({"holiday": nombre, "ds": pd.Timestamp(fecha),
                              "lower_window": int(lw), "upper_window": int(uw)})
        return pd.DataFrame(filas) if filas else None

    def _prever_prophet(self, serie, anio, mes, ndias):
        """Previsión con Prophet (opcional). Devuelve {dia: importe} o None."""
        try:
            from prophet import Prophet
        except Exception:
            return None
        if len(serie) < 30:  # datos insuficientes para un modelo fiable
            return None
        import logging
        import pandas as pd
        # Silenciar el logging ruidoso de cmdstanpy/Prophet.
        for nm in ("prophet", "cmdstanpy"):
            logging.getLogger(nm).setLevel(logging.CRITICAL)
        df = pd.DataFrame(serie, columns=["ds", "y"])
        df["ds"] = pd.to_datetime(df["ds"])
        # Eventos comerciales/estacionales como holidays (Prophet aprende su efecto).
        anios = sorted({d.year for d, _ in serie} | {anio})
        holi = self._eventos_comerciales_prophet(anios)
        m = Prophet(weekly_seasonality=True, yearly_seasonality=True,
                    daily_seasonality=False,
                    holidays=(holi if holi is not None and not holi.empty else None))
        try:
            m.add_country_holidays(country_name="ES")  # festivos oficiales de España
        except Exception:
            pass
        m.fit(df)
        fut = pd.DataFrame({"ds": pd.date_range(f"{anio}-{mes:02d}-01", periods=ndias)})
        fc = m.predict(fut)
        out = {}
        for _, r in fc.iterrows():
            out[r["ds"].day] = round(max(0.0, float(r["yhat"])), 2)
        return out

    def _prever_heuristico(self, serie, anio, mes, ndias):
        """Previsión estacional: combina la media del mismo día del mes con el
        factor del día de la semana, ponderada por tendencia reciente."""
        import datetime as _dt
        from collections import defaultdict
        prevision = {d: 0.0 for d in range(1, ndias + 1)}
        total = sum(v for _, v in serie)
        n = len(serie)
        if n == 0 or total <= 0:
            return prevision
        media_dia = total / n
        dow_sum = defaultdict(float); dow_cnt = defaultdict(int)
        dom_sum = defaultdict(float); dom_cnt = defaultdict(int)
        for fecha, v in serie:
            dow_sum[fecha.weekday()] += v; dow_cnt[fecha.weekday()] += 1
            dom_sum[fecha.day] += v; dom_cnt[fecha.day] += 1
        # Tendencia: ratio media de la última mitad vs primera mitad de la serie.
        mitad = n // 2
        tendencia = 1.0
        if mitad >= 1:
            prim = sum(v for _, v in serie[:mitad]) / mitad
            ult = sum(v for _, v in serie[mitad:]) / (n - mitad)
            if prim > 0:
                tendencia = max(0.5, min(2.0, ult / prim))
        factores = self._factores_estacionales(anio)   # Navidad, Black Friday, etc.
        for d in range(1, ndias + 1):
            fecha = _dt.date(anio, mes, d)
            base = (dom_sum[d] / dom_cnt[d]) if dom_cnt[d] else media_dia
            wd = fecha.weekday()
            factor_dow = (dow_sum[wd] / dow_cnt[wd]) / media_dia if dow_cnt[wd] and media_dia else 1.0
            estacional = factores.get(fecha, 1.0)
            pred = (base * 0.5 + media_dia * factor_dow * 0.5) * tendencia * estacional
            prevision[d] = round(max(0.0, pred), 2)
        return prevision

    def _rend_datos_auto(self, anio, mes):
        """Datos auto por día (TPV/autocobro + fichajes), con override manual guardado."""
        import calendar
        ndias = calendar.monthrange(anio, mes)[1]
        data = {d: {"fact": 0.0, "clientes": 0, "horas": 0.0, "prev": 0.0} for d in range(1, ndias + 1)}
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT DAY(fecha), COALESCE(SUM(total),0), COUNT(*) FROM ventas "
                    "WHERE YEAR(fecha)=%s AND MONTH(fecha)=%s GROUP BY DAY(fecha)", (anio, mes))
                for d, tot, cnt in cur.fetchall():
                    if d in data:
                        data[d]["fact"] = float(tot or 0); data[d]["clientes"] = int(cnt or 0)
                cur.execute(
                    "SELECT DAY(entrada), COALESCE(SUM(TIMESTAMPDIFF(MINUTE, entrada, salida)),0) "
                    "FROM fichajes WHERE salida IS NOT NULL AND YEAR(entrada)=%s AND MONTH(entrada)=%s "
                    "GROUP BY DAY(entrada)", (anio, mes))
                for d, mins in cur.fetchall():
                    if d in data:
                        data[d]["horas"] = round(float(mins or 0) / 60.0, 2)
                # Previsión IA: modelo predictivo alimentado por el histórico de
                # ventas subido en la pestaña HISTÓRICO DE VENTAS (prevision_historico).
                try:
                    pred = self._prever_facturacion_mensual(anio, mes)
                    for d, v in pred.items():
                        if d in data:
                            data[d]["prev"] = v
                except Exception:
                    pass
                try:
                    from src.db.empresa import empresa_actual_id
                    cur.execute(
                        "SELECT DAY(fecha), facturacion, clientes, horas, prevision FROM rendimiento_diario "
                        "WHERE id_empresa=%s AND YEAR(fecha)=%s AND MONTH(fecha)=%s",
                        (empresa_actual_id(), anio, mes))
                    for d, f, c, h, p in cur.fetchall():
                        if d in data:
                            if f is not None: data[d]["fact"] = float(f)
                            if c is not None: data[d]["clientes"] = int(c)
                            if h is not None: data[d]["horas"] = float(h)
                            if p is not None: data[d]["prev"] = float(p)
                except Exception:
                    pass
        except Exception:
            pass
        return data

    def _cargar_rendimiento(self):
        self._rend_updating = True
        try:
            if hasattr(self, "_rend_mes_lbl"):
                self._rend_mes_lbl.setText(f"{self._MESES_RND[self._rend_mes]} {self._rend_anio}")
            data = self._rend_datos_auto(self._rend_anio, self._rend_mes)
            dias = sorted(data.keys())
            t = self.tbl_rend
            t.setRowCount(len(dias))
            ncols = len(self._rend_cols)
            editables = {1, 3, 5, 9}  # facturación, clientes, horas, previsión
            for r, d in enumerate(dias):
                base = ["" for _ in range(ncols)]
                base[0] = str(d)
                base[1] = f"{data[d]['fact']:.2f}"
                base[3] = str(int(data[d]['clientes']))
                base[5] = f"{data[d]['horas']:.2f}"
                base[9] = f"{data[d]['prev']:.2f}"   # previsión IA (editable)
                for c in range(ncols):
                    it = QTableWidgetItem(base[c])
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    if c not in editables:
                        it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    t.setItem(r, c, it)
            self._rend_recalcular()
        finally:
            self._rend_updating = False

    def _rend_num(self, r, c):
        it = self.tbl_rend.item(r, c)
        s = (it.text() if it else "").replace(",", ".").strip()
        try:
            return float(s) if s else 0.0
        except ValueError:
            return 0.0

    def _rend_recalcular(self):
        t = self.tbl_rend
        fact_ac = 0.0; horas_ac = 0.0
        prev = self._rend_updating; self._rend_updating = True
        try:
            for r in range(t.rowCount()):
                fact = self._rend_num(r, 1); cli = self._rend_num(r, 3); horas = self._rend_num(r, 5)
                fact_ac += fact; horas_ac += horas
                tm = fact / cli if cli else 0.0
                prod = fact / horas if horas else 0.0
                prodac = fact_ac / horas_ac if horas_ac else 0.0

                def _setc(c, v):
                    it = t.item(r, c)
                    if it is None:
                        it = QTableWidgetItem(); it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable); t.setItem(r, c, it)
                    it.setText(f"{v:.2f}")
                _setc(2, fact_ac); _setc(4, tm); _setc(6, horas_ac); _setc(7, prod); _setc(8, prodac)
        finally:
            self._rend_updating = prev

    def _rend_on_edit(self, row, col):
        if getattr(self, "_rend_updating", False):
            return
        if col in (1, 3, 5):
            self._rend_recalcular()

    def _guardar_rendimiento(self):
        import datetime as _d
        try:
            from src.db.empresa import empresa_actual_id
            eid = empresa_actual_id()
        except Exception:
            eid = None
        t = self.tbl_rend
        try:
            with obtener_conexion() as conn:
                cur = conn.cursor()
                for r in range(t.rowCount()):
                    it0 = t.item(r, 0)
                    dia = (it0.text().strip() if it0 else "")
                    if not dia.isdigit():
                        continue
                    fecha = _d.date(self._rend_anio, self._rend_mes, int(dia))
                    fact = self._rend_num(r, 1); cli = int(self._rend_num(r, 3)); horas = self._rend_num(r, 5)
                    prev = self._rend_num(r, 9)   # previsión IA (última columna)
                    cur.execute(
                        "INSERT INTO rendimiento_diario (id_empresa, fecha, facturacion, clientes, horas, prevision) "
                        "VALUES (%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE "
                        "facturacion=VALUES(facturacion), clientes=VALUES(clientes), horas=VALUES(horas), "
                        "prevision=VALUES(prevision)",
                        (eid, fecha, fact, cli, horas, prev))
                conn.commit()
            mostrar_mensaje(self, tr("vta.perf_saved_t", default="Guardado"),
                            tr("vta.perf_saved", default="Rendimiento guardado correctamente."), "success")
        except Exception as e:
            mostrar_mensaje(self, tr("vta.error_title", default="Error"), str(e), "error")

    # ── Imprimir / Compartir tablas de RENDIMIENTO ────────────────────────────
    def _rend_acciones_row(self):
        row = QHBoxLayout(); row.addStretch()
        b_imp = QPushButton(tr("vta.perf_print", default="🖨  IMPRIMIR"))
        b_sh = QPushButton(tr("vta.perf_share", default="📧  COMPARTIR"))
        for b in (b_imp, b_sh):
            b.setStyleSheet(_SS_BTN_CIAN); b.setFixedHeight(34)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
        b_imp.clicked.connect(self._imprimir_tabla)
        b_sh.clicked.connect(self._compartir_tabla)
        row.addWidget(b_imp); row.addWidget(b_sh)
        return row

    def _exportar_tabla_pdf(self):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        t = self.tbl_rend
        nombre = "Rendimiento"
        titulo = f"RENDIMIENTO — {self._MESES_RND[self._rend_mes]} {self._rend_anio}"
        headers = [t.horizontalHeaderItem(c).text() for c in range(t.columnCount())]
        rows = [headers]
        for r in range(t.rowCount()):
            rows.append([(t.item(r, c).text() if t.item(r, c) else "") for c in range(t.columnCount())])
        carpeta = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../documentos/informes"))
        os.makedirs(carpeta, exist_ok=True)
        ruta = os.path.join(carpeta, f"{nombre}_{self._rend_anio}_{self._rend_mes:02d}.pdf")
        doc = SimpleDocTemplate(ruta, pagesize=landscape(A4),
                                leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1.2 * cm, bottomMargin=1 * cm)
        styles = getSampleStyleSheet()
        tbl = Table(rows, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3C88")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        doc.build([Paragraph(f"<b>{titulo}</b>", styles["Title"]), Spacer(1, 8), tbl])
        return ruta

    def _imprimir_tabla(self):
        try:
            ruta = self._exportar_tabla_pdf()
            mostrar_mensaje(self, tr("vta.perf_print_ok_t", default="Tabla exportada"),
                            tr("vta.perf_print_ok", default="Guardada en:\n{ruta}", ruta=ruta), "success")
            import platform
            import subprocess
            if platform.system() == "Windows":
                os.startfile(ruta)  # abre el visor → permite imprimir
            else:
                subprocess.Popen(["xdg-open", ruta])
        except Exception as e:
            mostrar_mensaje(self, tr("vta.error_title", default="Error"), str(e), "error")

    def _compartir_tabla(self):
        try:
            ruta = self._exportar_tabla_pdf()
            from src.gui.correo_corporativo import enviar_documento_por_correo
            asunto = f"Rendimiento {self._MESES_RND[self._rend_mes]} {self._rend_anio}"
            enviar_documento_por_correo(self, ruta, asunto)
        except Exception as e:
            mostrar_mensaje(self, tr("vta.error_title", default="Error"), str(e), "error")
